import os
import sys
import pandas as pd
import ast
from qgis.PyQt.QtCore import Qt,QUrl
from qgis.PyQt.QtWidgets import QApplication, QWidget, QPushButton, QFileDialog, QVBoxLayout, QLabel, QMessageBox, QInputDialog

from qgis.PyQt.QtGui import QDesktopServices
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.drawing.image import Image  # Import per gestire le immagini

class ArchaeologicalDataMapper(QWidget):
    def __init__(self, iface, parent=None):
        super(ArchaeologicalDataMapper, self).__init__(parent)
        self.iface = iface
        self.initUI()
        # Imposta la finestra del video player per stare sempre sopra
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)

    def initUI(self):
        layout = QVBoxLayout()

        self.label = QLabel('Seleziona il file di input e il file di output')
        layout.addWidget(self.label)

        self.btn_input = QPushButton('Seleziona file excel inventario materiali di pyArchInit', self)
        self.btn_input.clicked.connect(self.get_input_file)
        layout.addWidget(self.btn_input)

        self.btn_output = QPushButton('Salva il file di output', self)
        self.btn_output.clicked.connect(self.get_output_file)
        layout.addWidget(self.btn_output)

        self.btn_process = QPushButton('Elabora il mapping', self)
        self.btn_process.clicked.connect(self.process_data)
        layout.addWidget(self.btn_process)

        self.output_label = QLabel('')
        self.output_label.setOpenExternalLinks(True)
        self.output_label.linkActivated.connect(self.open_file)
        layout.addWidget(self.output_label)

        self.setLayout(layout)
        self.setGeometry(300, 300, 300, 150)
        self.setWindowTitle('Mapper Materiali da pyArchInit a ICA')
        self.show()


    def open_file(self, link):
        QDesktopServices.openUrl(QUrl(link))
    def get_input_file(self):
        self.input_file, _ = QFileDialog.getOpenFileName(self, 'Seleziona file excel dei materiali esportato da pyArchInit', '', 'Excel files (*.xlsx *.xls)')
        if self.input_file:
            self.label.setText(f'File di input: {self.input_file}')

    def get_output_file(self):
        self.output_file, _ = QFileDialog.getSaveFileName(self, 'Seleziona il nome del file di output', '', 'Excel files (*.xlsx *.xls)')
        if self.output_file:
            self.label.setText(f'File di output: {self.output_file}')
            #self.label.setOpenExternalLinks(True)

    def process_data(self):
        if hasattr(self, 'input_file') and hasattr(self, 'output_file'):
            try:
                self.create_archaeological_excel(self.input_file, self.output_file)
                QMessageBox.information(self, 'Successo', f'File Excel creato con successo: {self.output_file}')
                self.output_label.setText(f'<a href="file://{self.output_file}">Apri il file di output</a>')
            except Exception as e:
                QMessageBox.critical(self, 'Errore', f'Si è verificato un errore: {str(e)}')
        else:
            QMessageBox.warning(self, 'Attenzione', 'Seleziona prima il file di input e il nome del file di output')

    def load_input_data(self, input_file):
        df = pd.read_excel(input_file, engine='openpyxl')
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].apply(lambda x: ast.literal_eval(x) if isinstance(x, str) and x.startswith('[') else x)
        return df




    def map_data_to_template(self, input_df):
        template_columns = [
            "Codice ente competente per tutela",
            "Denominazione per esteso ente competente per tutela",
            "Tipo inventario",
            "Ente responsabile dell'indagine/concessionario",
            "Responsabile scientifico dell'indagine",
            "Anno indagine",
            "Luogo dell'indagine/contesto",
            "Denominazione indagine",
            "Metodo utilizzato per l'indagine",
            "Area/settore/ambiente/quadrato",
            "Unità Stratigrafica",
            "Deposizione funeraria",
            "Note sul contesto di ritrovamento",
            "Codice assegnato al reperto/ai materiali (sigla/numero cassetta)",
            "Definizione oggetto/lotto di materiali",
            "Tipologia/altre specifiche",
            "Categoria materiale",
            "Classe e produzione",
            "Classificazione/repertorio",
            "Quantità degli oggetti",
            "Descrizione",
            "Datazione/fascia cronologica-periodo",
            "Datazione specifica/da",
            "Datazione specifica/a",
            "Misure/tipo",
            "Misure/unità di misura",
            "Misure/valore",
            "Stato di conservazione",
            "Note stato conservazione",
            "Luogo di conservazione/Regione",
            "Luogo di conservazione/Provincia",
            "Luogo di conservazione/Comune",
            "Luogo di conservazione/ Denominazione deposito",
            "Luogo di conservazione/Indirizzo",
            "Luogo di conservazione/specifiche di collocazione",
            "Data inventario",
            "Riferimento immagine (nome file)",
            "Note e osservazioni"
        ]

        # Richiede l'inserimento manuale per i campi specifici
        #codice_ente, _ = QInputDialog.getText(self, 'Inserisci Codice', 'Codice ente competente per tutela:')
        denominazione_ente, _ = QInputDialog.getText(self, 'Inserisci Denominazione', 'Denominazione per esteso ente competente per tutela:')
        ente_responsabile, _ = QInputDialog.getText(self, 'Inserisci Ente Responsabile', 'Ente responsabile dell\'indagine/concessionario:')
        responsabile_scientifico, _ = QInputDialog.getText(self, 'Inserisci Responsabile', 'Responsabile scientifico dell\'indagine:')
        denominazione_indagine, _ = QInputDialog.getText(self, 'Inserisci Denominazione', 'Denominazione indagine:')
        metodo_indagine, _ = QInputDialog.getText(self, 'Inserisci Metodo Indagine', 'Metodo utilizzato per l\'indagine:')
        regione, _ = QInputDialog.getText(self, 'Inserisci la regione del deposito',
                                                  'Luogo di conservazione/Regione:')
        comune, _ = QInputDialog.getText(self, 'Inserisci la provincia del deposito',
                                                  'Luogo di conservazione/Provincia:')
        provincia, _ = QInputDialog.getText(self, 'Inserisci il comune del depoisto',
                                                  'Luogo di conservazione/Comune:')
        indirizzo, _ = QInputDialog.getText(self, 'Inserisci l\'indirizzo del deposito',
                                            'Luogo di conservazione/Indirizzo:')
        specifiche, _ = QInputDialog.getText(self, 'Inserisci specifiche di collocazione',
                                            'Luogo di conservazione/specifiche di collocazione:')

        mapped_data = []

        for _, row in input_df.iterrows():
            mapped_row = {
                "Codice ente competente per tutela": '',
                "Denominazione per esteso ente competente per tutela": denominazione_ente,
                "Ente responsabile dell'indagine/concessionario": ente_responsabile,
                "Responsabile scientifico dell'indagine": responsabile_scientifico,
                "Denominazione Indagine": denominazione_indagine,
                "Metodo utilizzato per l'indagine": metodo_indagine,
                "Luogo dell'indagine/contesto": row.get('sito', ''),
                "Tipo inventario": row.get('tipo_reperto', ''),
                "Tipologia/altre specifiche": row.get('tipo', ''),
                "Codice assegnato al reperto/ai materiali (sigla/numero cassetta)": row.get('n_reperto/nr_cassa', ''),
                "Definizione oggetto/lotto di materiali": row.get('definizione', ''),
                "Descrizione": row.get('descrizione', ''),
                "Area/settore/ambiente/quadrato": row.get('area', ''),
                "Unità Stratigrafica": row.get('us', ''),
                "Luogo di conservazione/specifiche di collocazione": f"Lavato: {row.get('lavato', '')}, Nr. cassa: {row.get('nr_cassa', '')}",
                "Luogo di conservazione/ Denominazione deposito": row.get('luogo_conservazione', ''),
                "Stato di conservazione": row.get('stato_conservazione', ''),
                "Datazione/fascia cronologica-periodo": row.get('datazione_reperto', ''),
                "Luogo di conservazione/Regione":regione,
                "Luogo di conservazione/Provincia":provincia,
                "Luogo di conservazione/Comune":comune,
                "Luogo di conservazione/Indirizzo":indirizzo,
                "Luogo di conservazione / specifiche di collocazione": specifiche,
                "Anno indagine": row.get('years', ''),
                "Riferimento immagine (nome file)": row.get('media_names', '')
            }


            def process_quantity(q):
                if not q or q == []:
                    return [], "", []

                t, u, v = [], [], []
                for measurement in q:
                    if isinstance(measurement, list) and len(measurement) == 3:
                        t.append(str(measurement[0]))
                        u.append(str(measurement[1]))
                        v.append(str(measurement[2]))

                # Se tutte le unità sono uguali, usa solo la prima
                unique_unit = u[0] if u and len(set(u)) == 1 else ", ".join(u)

                return t, unique_unit, v
            # Elabora le misurazioni
            if 'elementi_reperto' in row:
                t, u, v = process_quantity(row['elementi_reperto'])
                mapped_row["Quantità degli oggetti"] = "/".join(v) if v else ""

            else:
                mapped_row["Quantità degli oggetti"] = ""




            def process_measurements(measurements):
                if not measurements or measurements == []:
                    return [], "", []

                types, units, values = [], [], []
                for measurement in measurements:
                    if isinstance(measurement, list) and len(measurement) == 3:
                        types.append(str(measurement[0]))
                        units.append(str(measurement[1]))
                        values.append(str(measurement[2]))

                # Se tutte le unità sono uguali, usa solo la prima
                unique_unit = units[0] if units and len(set(units)) == 1 else ", ".join(units)

                return types, unique_unit, values
            # Elabora le misurazioni
            if 'misurazioni' in row:
                types, unit, values = process_measurements(row['misurazioni'])
                mapped_row["Misure/tipo"] = "/".join(types) if types else ""
                mapped_row["Misure/unità di misura"] = unit
                mapped_row["Misure/valore"] = "/".join(values) if values else ""
            else:
                mapped_row["Misure/tipo"] = ""
                mapped_row["Misure/unità di misura"] = ""
                mapped_row["Misure/valore"] = ""
            if isinstance(row.get('elementi_reperto'), list) and len(row['elementi_reperto']) > 0:
                mapped_row["Note e osservazioni"] = ", ".join([f"{item[0]}: {item[1]}" for item in row['elementi_reperto']])

            mapped_data.append(mapped_row)

        result_df = pd.DataFrame(mapped_data, columns=template_columns)
        # Verifica se il DataFrame risultante è vuoto
        if result_df.empty:
            print("Attenzione: Nessun dato valido da esportare.")
            return None

        return result_df
    def create_archaeological_excel(self, input_file, output_file):
        try:
            input_df = self.load_input_data(input_file)
            mapped_df = self.map_data_to_template(input_df)

            if mapped_df is None or mapped_df.empty:
                QMessageBox.warning(self, "Attenzione", "Non ci sono dati da esportare in Excel.")
                return False

            # Ottieni la directory corrente del file
            current_dir = os.path.dirname(os.path.abspath(__file__))

            # Costruisci i percorsi relativi
            template_path = os.path.join(current_dir, 'template_materiali.xlsx')
            logo_path = os.path.join(current_dir, 'resized_logo.jpeg')

            # Verifica che i file necessari esistano
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"Il file template non è stato trovato: {template_path}")
            if not os.path.exists(logo_path):
                raise FileNotFoundError(f"Il file del logo non è stato trovato: {logo_path}")

            # Carica il template esistente
            wb = load_workbook(template_path)
            ws = wb.active

            # Inserisci il logo ridimensionato nella cella A1
            logo_img = Image(logo_path)
            ws.add_image(logo_img, 'A1')

            # Copia la larghezza delle colonne dalla prima riga e l'altezza delle righe dalla seconda riga
            first_row = ws.row_dimensions[1].height
            first_col_widths = {col_letter: ws.column_dimensions[col_letter].width for col_letter in
                                ws.column_dimensions}

            # Imposta l'altezza delle righe e la larghezza delle colonne uniformemente per tutte le righe e colonne
            for row_num in range(3, ws.max_row + len(mapped_df) + 1):
                ws.row_dimensions[row_num].height = first_row

            for col_letter, col_width in first_col_widths.items():
                ws.column_dimensions[col_letter].width = col_width

            # Determina la riga di partenza per l'inserimento dei dati
            start_row = 3  # I dati iniziano dalla terza riga

            # Inserisci i dati mappati nel template
            for index, row in mapped_df.iterrows():
                for col_num, value in enumerate(row, start=1):
                    cell = ws.cell(row=start_row + index, column=col_num, value=value)

                    # Aggiungi il bordo nero a tutte le celle
                    thin_border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                    cell.border = thin_border

            # Salva il nuovo file
            wb.save(output_file)
            QMessageBox.information(self, "Successo", f"File Excel salvato con successo in:\n{output_file}")
            return True

        except FileNotFoundError as e:
            QMessageBox.critical(self, "Errore", str(e))
        except Exception as e:
            QMessageBox.critical(self, "Errore",
                                 f"Si è verificato un errore durante l'esportazione in Excel:\n{str(e)}")

        return False


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ArchaeologicalDataMapper(None)  # Passa None come parent
    ex.show()  # Mostra l'interfaccia
    sys.exit(app.exec())
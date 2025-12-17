#! /usr/bin/env python
# -*- coding: utf 8 -*-
"""
/***************************************************************************
        pyArchInit Plugin  - A QGIS plugin to manage archaeological dataset
                             stored in Postgres
                             -------------------
    begin                : 2007-12-01
    copyright            : (C) 2008 by Luca Mandolesi; Enzo Cocca <enzo.ccc@gmail.com>
    email                : mandoluca at gmail.com
 ***************************************************************************/
/***************************************************************************
 *                                                                          *
 *   This program is free software; you can redistribute it and/or modify   *
 *   it under the terms of the GNU General Public License as published by   *
 *   the Free Software Foundation; either version 2 of the License, or      *
 *   (at your option) any later version.                                    *
 ***************************************************************************/
"""
from __future__ import absolute_import

import platform
import re
import subprocess
import time

import cv2
from builtins import range
import math
from datetime import date
from qgis.core import *
import numpy as np
import urllib.parse
import pyvista as pv
import vtk
from pyvistaqt import QtInteractor
import functools
from qgis.PyQt.QtCore import *
from qgis.PyQt.QtGui import QColor, QIcon, QPixmap
from qgis.PyQt.QtWidgets import *
from collections import OrderedDict
from qgis.PyQt.uic import loadUiType

from ..modules.utility.VideoPlayerPottery import VideoPlayerWindow
from ..modules.utility.pyarchinit_media_utility import *
from ..gui.imageViewer import ImageViewer
from ..gui.sortpanelmain import SortPanelMain
from ..gui.quantpanelmain_pottery import QuantPanelMain
from ..modules.utility.pyarchinit_exp_POTTERYsheet_pdf import generate_POTTERY_pdf
from ..modules.utility.csv_writer import UnicodeWriter


from ..modules.db.pyarchinit_conn_strings import Connection
from ..modules.db.pyarchinit_db_manager import get_db_manager
from ..modules.db.concurrency_manager import ConcurrencyManager, RecordLockIndicator

from ..modules.utility.pyarchinit_error_check import Error_check
from ..modules.db.pyarchinit_utility import Utility
from ..gui.pyarchinitConfigDialog import pyArchInitDialog_Config
from ..modules.utility.remote_image_loader import load_icon, get_image_path, initialize as init_remote_loader

# Pottery Visual Similarity Search - lazy import to avoid startup errors
try:
    from ..modules.utility.pottery_similarity import (
        PotterySimilaritySearchEngine,
        PotterySimilarityWorker,
        get_available_models,
        get_embedding_updater,
        set_auto_update_enabled
    )
    HAS_SIMILARITY_SEARCH = True
except ImportError as e:
    print(f"Pottery similarity search not available: {e}")
    HAS_SIMILARITY_SEARCH = False

# Remote image loader for Cloudinary support
try:
    from ..modules.utility.remote_image_loader import RemoteImageLoader
    HAS_REMOTE_LOADER = True
except ImportError:
    HAS_REMOTE_LOADER = False



MAIN_DIALOG_CLASS, _ = loadUiType(
    os.path.join(os.path.dirname(__file__), os.pardir, 'gui', 'ui', 'pyarchinit_Pottery_ui.ui'))

class pyarchinit_Pottery(QDialog, MAIN_DIALOG_CLASS):
    L = QgsSettings().value("locale/userLocale")[0:2]
    if L == 'it':
        MSG_BOX_TITLE = "PyArchInit - Scheda Ceramica"
    elif L == 'en':
        MSG_BOX_TITLE = "PyArchInit - Pottery form"
    elif L == 'de':
        MSG_BOX_TITLE = "PyArchInit - Pottery formular"
    DATA_LIST = []
    DATA_LIST_REC_CORR = []
    DATA_LIST_REC_TEMP = []
    REC_CORR = 0
    REC_TOT = 0
    SITO = pyArchInitDialog_Config

    if L == 'it':
        STATUS_ITEMS = {"b": "Usa", "f": "Trova", "n": "Nuovo Record"}

    if L == 'de':
        STATUS_ITEMS = {"b": "Aktuell ", "f": "Finden", "n": "Neuer Rekord"}

    else:
        STATUS_ITEMS = {"b": "Current", "f": "Find", "n": "New Record"}
    BROWSE_STATUS = "b"
    SORT_MODE = 'asc'
    if L == 'it':
        SORTED_ITEMS = {"n": "Non ordinati", "o": "Ordinati"}
    if L == 'de':
        SORTED_ITEMS = {"n": "Nicht sortiert", "o": "Sortiert"}
    else:
        SORTED_ITEMS = {"n": "Not sorted", "o": "Sorted"}
    SORT_STATUS = "n"
    SORT_ITEMS_CONVERTED = ''
    UTILITY = Utility()
    DB_MANAGER = ""
    TABLE_NAME = 'pottery_table'
    MAPPER_TABLE_CLASS = "POTTERY"
    HOME = os.environ['PYARCHINIT_HOME']
    PDFFOLDER = '{}{}{}'.format(HOME, os.sep, "pyarchinit_PDF_folder")
    NOME_SCHEDA = "Scheda Pottery"
    ID_TABLE = "id_rep"
    ID_SITO = "sito"

    CONVERSION_DICT = {
    ID_TABLE:ID_TABLE,
    "ID Number":"id_number",
    "Sito":"sito",
    "Area":"area",
    "US":"us",
    "Box":"box",
    "Photo":"photo",
    "Drawing":"drawing",
    "Year":"anno",
    "Fabric":"fabric",
    "Percent":"percent",
    "Material":"material",
    "Shape":"form",
    "Specific form":"specific_form",
    "Ware":"ware",
    "Munsell color":"munsell",
    "Surface treatment":"surf_trat",
    "External decoration":"exdeco",
    "Internal decoration":"intdeco",
    "Wheel made":"wheel_made",
    "Description external decoration":"descrip_ex_deco",
    "Description internal decoration":"descrip_in_deco",
    "Note":"note",
    "Diameter Max":"diametro_max",
    "QTY":"qty",
    "Diameter Rim":"diametro_rim",
    "Diameter Bottom":"diametro_bottom",
    "Total height":"diametro_height",
    "Preserved height":"diametro_preserved",
    "Specific shape":"specific_shape",
    "Bag": "bag",
    "Sector": "sector",
    "Decoration type": "decoration_type",
    "Decoration motif": "decoration_motif",
    "Decoration position": "decoration_position",
    "Datazione": "datazione"
    }

    SORT_ITEMS = [
                ID_TABLE,
                "ID Number",
                "Sito",
                "Area",
                "US",
                "Box",
                "Photo",
                "Drawing",
                "Year",
                "Fabric",
                "Percent",
                "Material",
                "Shape",
                "Specific form",
                "Ware",
                "Munsell color",
                "Surface treatment",
                "External decoration",
                "Internal decoration",
                "Wheel made",
                "Description external decoration",
                "Description internal decoration",
                "Note",
                "Diameter Max",
                "QTY",
                "Diameter Rim",
                "Diameter Bottom",
                "Total height",
                "Preserved height",
                "Specific shape",
                "Bag",
                "Sector"
                ]
    QUANT_ITEMS = [
                'Fabric',
                'US',
                'Area',
                'Material',
                'Percent',
                'Shape',
                'Specific form',
                'Ware',
                'Munsell color',
                'Surface treatment',
                'External decoration',
                'Internal decoration',
                'Wheel made',
                ]

    TABLE_FIELDS_UPDATE = [
                    "id_number",
                    "sito",
                    "area",
                    "us",
                    "box",
                    "photo",
                    "drawing",
                    "anno",
                    "fabric",
                    "percent",
                    "material",
                    "form",
                    "specific_form",
                    "ware",
                    "munsell",
                    "surf_trat",
                    "exdeco",
                    "intdeco",
                    "wheel_made",
                    "descrip_ex_deco",
                    "descrip_in_deco",
                    "note",
                    "diametro_max",
                    "qty",
                    "diametro_rim",
                    "diametro_bottom",
                    "diametro_height",
                    "diametro_preserved",
                    "specific_shape",
                    "bag",
                    "sector",
                    "decoration_type",
                    "decoration_motif",
                    "decoration_position",
                    "datazione"
                    ]
    TABLE_FIELDS = [
                    "id_number",
                    "sito",
                    "area",
                    "us",
                    "box",
                    "photo",
                    "drawing",
                    "anno",
                    "fabric",
                    "percent",
                    "material",
                    "form",
                    "specific_form",
                    "ware",
                    "munsell",
                    "surf_trat",
                    "exdeco",
                    "intdeco",
                    "wheel_made",
                    "descrip_ex_deco",
                    "descrip_in_deco",
                    "note",
                    "diametro_max",
                    "qty",
                    "diametro_rim",
                    "diametro_bottom",
                    "diametro_height",
                    "diametro_preserved",
                    "specific_shape",
                    "bag",
                    "sector",
                    "decoration_type",
                    "decoration_motif",
                    "decoration_position",
                    "datazione"
                    ]
    LANG = {
        "IT": ['it_IT', 'IT', 'it', 'IT_IT', 'it_CH'],
        "EN": ['en_US', 'EN_US', 'en', 'EN', 'en_GB', 'en_AU', 'en_CA', 'en_NZ', 'en_IE', 'en_ZA'],
        "DE": ['de_DE', 'de', 'DE', 'DE_DE', 'de_AT', 'de_CH']
    }


    PDFFOLDER = '{}{}{}'.format(HOME, os.sep, "pyarchinit_PDF_folder")
    DB_SERVER = "not defined"  ####nuovo sistema sort
    QUANT_PATH = '{}{}{}'.format(HOME, os.sep, "pyarchinit_Quantificazioni_folder")

    def __init__(self, iface):
        super().__init__()
        self.iface = iface

        self.setupUi(self)
        self.mDockWidget_4.setHidden(True)
        self.mDockWidget_export.setHidden(True)
        self.setAcceptDrops(True)
        self.iconListWidget.setDragDropMode(QAbstractItemView.DragDrop)
        # Dizionario per memorizzare le immagini in cache
        self.image_cache = OrderedDict()
        self.video_player=None
        # Numero massimo di elementi nella cache
        self.cache_limit = 100
        self.currentLayerId = None
        try:
            self.on_pushButton_connect_pressed()
        except Exception as e:
            QMessageBox.warning(self, "Sistema di connessione", str(e),  QMessageBox.Ok)
        if len(self.DATA_LIST)==0:
            self.comboBox_sito.setCurrentIndex(0)
        else:
            self.comboBox_sito.setCurrentIndex(1)
        sito = self.comboBox_sito.currentText()
        self.comboBox_sito.setEditText(sito)

        self.fill_fields()
        self.msg_sito()
        self.set_sito()
        self.show()
        self.customize_GUI()
        self.toolButton_pdfpath.clicked.connect(self.setPathpdf)
        self.pbnOpenpdfDirectory.clicked.connect(self.openpdfDir)
        self.setnone()

        # Initialize remote image loader with credentials from QGIS settings
        init_remote_loader()

    def get_images_for_entities(self, entity_ids, log_signal=None):
        def log(message, level="info"):
            if log_signal:
                log_signal.emit(message, level)
        """Recupera le immagini dalla tabella mediaentity in base agli ID forniti."""
        log(f"Called get_images_for_entities with entity_ids: {entity_ids}")

        if not entity_ids:
            return []

        try:
            images = []
            conn = Connection()
            thumb_resize = conn.thumb_resize()
            thumb_resize_str = thumb_resize['thumb_resize']

            for entity_id in entity_ids:
                log(f"Called id table: {self.ID_TABLE}")
                # Usa la stessa logica di loadMediaPreview
                rec_list = self.ID_TABLE + " = " + str(entity_id)
                log(f"Called rec list: {rec_list}")
                # Usa la stessa logica di loadMediaPreview
                search_dict = {
                    'id_entity': "'" + str(entity_id) + "'",
                    'entity_type': "'CERAMICA'"
                }

                record_us_list = self.DB_MANAGER.query_bool(search_dict, 'MEDIATOENTITY')
                log(f"Found {len(record_us_list)} records for entity_id {entity_id}")  # Debug log

                for media_record in record_us_list:
                    search_dict = {'id_media': "'" + str(media_record.id_media) + "'"}
                    u = Utility()
                    search_dict = u.remove_empty_items_fr_dict(search_dict)
                    mediathumb_data = self.DB_MANAGER.query_bool(search_dict, "MEDIA_THUMB")
                    log(f"Found {len(mediathumb_data)} thumbs for media_id {media_record.id_media}")  # Debug log

                    if mediathumb_data:
                        thumb_path = str(mediathumb_data[0].path_resize)
                        images.append({
                            'id': entity_id,
                            'url': thumb_resize_str + thumb_path,
                            'caption': media_record.media_name
                        })
                        log(f"Added image with path: {thumb_resize_str + thumb_path}")  # Debug log

            log(f"Returning total of {len(images)} images")  # Debug log
            return images

        except Exception as e:
            log(f"Error in get_images_for_entities: {str(e)}")
            traceback.print_exc()  # Questo mostrerà lo stack trace completo
            return []


    def setnone(self):
        if self.lineEdit_diametro_max.text=='None' or None or 'NULL'or 'Null':
            self.lineEdit_diametro_max.clear()
            self.lineEdit_diametro_max.setText('')
            self.lineEdit_diametro_max.update()
        if self.lineEdit_diametro_rim.text()=='None' or None or 'NULL'or 'Null':
            self.lineEdit_diametro_rim.clear()
            self.lineEdit_diametro_rim.setText('')
            self.lineEdit_diametro_rim.update()
        if self.lineEdit_diametro_bottom.text=='None' or None or 'NULL'or 'Null':
            self.lineEdit_diametro_bottom.clear()
            self.lineEdit_diametro_bottom.setText('')
            self.lineEdit_diametro_bottom.update()
        if self.lineEdit_diametro_preserved.text == 'None' or None or 'NULL'or 'Null':
            self.lineEdit_diametro_preserved.clear()
            self.lineEdit_diametro_preserved.setText('')
            self.lineEdit_diametro_preserved.update()
        if self.lineEdit_diametro_height.text == 'None' or None or 'NULL'or 'Null':
            self.lineEdit_diametro_height.clear()
            self.lineEdit_diametro_height.setText('')
            self.lineEdit_diametro_height.update()



    def generate_list_foto(self):
        data_list_foto = []
        for i in range(len(self.DATA_LIST)):
            conn = Connection()
            thumb_path = conn.thumb_path()
            thumb_path_str = thumb_path['thumb_path']

            if thumb_path_str == '':
                if self.L == 'it':
                    QMessageBox.information(self, "Info",
                                            "devi settare prima la path per salvare le thumbnail . Vai in impostazioni di sistema/ path setting ")
                elif self.L == 'de':
                    QMessageBox.information(self, "Info",
                                            "müssen Sie zuerst den Pfad zum Speichern der Miniaturansichten und Videos festlegen. Gehen Sie zu System-/Pfad-Einstellung")
                else:
                    QMessageBox.information(self, "Message",
                                            "you must first set the path to save the thumbnails and videos. Go to system/path setting")
            else:
                search_dict = {'id_entity': "'" + str(eval("self.DATA_LIST[i].id_rep")) + "'", 'entity_type': "'CERAMICA'"}
                record_doc_list = self.DB_MANAGER.query_bool(search_dict, 'MEDIAVIEW')
                for media in record_doc_list:
                    thumbnail = (thumb_path_str + media.filepath)
                    foto = (media.id_media)

                    # Get photo and drawing fields from pottery record
                    photo_val = ''
                    drawing_val = ''
                    if hasattr(self.DATA_LIST[i], 'photo') and self.DATA_LIST[i].photo:
                        photo_val = str(self.DATA_LIST[i].photo)
                    if hasattr(self.DATA_LIST[i], 'drawing') and self.DATA_LIST[i].drawing:
                        drawing_val = str(self.DATA_LIST[i].drawing)

                    data_list_foto.append([
                        str(self.DATA_LIST[i].sito),  # 0 - Sito
                        str(self.DATA_LIST[i].area),  # 1 - Area
                        str(self.DATA_LIST[i].us),  # 2 - US
                        str(self.DATA_LIST[i].sector),  # 3 - Sector
                        str(self.DATA_LIST[i].anno),  # 4 - Anno
                        str(self.DATA_LIST[i].id_number),  # 5 - ID Number
                        str(self.DATA_LIST[i].note),  # 6 - Note
                        str(foto),  # 7 - Foto (media_id)
                        str(thumbnail),  # 8 - Thumbnail
                        photo_val,  # 9 - Photo (from pottery_table)
                        drawing_val])  # 10 - Drawing (from pottery_table)

        return data_list_foto

        # #####################fine########################

    # def generate_list_pdf(self):
    #     data_list = []
    #     for i in range(len(self.DATA_LIST)):
    #         data_list.append([
    #             str(self.DATA_LIST[i].divelog_id),
    #             str(self.DATA_LIST[i].artefact_id),
    #             str(self.DATA_LIST[i].site),
    #             str(self.DATA_LIST[i].area),
    #             str(self.DATA_LIST[i].inclusions),
    #             str(self.DATA_LIST[i].form),
    #             str(self.DATA_LIST[i].specific_part),
    #             str(self.DATA_LIST[i].category),
    #             str(self.DATA_LIST[i].typology),
    #             str(self.DATA_LIST[i].depth),
    #             str(self.DATA_LIST[i].retrieved),
    #             str(self.DATA_LIST[i].percent_inclusion),
    #             str(self.DATA_LIST[i].provenance),
    #             str(self.DATA_LIST[i].munsell_clay),
    #             str(self.DATA_LIST[i].munsell_surf),
    #             str(self.DATA_LIST[i].surf_treatment),
    #             str(self.DATA_LIST[i].conservation),
    #             str(self.DATA_LIST[i].storage_),
    #             str(self.DATA_LIST[i].period),
    #             str(self.DATA_LIST[i].state),
    #             str(self.DATA_LIST[i].samples),
    #             str(self.DATA_LIST[i].washed),
    #             str(self.DATA_LIST[i].dm),
    #             str(self.DATA_LIST[i].dr),
    #             str(self.DATA_LIST[i].db),
    #             str(self.DATA_LIST[i].th),
    #             str(self.DATA_LIST[i].ph),
    #             str(self.DATA_LIST[i].bh),
    #             str(self.DATA_LIST[i].thickmin),
    #             str(self.DATA_LIST[i].thickmax),
    #             str(self.DATA_LIST[i].date_),
    #             str(self.DATA_LIST[i].years),
    #             str(self.DATA_LIST[i].description),
    #             str(self.DATA_LIST[i].photographed),
    #             str(self.DATA_LIST[i].drawing),
    #             str(self.DATA_LIST[i].wheel_made)
    #         ])
    #     return data_list

    def on_pushButton_print_pressed(self):

        # if self.checkBox_s_pottery.isChecked():
        #     pottery_pdf_sheet = generate_POTTERY_pdf()
        #     data_list = self.generate_list_pdf()
        #     pottery_pdf_sheet.build_POTTERY_sheets(data_list)
        #     QMessageBox.warning(self, 'Ok', "Export completed", QMessageBox.Ok)
        # else:
        #     pass



        if self.checkBox_e_foto_t.isChecked():
            POTTERY_index_pdf = generate_POTTERY_pdf()
            data_list_foto = self.generate_list_foto()

            try:
                if bool(data_list_foto):
                    POTTERY_index_pdf.build_index_Foto_2(data_list_foto, data_list_foto[0][0])
                    QMessageBox.warning(self, 'Ok', "Export completed", QMessageBox.Ok)

                else:
                    QMessageBox.warning(self, 'Warning',
                                        "Pottery list photo can't to be exported, you must tag before the pics",
                                        QMessageBox.Ok)
            except Exception as e:
                QMessageBox.warning(self, 'Warning', str(e), QMessageBox.Ok)

        if self.checkBox_e_foto.isChecked():
            POTTERY_index_pdf = generate_POTTERY_pdf()
            data_list_foto = self.generate_list_foto()

            try:
                if bool(data_list_foto):
                    POTTERY_index_pdf.build_index_Foto(data_list_foto, data_list_foto[0][0])
                    QMessageBox.warning(self, 'Ok', "Export completed", QMessageBox.Ok)

                else:
                    QMessageBox.warning(self, 'Warniong',
                                        "Pottery list photo can't to be exported because the image are not tagged",
                                        QMessageBox.Ok)
            except Exception as e:
                QMessageBox.warning(self, 'Warning', str(e), QMessageBox.Ok)

    def setPathpdf(self):
        s = QgsSettings()
        dbpath = QFileDialog.getOpenFileName(
            self,
            "Set file name",
            self.PDFFOLDER,
            " PDF (*.pdf)"
        )[0]
        # filename=dbpath.split("/")[-1]
        if dbpath:
            self.lineEdit_pdf_path.setText(dbpath)
            s.setValue('', dbpath)

    def openpdfDir(self):

        path = '{}{}{}'.format(self.HOME, os.sep, "pyarchinit_PDF_folder")
        if platform.system() == "Windows":
            os.startfile(path)
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])

    # def on_pushButton_convert_pressed(self):
    #     # if not bool(self.setPathpdf()):
    #     # QMessageBox.warning(self, "INFO", "devi scegliere un file pdf",
    #     # QMessageBox.Ok)
    #     try:
    #         pdf_file = self.lineEdit_pdf_path.text()
    #         filename = pdf_file.split("/")[-1]
    #         docx_file = self.PDFFOLDER + '/' + filename + '.docx'
    #         # convert pdf to docx
    #         parse(pdf_file, docx_file, start=self.lineEdit_pag1.text(), end=self.lineEdit_pag2.text())
    #
    #         QMessageBox.information(self, "INFO", "Conversion completed",
    #                                 QMessageBox.Ok)
    #     except Exception as e:
    #         QMessageBox.warning(self, "Error", str(e),
    #                             QMessageBox.Ok)

    def msg_sito(self):
        # self.model_a.database().close()
        conn = Connection()
        sito_set = conn.sito_set()
        sito_set_str = sito_set['sito_set']
        if bool(self.comboBox_sito.currentText()) and self.comboBox_sito.currentText() == sito_set_str:

            if self.L == 'it':
                QMessageBox.information(self, "OK", "Sei connesso al sito: %s" % str(sito_set_str), QMessageBox.Ok)

            elif self.L == 'de':
                QMessageBox.information(self, "OK",
                                        "Sie sind mit der archäologischen Stätte verbunden: %s" % str(sito_set_str),
                                        QMessageBox.Ok)

            else:
                QMessageBox.information(self, "OK", "You are connected to the site: %s" % str(sito_set_str),
                                        QMessageBox.Ok)

        elif sito_set_str == '':
            if self.L == 'it':
                msg = QMessageBox.information(self, "Attenzione",
                                              "Non hai settato alcun sito. Vuoi settarne uno? click Ok altrimenti Annulla per  vedere tutti i record",
                                              QMessageBox.Ok | QMessageBox.Cancel)
            elif self.L == 'de':
                msg = QMessageBox.information(self, "Achtung",
                                              "Sie haben keine archäologischen Stätten eingerichtet. Klicken Sie auf OK oder Abbrechen, um alle Datensätze zu sehen",
                                              QMessageBox.Ok | QMessageBox.Cancel)
            else:
                msg = QMessageBox.information(self, "Warning",
                                              "You have not set up any archaeological site. Do you want to set one? click Ok otherwise Cancel to see all records",
                                              QMessageBox.Ok | QMessageBox.Cancel)
            if msg == QMessageBox.Cancel:
                pass
            else:
                dlg = pyArchInitDialog_Config(self)
                dlg.charge_list()
                dlg.exec_()
    def set_sito(self):
        # self.model_a.database().close()
        conn = Connection()
        sito_set = conn.sito_set()
        sito_set_str = sito_set['sito_set']
        try:
            if bool(sito_set_str):
                search_dict = {
                    'sito': "'" + str(sito_set_str) + "'"}  # 1 - Sito
                u = Utility()
                search_dict = u.remove_empty_items_fr_dict(search_dict)
                res = self.DB_MANAGER.query_bool(search_dict, self.MAPPER_TABLE_CLASS)
                self.DATA_LIST = []
                for i in res:
                    self.DATA_LIST.append(i)
                self.REC_TOT, self.REC_CORR = len(self.DATA_LIST), 0
                self.DATA_LIST_REC_TEMP = self.DATA_LIST_REC_CORR = self.DATA_LIST[0]  ####darivedere
                self.fill_fields()
                self.BROWSE_STATUS = "b"
                self.SORT_STATUS = "n"
                self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                self.set_rec_counter(len(self.DATA_LIST), self.REC_CORR + 1)
                self.setComboBoxEnable(["self.comboBox_sito"], "False")
            else:
                pass  #
        except Exception as e:
            if self.L == 'it':

                QMessageBox.information(self, "Attenzione", "Non esiste questo sito: "'"' + str(
                    sito_set_str) + '"'" in questa scheda, Per favore distattiva la 'scelta sito' dalla scheda di configurazione plugin per vedere tutti i record oppure crea la scheda",
                                        QMessageBox.Ok)
            elif self.L == 'de':

                QMessageBox.information(self, "Warnung", "Es gibt keine solche archäologische Stätte: "'""' + str(
                    sito_set_str) + '"'" in dieser Registerkarte, Bitte deaktivieren Sie die 'Site-Wahl' in der Plugin-Konfigurationsregisterkarte, um alle Datensätze zu sehen oder die Registerkarte zu erstellen",
                                        QMessageBox.Ok)
            else:

                QMessageBox.information(self, "Warning", "There is no such site: "'"' + str(
                    sito_set_str) + '"'" in this tab, Please disable the 'site choice' from the plugin configuration tab to see all records or create the tab",
                                        QMessageBox.Ok)

    def on_pushButtonQuant_pressed(self):
        dlg = QuantPanelMain(self)
        dlg.insertItems(self.QUANT_ITEMS)
        dlg.exec_()
        dataset = []
        parameter1 = dlg.TYPE_QUANT
        parameters2 = dlg.ITEMS
        contatore = 0
        if parameter1 == 'QTY':
            for i in range(len(self.DATA_LIST)):
                temp_dataset = ()
                try:
                    temp_dataset = (self.parameter_quant_creator(parameters2, i), int(self.DATA_LIST[i].qty))
                    contatore += int(self.DATA_LIST[i].qty) #conteggio totale
                    dataset.append(temp_dataset)
                except  Exception as e:
                    #QMessageBox.warning(self, "Error", str(e),  QMessageBox.Ok)
                    print(e)
            #QMessageBox.warning(self, "Totale", str(contatore),  QMessageBox.Ok)
            if bool(dataset):
                dataset_sum = self.UTILITY.sum_list_of_tuples_for_value(dataset)
                csv_dataset = []
                for sing_tup in dataset_sum:
                    sing_list = [sing_tup[0], str(sing_tup[1])]
                    csv_dataset.append(sing_list)
                filename = ('%s%squant_qty.csv') % (self.QUANT_PATH, os.sep)
                #QMessageBox.warning(self, "Esportazione", str(filename), MessageBox.Ok)
                f = open(filename, 'wb')
                Uw = UnicodeWriter(f)
                Uw.writerows(csv_dataset)
                f.close()
                self.plot_chart(dataset_sum, 'Frequency analisys', 'Qty')
            else:
                QMessageBox.warning(self, "Warning", "The datas not are present",  QMessageBox.Ok)
    def parameter_quant_creator(self, par_list, n_rec):
        self.parameter_list = par_list
        self.record_number = n_rec
        converted_parameters = []
        for par in self.parameter_list:
            converted_parameters.append(self.CONVERSION_DICT[par])
        parameter2 = ''
        for sing_par_conv in range(len(converted_parameters)):
            exec_str =  ('str(self.DATA_LIST[%d].%s)') % (self.record_number, converted_parameters[sing_par_conv])
            paramentro = str(self.parameter_list[sing_par_conv])
            exec_str = ' -' + paramentro[:4] + ": " + eval(exec_str)
            parameter2 += exec_str
        return parameter2
    def plot_chart(self, d, t, yl):
        self.data_list = d
        self.title = t
        self.ylabel = yl
        if type(self.data_list) == list:
            data_diz = {}
            for item in self.data_list:
                data_diz[item[0]] = item[1]
        x = list(range(len(data_diz)))
        n_bars = len(data_diz)
        values = list(data_diz.values())
        teams = list(data_diz.keys())
        ind = np.arange(n_bars)
        #randomNumbers = random.sample(range(0, 10), 10)
        self.widget.canvas.ax.clear()
        #QMessageBox.warning(self, "Alert", str(teams) ,  QMessageBox.Ok)
        bars = self.widget.canvas.ax.bar(x, height=values, width=0.5, align='center', alpha=0.4,picker=5)
        #guardare il metodo barh per barre orizzontali
        self.widget.canvas.ax.set_title(self.title)
        self.widget.canvas.ax.set_ylabel(self.ylabel)
        l = []
        for team in teams:
            l.append('""')
        #self.widget.canvas.ax.set_xticklabels(x , ""   ,size = 'x-small', rotation = 0)
        n = 0
        for bar in bars:
            val = int(bar.get_height())
            x_pos = bar.get_x() + 0.25
            label  = teams[n]+ ' - ' + str(val)
            y_pos = 0.1 #bar.get_height() - bar.get_height() + 1
            self.widget.canvas.ax.tick_params(axis='x', labelsize=8)
            #self.widget.canvas.ax.set_xticklabels(ind + x, ['fg'], position = (x_pos,y_pos), xsize = 'small', rotation = 90)
            self.widget.canvas.ax.text(x_pos, y_pos, label,zorder=0, ha='center', va='bottom',size = 'x-small', rotation = 90)
            n+=1
        #self.widget.canvas.ax.plot(randomNumbers)
        self.widget.canvas.draw()

    def enable_button(self, n):
        self.pushButton_connect.setEnabled(n)

        self.pushButton_new_rec.setEnabled(n)

        self.pushButton_view_all.setEnabled(n)

        self.pushButton_first_rec.setEnabled(n)

        self.pushButton_last_rec.setEnabled(n)

        self.pushButton_prev_rec.setEnabled(n)

        self.pushButton_next_rec.setEnabled(n)

        self.pushButton_delete.setEnabled(n)

        self.pushButton_new_search.setEnabled(n)

        self.pushButton_search_go.setEnabled(n)

        self.pushButton_sort.setEnabled(n)

    def enable_button_search(self, n):
        self.pushButton_connect.setEnabled(n)

        self.pushButton_new_rec.setEnabled(n)

        self.pushButton_view_all.setEnabled(n)

        self.pushButton_first_rec.setEnabled(n)

        self.pushButton_last_rec.setEnabled(n)

        self.pushButton_prev_rec.setEnabled(n)

        self.pushButton_next_rec.setEnabled(n)

        self.pushButton_delete.setEnabled(n)

        self.pushButton_save.setEnabled(n)

        self.pushButton_sort.setEnabled(n)




    def on_pushButton_connect_pressed(self):

        conn = Connection()
        conn_str = conn.conn_str()
        test_conn = conn_str.find('sqlite')
        if test_conn == 0:
            self.DB_SERVER = "sqlite"

        try:
            self.DB_MANAGER = get_db_manager(conn_str, use_singleton=True)

            # Initialize concurrency manager with current username
            self.concurrency_manager = ConcurrencyManager(self)
            # Get username from connection settings
            s = QgsSettings()
            db_username = s.value('pyArchInit/current_user', '', type=str)
            if not db_username:
                # Fallback to lineEdit_User from config if available
                db_username = conn.conn_str().split('/')[-1].split('@')[0] if '@' in conn.conn_str() else ''
            if not db_username:
                import getpass
                db_username = getpass.getuser()
            self.concurrency_manager.set_username(db_username)
            print(f"Pottery form: ConcurrencyManager initialized with username: {db_username}")

            # Initialize lock indicator (a simple status object)
            self.lock_indicator = RecordLockIndicator(self)

            self.charge_records()  # charge records from DB
            # check if DB is empty
            if self.DATA_LIST:
                self.REC_TOT, self.REC_CORR = len(self.DATA_LIST), 0
                self.DATA_LIST_REC_TEMP = self.DATA_LIST_REC_CORR = self.DATA_LIST[0]
                self.BROWSE_STATUS = 'b'
                self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                self.label_sort.setText(self.SORTED_ITEMS["n"])
                self.set_rec_counter(len(self.DATA_LIST), self.REC_CORR + 1)
                self.charge_list()
                self.fill_fields()
                #self.setComboBoxEnable(["self.comboBox_area"], "False")
                #self.setComboBoxEnable(["self.lineEdit_us"], "False")
            else:
                if self.L == 'it':
                    QMessageBox.warning(self, "BENVENUTO",
                                        "Benvenuto in pyArchInit " + self.NOME_SCHEDA + ". Il database e' vuoto. Premi 'Ok' e buon lavoro!",
                                        QMessageBox.Ok)
                elif self.L == 'de':
                    QMessageBox.warning(self, "WILLKOMMEN",
                                        "WILLKOMMEN in pyArchInit" + "SE-MSE formular" + ". Die Datenbank ist leer. Tippe 'Ok' und aufgehts!",
                                        QMessageBox.Ok)
                else:
                    QMessageBox.warning(self, "WELCOME",
                                        "Welcome in pyArchInit" + "Samples SU-WSU" + ". The DB is empty. Push 'Ok' and Good Work!",
                                        QMessageBox.Ok)
                self.charge_list()
                self.BROWSE_STATUS = 'x'
                #self.setComboBoxEnable(["self.comboBox_area"], "True")
                #self.setComboBoxEnable(["self.lineEdit_us"], "True")
                self.on_pushButton_new_rec_pressed()
        except Exception as e:
            e = str(e)
            if e.find("no such table"):
                if self.L == 'it':
                    msg = "La connessione e' fallita {}. " \
                          "E' NECESSARIO RIAVVIARE QGIS oppure rilevato bug! Segnalarlo allo sviluppatore".format(
                        str(e))
                    self.iface.messageBar().pushMessage(self.tr(msg), Qgis.Warning, 0)

                elif self.L == 'de':
                    msg = "Verbindungsfehler {}. " \
                          " QGIS neustarten oder es wurde ein bug gefunden! Fehler einsenden".format(str(e))
                    self.iface.messageBar().pushMessage(self.tr(msg), Qgis.Warning, 0)
                else:
                    msg = "The connection failed {}. " \
                          "You MUST RESTART QGIS or bug detected! Report it to the developer".format(str(e))
            else:
                if self.L == 'it':
                    msg = "Attenzione rilevato bug! Segnalarlo allo sviluppatore. Errore: ".format(str(e))
                    self.iface.messageBar().pushMessage(self.tr(msg), Qgis.Warning, 0)
                elif self.L == 'de':
                    msg = "ACHTUNG. Es wurde ein bug gefunden! Fehler einsenden: ".format(str(e))
                    self.iface.messageBar().pushMessage(self.tr(msg), Qgis.Warning, 0)
                else:
                    msg = "Warning bug detected! Report it to the developer. Error: ".format(str(e))
                    self.iface.messageBar().pushMessage(self.tr(msg), Qgis.Warning, 0)


    def on_toolButtonPreview_toggled(self):
        if self.L=='it':
            if self.toolButtonPreview.isChecked():
                QMessageBox.warning(self, "Messaggio",
                                    "Modalita' Preview US attivata. Le piante delle US saranno visualizzate nella sezione Piante",
                                    QMessageBox.Ok)
                self.tabWidget.setCurrentIndex(10)  # Set the current tab to the map preview tab
                self.loadMapPreview()
            else:
                self.loadMapPreview(1)
        elif self.L=='de':
            if self.toolButtonPreview.isChecked():
                QMessageBox.warning(self, "Message",
                                    "Modalität' Preview der aktivierten SE. Die Plana der SE werden in der Auswahl der Plana visualisiert",
                                    QMessageBox.Ok)
                self.tabWidget.setCurrentIndex(10)  # Set the current tab to the map preview tab
                self.loadMapPreview()
            else:
                self.tabWidget.setCurrentIndex(0)
                self.loadMapPreview(1)
        else:
            if self.toolButtonPreview.isChecked():
                QMessageBox.warning(self, "Message",
                                    "Preview SU mode enabled. US plants will be displayed in the Plants section",
                                    QMessageBox.Ok)
                self.tabWidget.setCurrentIndex(10)  # Set the current tab to the map preview tab
                self.loadMapPreview()
            else:
                self.tabWidget.setCurrentIndex(0)
                self.loadMapPreview(1)
    def customize_GUI(self):
        # media prevew system
        self.iconListWidget.setDragEnabled(True)
        self.iconListWidget.setAcceptDrops(True)
        self.iconListWidget.setDropIndicatorShown(True)

        self.iconListWidget.setLineWidth(2)
        self.iconListWidget.setMidLineWidth(2)
        # self.iconListWidget.setProperty("showDropIndicator", False)
        self.iconListWidget.setIconSize(QSize(430, 570))

        self.iconListWidget.setUniformItemSizes(True)
        self.iconListWidget.setObjectName("iconListWidget")
        self.iconListWidget.SelectionMode()
        self.iconListWidget.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.iconListWidget.itemDoubleClicked.connect(self.openWide_image)

        # Setup Statistics Tab
        self.setup_statistics_tab()

        # Setup Visual Similarity Search
        self.setup_similarity_search_ui()

        # Setup Embedding Index Auto-Updater
        self.setup_embedding_auto_updater()

        # Setup Filter Button for ID Number selection
        self.setup_filter_button()

        # Setup Auto-Populate Photo/Drawing button
        self.setup_auto_populate_button()

    def setup_auto_populate_button(self):
        """Setup the button to auto-populate photo and drawing fields from media associations"""
        try:
            from qgis.PyQt.QtWidgets import QPushButton
            from qgis.PyQt.QtCore import QSize

            # Create the auto-populate button
            self.pushButton_auto_populate_photo = QPushButton(self)
            if self.L == 'it':
                self.pushButton_auto_populate_photo.setText("Popola Foto/Disegni")
                self.pushButton_auto_populate_photo.setToolTip(
                    "Auto-popola i campi Photo e Drawing dai media associati")
            elif self.L == 'de':
                self.pushButton_auto_populate_photo.setText("Fotos/Zeich. füllen")
                self.pushButton_auto_populate_photo.setToolTip(
                    "Foto- und Zeichnungsfelder aus Medienverknüpfungen automatisch füllen")
            else:
                self.pushButton_auto_populate_photo.setText("Populate Photo/Drawing")
                self.pushButton_auto_populate_photo.setToolTip(
                    "Auto-populate Photo and Drawing fields from media associations")

            # Style the button
            self.pushButton_auto_populate_photo.setMinimumSize(QSize(120, 25))
            self.pushButton_auto_populate_photo.setMaximumSize(QSize(150, 25))

            # Find the parent layout of pushButton_print and insert after it
            try:
                parent_layout = self.pushButton_print.parent().layout()
                if parent_layout:
                    index = parent_layout.indexOf(self.pushButton_print)
                    if index >= 0:
                        parent_layout.insertWidget(index + 1, self.pushButton_auto_populate_photo)
                    else:
                        parent_layout.addWidget(self.pushButton_auto_populate_photo)
                else:
                    # Alternative: position next to lineEdit_photo
                    photo_geom = self.lineEdit_photo.geometry()
                    self.pushButton_auto_populate_photo.setGeometry(
                        photo_geom.x() + photo_geom.width() + 5,
                        photo_geom.y(),
                        140,
                        photo_geom.height()
                    )
            except:
                # Fallback: position it relative to lineEdit_photo if available
                if hasattr(self, 'lineEdit_photo'):
                    photo_geom = self.lineEdit_photo.geometry()
                    self.pushButton_auto_populate_photo.setGeometry(
                        photo_geom.x() + photo_geom.width() + 5,
                        photo_geom.y(),
                        140,
                        photo_geom.height()
                    )

            # Connect the button
            self.pushButton_auto_populate_photo.clicked.connect(self.populate_photo_drawing_from_media)

        except Exception as e:
            print(f"Error setting up auto-populate button: {e}")

    def setup_filter_button(self):
        """Setup the filter button for ID Number selection"""
        # Find the layout where pushButton_view_all is located
        # and add a filter button next to it
        try:
            # Create the filter button programmatically
            self.pushButton_filter_pottery = QPushButton(self)
            if self.L == 'it':
                self.pushButton_filter_pottery.setText("Filtra ID")
                self.pushButton_filter_pottery.setToolTip("Filtra i record per ID Number")
            elif self.L == 'de':
                self.pushButton_filter_pottery.setText("Filter ID")
                self.pushButton_filter_pottery.setToolTip("Datensätze nach ID-Nummer filtern")
            else:
                self.pushButton_filter_pottery.setText("Filter ID")
                self.pushButton_filter_pottery.setToolTip("Filter records by ID Number")

            # Style the button to match existing buttons
            self.pushButton_filter_pottery.setMinimumSize(QSize(50, 25))
            self.pushButton_filter_pottery.setMaximumSize(QSize(100, 25))

            # Find the parent layout of pushButton_view_all and insert the filter button
            parent_layout = self.pushButton_view_all.parent().layout()
            if parent_layout:
                # Find the index of pushButton_view_all
                index = parent_layout.indexOf(self.pushButton_view_all)
                if index >= 0:
                    parent_layout.insertWidget(index + 1, self.pushButton_filter_pottery)
                else:
                    # Fallback: add to the end
                    parent_layout.addWidget(self.pushButton_filter_pottery)
            else:
                # Alternative: add next to view_all button manually
                view_all_geom = self.pushButton_view_all.geometry()
                self.pushButton_filter_pottery.setGeometry(
                    view_all_geom.x() + view_all_geom.width() + 5,
                    view_all_geom.y(),
                    80,
                    view_all_geom.height()
                )

            # Connect the button to the filter method
            self.pushButton_filter_pottery.clicked.connect(self.on_pushButton_filter_pottery_pressed)

        except Exception as e:
            print(f"Error setting up filter button: {e}")

    def on_pushButton_filter_pottery_pressed(self):
        """Open the filter dialog for ID Number and Year selection"""
        self.empty_fields()
        # Create and show the dialog
        filter_dialog = PotteryFilterDialog(self.DB_MANAGER, self)
        result = filter_dialog.exec_()

        if result:
            # Get the selected ID numbers and year from the dialog
            selected_ids = filter_dialog.get_selected_ids()
            selected_year = filter_dialog.get_selected_year()

            if not selected_ids:
                if self.L == 'it':
                    QMessageBox.information(self, 'Info', "Nessun ID Number selezionato.", QMessageBox.Ok)
                else:
                    QMessageBox.information(self, 'Info', "No ID Number selected.", QMessageBox.Ok)
                return

            # Filter DATA_LIST based on selected IDs and optionally year
            if selected_year is not None:
                # Filter by both year and ID numbers
                filtered_data_list = [
                    record for record in self.DATA_LIST
                    if record.id_number in selected_ids and record.anno == selected_year
                ]
            else:
                # Filter by ID numbers only
                filtered_data_list = [
                    record for record in self.DATA_LIST
                    if record.id_number in selected_ids
                ]

            # Sort filtered list based on the order of selected IDs
            filtered_data_list = sorted(
                filtered_data_list,
                key=lambda record: selected_ids.index(record.id_number) if record.id_number in selected_ids else -1
            )

            # Update the UI with the filtered and sorted data
            if filtered_data_list:
                self.DATA_LIST = filtered_data_list
                self.REC_TOT, self.REC_CORR = len(self.DATA_LIST), 0
                self.set_rec_counter(len(self.DATA_LIST), self.REC_CORR + 1)
                self.DATA_LIST_REC_TEMP = self.DATA_LIST_REC_CORR = self.DATA_LIST[0]
                self.fill_fields()
                self.BROWSE_STATUS = "b"
                self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
            else:
                if self.L == 'it':
                    QMessageBox.information(self, 'Nessun Risultato', "Nessun record corrisponde ai filtri selezionati.", QMessageBox.Ok)
                else:
                    QMessageBox.information(self, 'No Results', "No records match the selected filters.", QMessageBox.Ok)

    def dropEvent(self, event):
        mimeData = event.mimeData()
        accepted_formats = ["jpg", "jpeg", "png", "tiff", "tif", "bmp", "mp4", "avi", "mov", "mkv", "flv", "obj", "stl",
                            "ply", "fbx", "3ds"]
        if mimeData.hasUrls():
            for url in mimeData.urls():
                try:
                    path = url.toLocalFile()
                    if os.path.isfile(path):
                        filename = os.path.basename(path)
                        filetype = filename.split(".")[-1]
                        if filetype.lower() in accepted_formats:
                            self.load_and_process_image(path)
                        else:
                            QMessageBox.warning(self, "Error", f"Unsupported file type: {filetype}", QMessageBox.Ok)
                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Failed to process the file: {str(e)}", QMessageBox.Ok)
        super().dropEvent(event)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        event.acceptProposedAction()

    def insert_record_media(self, mediatype, filename, filetype, filepath):
        self.mediatype = mediatype
        self.filename = filename
        self.filetype = filetype
        self.filepath = filepath
        try:
            data = self.DB_MANAGER.insert_media_values(
                self.DB_MANAGER.max_num_id('MEDIA', 'id_media') + 1,
                str(self.mediatype),  # 1 - mediatyype
                str(self.filename),  # 2 - filename
                str(self.filetype),  # 3 - filetype
                str(self.filepath),  # 4 - filepath
                str('Insert description'),  # 5 - descrizione
                str("['imagine']"))  # 6 - tags
            try:
                self.DB_MANAGER.insert_data_session(data)
                return 1
            except Exception as e:
                e_str = str(e)
                if e_str.__contains__("Integrity"):
                    msg = self.filename + ": Image already in the database"
                else:
                    msg = e
                # QMessageBox.warning(self, "Errore", "Warning 1 ! \n"+ str(msg),  QMessageBox.Ok)
                return 0
        except Exception as e:
            QMessageBox.warning(self, "Error", "Warning 2 ! \n" + str(e), QMessageBox.Ok)
            return 0

    def insert_record_mediathumb(self, media_max_num_id, mediatype, filename, filename_thumb, filetype, filepath_thumb,
                                 filepath_resize):
        self.media_max_num_id = media_max_num_id
        self.mediatype = mediatype
        self.filename = filename
        self.filename_thumb = filename_thumb
        self.filetype = filetype
        self.filepath_thumb = filepath_thumb
        self.filepath_resize = filepath_resize
        try:
            data = self.DB_MANAGER.insert_mediathumb_values(
                self.DB_MANAGER.max_num_id('MEDIA_THUMB', 'id_media_thumb') + 1,
                str(self.media_max_num_id),  # 1 - media_max_num_id
                str(self.mediatype),  # 2 - mediatype
                str(self.filename),  # 3 - filename
                str(self.filename_thumb),  # 4 - filename_thumb
                str(self.filetype),  # 5 - filetype
                str(self.filepath_thumb),  # 6 - filepath_thumb
                str(self.filepath_resize))  # 6 - filepath_thumb
            try:
                self.DB_MANAGER.insert_data_session(data)
                return 1
            except Exception as e:
                e_str = str(e)
                if e_str.__contains__("Integrity"):
                    msg = self.filename + ": thumb already present into the database"
                else:
                    msg = e
                # QMessageBox.warning(self, "Error", "warming 1 ! \n"+ str(msg),  QMessageBox.Ok)
                return 0
        except Exception as e:
            QMessageBox.warning(self, "Error", "Warning 2 ! \n" + str(e), QMessageBox.Ok)
            return 0

    def insert_mediaToEntity_rec(self, id_entity, entity_type, table_name, id_media, filepath, media_name):
        """
        id_mediaToEntity,
        id_entity,
        entity_type,
        table_name,
        id_media,
        filepath,
        media_name"""
        self.id_entity = id_entity
        self.entity_type = entity_type
        self.table_name = table_name
        self.id_media = id_media
        self.filepath = filepath
        self.media_name = media_name
        try:
            data = self.DB_MANAGER.insert_media2entity_values(
                self.DB_MANAGER.max_num_id('MEDIATOENTITY', 'id_mediaToEntity') + 1,
                int(self.id_entity),  # 1 - id_entity
                str(self.entity_type),  # 2 - entity_type
                str(self.table_name),  # 3 - table_name
                int(self.id_media),  # 4 - us
                str(self.filepath),  # 5 - filepath
                str(self.media_name))  # 6 - media_name
            try:
                self.DB_MANAGER.insert_data_session(data)
                return 1
            except Exception as e:
                e_str = str(e)
                if e_str.__contains__("Integrity"):
                    msg = self.ID_TABLE + " already present into the database"
                else:
                    msg = e
                QMessageBox.warning(self, "Error", "Warning 1 ! \n" + str(msg), QMessageBox.Ok)
                return 0
        except Exception as e:
            QMessageBox.warning(self, "Error", "Warning 2 ! \n" + str(e), QMessageBox.Ok)
            return 0

    def delete_mediaToEntity_rec(self, id_entity, entity_type, table_name, id_media, filepath, media_name):
        """
        id_mediaToEntity,
        id_entity,
        entity_type,
        table_name,
        id_media,
        filepath,
        media_name"""
        self.id_entity = id_entity
        self.entity_type = entity_type
        self.table_name = table_name
        self.id_media = id_media
        self.filepath = filepath
        self.media_name = media_name
        try:
            data = self.DB_MANAGER.insert_media2entity_values(
                self.DB_MANAGER.max_num_id('MEDIATOENTITY', 'id_mediaToEntity') + 1,
                int(self.id_entity),  # 1 - id_entity
                str(self.entity_type),  # 2 - entity_type
                str(self.table_name),  # 3 - table_name
                int(self.id_media),  # 4 - us
                str(self.filepath),  # 5 - filepath
                str(self.media_name))
        except Exception as e:
            QMessageBox.warning(self, "Error", "Warning 2 ! \n" + str(e), QMessageBox.Ok)
            return 0

    def generate_reperti(self):
        # tags_list = self.table2dict('self.tableWidgetTags_US')
        record_rep_list = []
        sito = self.comboBox_sito.currentText()
        area = self.comboBox_area.currentText()
        us = self.lineEdit_us.text()
        nv = self.lineEdit_id_number.text()
        # for sing_tags in tags_list:
        search_dict = {
            'id_number': "'" + str(nv) + "'",
            'sito': "'" + str(sito) + "'",
            'area': "'" + str(area) + "'",
            'us': "'" + str(us) + "'",
        }
        j = self.DB_MANAGER.query_bool(search_dict, 'POTTERY')
        record_rep_list.append(j)
        # QMessageBox.information(self, 'search db', str(record_us_list))
        rep_list = []
        for r in record_rep_list:
            rep_list.append([r[0].id_rep, 'CERAMICA', 'pottery_table'])
        return rep_list

    def assignTags_reperti(self, item):
        """
        id_mediaToEntity,
        id_entity,
        entity_type,
        table_name,
        id_media,
        filepath,
        media_name
        """
        rep_list = self.generate_reperti()
        # QMessageBox.information(self,'search db',str(us_list))
        if not rep_list:
            return

        for rep_data in rep_list:
            id_orig_item = item.text()  # return the name of original file
            search_dict = {'filename': "'" + str(id_orig_item) + "'"}
            media_data = self.DB_MANAGER.query_bool(search_dict, 'MEDIA')
            self.insert_mediaToEntity_rec(rep_data[0], rep_data[1], rep_data[2], media_data[0].id_media,
                                          media_data[0].filepath, media_data[0].filename)

            # Trigger automatic embedding index update
            if hasattr(self, 'embedding_updater') and self.embedding_updater is not None:
                try:
                    pottery_id = rep_data[0]  # id_rep
                    media_id = media_data[0].id_media
                    self.embedding_updater.on_image_added(pottery_id, media_id)
                except Exception as e:
                    print(f"[Pottery] Error triggering embedding update on image add: {e}")

    def load_and_process_image(self, filepath):
        media_resize_suffix = ''
        media_thumb_suffix = ''
        conn = Connection()
        thumb_path = conn.thumb_path()
        thumb_path_str = thumb_path['thumb_path']
        if thumb_path_str == '':
            if self.L == 'it':
                QMessageBox.information(self, "Info",
                                        "devi settare prima la path per salvare le thumbnail e i video. Vai in impostazioni di sistema/ path setting ")
            elif self.L == 'de':
                QMessageBox.information(self, "Info",
                                        "müssen Sie zuerst den Pfad zum Speichern der Miniaturansichten und Videos festlegen. Gehen Sie zu System-/Pfad-Einstellung")
            else:
                QMessageBox.information(self, "Message",
                                        "you must first set the path to save the thumbnails and videos. Go to system/path setting")
        else:
            filename = os.path.basename(filepath)
            filename, filetype = filename.split(".")[0], filename.split(".")[1]
            # Check the media type based on the file extension
            accepted_image_formats = ["jpg", "jpeg", "png", "tiff", "tif", "bmp"]
            accepted_video_formats = ["mp4", "avi", "mov", "mkv", "flv"]
            accepted_3d_formats = ["obj", "stl", "ply", "fbx", "3ds"]

            if filetype.lower() in accepted_image_formats:
                mediatype = 'image'
                media_thumb_suffix = '_thumb.png'
                media_resize_suffix = '.png'
            elif filetype.lower() in accepted_video_formats:
                mediatype = 'video'
                media_thumb_suffix = '_video.png'
                media_resize_suffix = '.' + filetype.lower()
            elif filetype.lower() in accepted_3d_formats:
                mediatype = '3d_model'
                media_thumb_suffix = '_3d_thumb.png'
                media_resize_suffix = '.' + filetype.lower()
            else:
                raise ValueError(f"Unrecognized media type for file {filename}.{filetype}")

            if mediatype == 'video':
                if filetype.lower() == 'mp4':
                    media_resize_suffix = '.mp4'
                elif filetype.lower() == 'avi':
                    media_resize_suffix = '.avi'
                elif filetype.lower() == 'mov':
                    media_resize_suffix = '.mov'
                elif filetype.lower() == 'mkv':
                    media_resize_suffix = '.mkv'
                elif filetype.lower() == 'flv':
                    media_resize_suffix = '.flv'

            elif mediatype == '3d_model':
                if filetype.lower() == 'obj':
                    media_resize_suffix = '.obj'
                elif filetype.lower() == 'ply':
                    media_resize_suffix = '.ply'
                elif filetype.lower() == 'fbx':
                    media_resize_suffix = '.fbx'
                elif filetype.lower() == '3ds':
                    media_resize_suffix = '.3ds'
                elif filetype.lower() == 'stl':
                    media_resize_suffix = '.stl'
            # Check and insert record in the database
            idunique_image_check = self.db_search_check('MEDIA', 'filepath', filepath)

            try:
                if bool(idunique_image_check):

                    return
                else:
                    # mediatype = 'image'
                    self.insert_record_media(mediatype, filename, filetype, filepath)
                    MU = Media_utility()
                    MUR = Media_utility_resize()
                    MU_video = Video_utility()
                    MUR_video = Video_utility_resize()
                    media_max_num_id = self.DB_MANAGER.max_num_id('MEDIA', 'id_media')
                    thumb_path = conn.thumb_path()
                    thumb_path_str = thumb_path['thumb_path']
                    thumb_resize = conn.thumb_resize()
                    thumb_resize_str = thumb_resize['thumb_resize']
                    filenameorig = filename
                    filename_thumb = str(media_max_num_id) + "_" + filename + media_thumb_suffix
                    filename_resize = str(media_max_num_id) + "_" + filename + media_resize_suffix
                    filepath_thumb = filename_thumb
                    filepath_resize = filename_resize
                    self.SORT_ITEMS_CONVERTED = []

                    try:

                        if mediatype == '3d_model':
                            self.process_3d_model(media_max_num_id, filepath, filename, thumb_path_str,
                                                  thumb_resize_str,
                                                  media_thumb_suffix, media_resize_suffix)

                        elif mediatype == 'video':
                            vcap = cv2.VideoCapture(filepath)
                            res, im_ar = vcap.read()
                            while im_ar.mean() < 1 and res:
                                res, im_ar = vcap.read()
                            im_ar = cv2.resize(im_ar, (100, 100), 0, 0, cv2.INTER_LINEAR)
                            # to save we have two options
                            outputfile = '{}.png'.format(os.path.dirname(filepath) + '/' + filename)
                            cv2.imwrite(outputfile, im_ar)
                            MU_video.resample_images(media_max_num_id, outputfile, filenameorig, thumb_path_str,
                                                     media_thumb_suffix)
                            MUR_video.resample_images(media_max_num_id, filepath, filenameorig, thumb_resize_str,
                                                      media_resize_suffix)
                        elif mediatype == 'image':
                            MU.resample_images(media_max_num_id, filepath, filenameorig, thumb_path_str,
                                               media_thumb_suffix)
                            MUR.resample_images(media_max_num_id, filepath, filenameorig, thumb_resize_str,
                                                media_resize_suffix)
                    except Exception as e:
                        QMessageBox.warning(self, "Cucu", str(e), QMessageBox.Ok)
                    self.insert_record_mediathumb(media_max_num_id, mediatype, filename, filename_thumb, filetype,
                                                  filepath_thumb, filepath_resize)

                    item = QListWidgetItem(str(filenameorig))
                    item.setData(Qt.UserRole, str(media_max_num_id))
                    icon = load_icon(get_image_path(str(thumb_path_str), filepath_thumb))
                    item.setIcon(icon)
                    self.iconListWidget.addItem(item)

                self.assignTags_reperti(item)




            except AssertionError as e:

                if self.L == 'it':
                    QMessageBox.warning(self, "Warning", "controlla che il nome del file non abbia caratteri speciali",

                                        QMessageBox.Ok)

                if self.L == 'de':

                    QMessageBox.warning(self, "Warning", "prüfen, ob der Dateiname keine Sonderzeichen enthält",
                                        QMessageBox.Ok)

                else:

                    QMessageBox.warning(self, "Warning", str(e), QMessageBox.Ok)

    def db_search_check(self, table_class, field, value):
        self.table_class = table_class
        self.field = field
        self.value = value
        search_dict = {self.field: "'" + str(self.value) + "'"}
        u = Utility()
        search_dict = u.remove_empty_items_fr_dict(search_dict)
        res = self.DB_MANAGER.query_bool(search_dict, self.table_class)
        return res

    def _remove_image_tag_and_embedding(self, pottery_id, filename):
        """
        Helper method to remove image tag and trigger embedding removal.

        Args:
            pottery_id: The pottery record ID (id_rep)
            filename: The original filename of the image
        """
        # Get media_id before removing the tag
        media_id = None
        try:
            search_dict = {'filename': "'" + str(filename) + "'"}
            media_data = self.DB_MANAGER.query_bool(search_dict, 'MEDIA')
            if media_data:
                media_id = media_data[0].id_media
        except Exception as e:
            print(f"[Pottery] Error getting media_id for embedding removal: {e}")

        # Remove the tag from database
        self.DB_MANAGER.remove_tags_from_db_sql_scheda(pottery_id, filename)

        # Trigger embedding removal
        if media_id is not None and hasattr(self, 'embedding_updater') and self.embedding_updater is not None:
            try:
                self.embedding_updater.on_image_removed(pottery_id, media_id)
            except Exception as e:
                print(f"[Pottery] Error triggering embedding removal: {e}")

    def on_pushButton_removetags_pressed(self):
        def r_id():
            record_rep_list = []
            sito = self.comboBox_sito.currentText()
            area = self.comboBox_area.currentText()
            us = self.lineEdit_us.text()
            nv = self.lineEdit_id_number.text()
            # for sing_tags in tags_list:
            search_dict = {
                'id_number': "'" + str(nv) + "'",
                'sito': "'" + str(sito) + "'",
                'area': "'" + str(area) + "'",
                'us': "'" + str(us) + "'",
            }
            j = self.DB_MANAGER.query_bool(search_dict, 'POTTERY')
            record_rep_list.append(j)
            # QMessageBox.information(self, 'search db', str(record_us_list))
            a=None
            for r in record_rep_list:
                a = r[0].id_rep
            # QMessageBox.information(self,'ok',str(a))# QMessageBox.information(self, "Scheda US", str(us_list), QMessageBox.Ok)
            return a

        items_selected = self.iconListWidget.selectedItems()
        if not bool(items_selected):
            if self.L == 'it':

                msg = QMessageBox.warning(self, "Attenzione!!!",
                                          "devi selezionare prima l'immagine",
                                          QMessageBox.Ok)

            elif self.L == 'de':

                msg = QMessageBox.warning(self, "Warnung",
                                          "moet je eerst de afbeelding selecteren",
                                          QMessageBox.Ok)
            else:

                msg = QMessageBox.warning(self, "Warning",
                                          "you must first select an image",
                                          QMessageBox.Ok)
        else:
            if self.L == 'it':
                msg = QMessageBox.warning(self, "Warning",
                                          "Vuoi veramente cancellare i tags dalle thumbnail selezionate? \n L'azione è irreversibile",
                                          QMessageBox.Ok | QMessageBox.Cancel)
                if msg == QMessageBox.Cancel:
                    QMessageBox.warning(self, "Messaggio!!!", "Azione Annullata!")
                else:
                    # items_selected = self.iconListWidget.selectedItems()
                    pottery_id = r_id()
                    for item in items_selected:
                        id_orig_item = item.text()  # return the name of original file

                        # s = self.iconListWidget.item(0, 0).text()
                        self._remove_image_tag_and_embedding(pottery_id, id_orig_item)
                        row = self.iconListWidget.row(item)
                        self.iconListWidget.takeItem(row)
                    QMessageBox.warning(self, "Info", "Tags rimossi!")
            elif self.L == 'de':
                msg = QMessageBox.warning(self, "Warning",
                                          "Wollen Sie wirklich die Tags aus den ausgewählten Miniaturbildern löschen? \n Die Aktion ist unumkehrbar",
                                          QMessageBox.Ok | QMessageBox.Cancel)
                if msg == QMessageBox.Cancel:
                    QMessageBox.warning(self, "Warnung", "Azione Annullata!")
                else:
                    # items_selected = self.iconListWidget.selectedItems()
                    pottery_id = r_id()
                    for item in items_selected:
                        id_orig_item = item.text()  # return the name of original file

                        # s = self.iconListWidget.item(0, 0).text()
                        self._remove_image_tag_and_embedding(pottery_id, id_orig_item)
                        row = self.iconListWidget.row(item)
                        self.iconListWidget.takeItem(row)
                    QMessageBox.warning(self, "Info", "Tags entfernt")

            else:
                msg = QMessageBox.warning(self, "Warning",
                                          "Do you really want to delete the tags from the selected thumbnails? \n The action is irreversible",
                                          QMessageBox.Ok | QMessageBox.Cancel)
                if msg == QMessageBox.Cancel:
                    QMessageBox.warning(self, "Warning", "Action cancelled")
                else:
                    # items_selected = self.iconListWidget.selectedItems()
                    pottery_id = r_id()
                    for item in items_selected:
                        id_orig_item = item.text()  # return the name of original file

                        # s = self.iconListWidget.item(0, 0).text()
                        self._remove_image_tag_and_embedding(pottery_id, id_orig_item)
                        row = self.iconListWidget.row(item)
                        self.iconListWidget.takeItem(row)  # remove the item from the list

                    QMessageBox.warning(self, "Info", "Tags removed")

    def on_pushButton_all_images_pressed(self):
        record_us_list = self.DB_MANAGER.query('MEDIA_THUMB')

        et = {'entity_type': "'CERAMICA'"}
        ser = self.DB_MANAGER.query_bool(et, 'MEDIATOENTITY')
        # Verifica se record_us_list è vuota
        if not record_us_list and not ser:
            QMessageBox.information(self, "Informazione", "Non ci sono immagini da mostrare.")
            return  # Termina la funzione

        # Inizializza la QListWidget fuori dal ciclo
        self.new_list_widget = QListWidget()
        # ##self.new_list_widget.setFixedSize(200, 300)
        self.new_list_widget.setSelectionMode(QAbstractItemView.SingleSelection)  # Permette selezioni multiple

        done_button = QPushButton("TAG")

        def update_done_button():
            if not self.new_list_widget.selectedItems():
                done_button.setHidden(True)
            else:
                done_button.setHidden(False)
                done_button.clicked.connect(self.on_done_selecting_all)

        self.new_list_widget.itemSelectionChanged.connect(
            update_done_button)  # Aggiungi un layout per le etichette dei numeri delle pagine
        self.pageLayout = QHBoxLayout()
        self.current_page_label = QLabel()  # Creiamo l'etichetta per la pagina corrente
        self.total_pages_label = QLabel()  # Creiamo l'etichetta per il totale delle pagine

        self.pageLayout.addWidget(self.current_page_label)  # Aggiungiamo l'etichetta della pagina corrente al layout
        self.pageLayout.addWidget(self.total_pages_label)  # Aggiungiamo l'etichetta del totale delle pagine al layout

        # Aggiungi un pulsante "Indietro"
        self.prevButton = QPushButton("<<")
        self.prevButton.clicked.connect(self.go_to_previous_page)
        self.pageLayout.addWidget(self.prevButton)

        # Aggiungi le etichette dei numeri delle pagine
        self.pageLabels = []
        for i in range(1, 6):
            label = QLabel(str(i))
            label.setAlignment(Qt.AlignCenter)
            label.setMinimumWidth(30)
            label.setFrameStyle(QFrame.Panel | QFrame.Sunken)
            label.setMargin(2)
            label.mousePressEvent = functools.partial(self.on_page_label_clicked, i)
            self.pageLabels.append(label)
            self.pageLayout.addWidget(label)

        # Aggiungi un pulsante "Avanti"
        self.nextButton = QPushButton(">>")
        self.nextButton.clicked.connect(self.go_to_next_page)
        self.pageLayout.addWidget(self.nextButton)

        layout = QVBoxLayout()
        # Crea un campo di input per la ricerca
        self.search_field = QLineEdit()
        self.search_field.setPlaceholderText("Cerca...poi schiaccia invio")
        self.current_filter_text = ""

        self.page_size = 10  # Numero di immagini per pagina
        self.current_page = 1  # Pagina corrente
        self.total_pages = 0  # Numero totale di pagine

        # Aggiungi il campo di ricerca al layout sopra la QListWidget
        layout.insertWidget(0, self.search_field)

        layout.addLayout(self.pageLayout)
        layout.addWidget(self.new_list_widget)
        layout.addWidget(done_button)

        # Imposta il fattore di estensione per i widget nel layout
        # Il primo parametro è l'indice del widget e il secondo parametro è il fattore di estensione
        # In questo caso, new_list_widget ha un indice di 0 e done_button ha un indice di 1
        layout.setStretchFactor(self.new_list_widget, 5)  # new_list_widget avrà 3 volte più spazio di done_button
        layout.setStretchFactor(done_button, 1)  # done_button avrà 1/3 dello spazio di new_list_widget

        # Imposta il layout sulla tua finestra o su un altro widget
        self.setLayout(layout)

        # Crea un nuovo widget per contenere la QListWidget e il pulsante, e applica il layout
        self.widget = QWidget()
        self.widget.setLayout(layout)
        self.widget.adjustSize()
        self.widget.show()

        self.load_images()

        # Connette il campo di ricerca a una funzione di filtraggio
        self.search_field.returnPressed.connect(self.filter_items)

    def load_images(self, filter_text=None):
        conn = Connection()
        thumb_path = conn.thumb_path()
        thumb_path_str = thumb_path['thumb_path']
        u = Utility()

        # Calcola l'offset per la pagina corrente
        # offset = (self.current_page - 1) * self.page_size

        # Ottieni tutti i record delle immagini
        all_images = self.DB_MANAGER.query('MEDIA_THUMB')

        # Ottieni tutte le immagini taggate
        tagged_images = self.DB_MANAGER.query('MEDIATOENTITY')

        # Ottieni gli id_media di tutte le immagini taggate
        tagged_ids = [i.id_media for i in tagged_images]

        # Filtra tutte le immagini per ottenere solo quelle non taggate
        untagged_images = [i for i in all_images if i.id_media not in tagged_ids]

        # Inizializza l'elenco delle immagini 'US' come un duplicato delle immagini non taggate
        us_images = untagged_images[:]

        if len(all_images) > 100:

            if filter_text:  # se il filtro è attivo
                filtered_images = [i for i in untagged_images if filter_text.lower() in i.media_filename.lower()]
            else:
                filtered_images = us_images
            # Calcola gli indici di inizio e fine per la pagina corrente
            start_index = (self.current_page - 1) * self.page_size
            end_index = start_index + self.page_size

            # Ottieni i record delle immagini per la pagina corrente
            self.record_us_list = filtered_images[start_index:end_index]
            # Pulisci la QListWidget prima di aggiungere le nuove immagini
            self.new_list_widget.clear()
            # Aggiungi l'intestazione alla QListWidget
            header_item = QListWidgetItem(
                "Le righe selezionate in giallo indicano immagini non taggate\n Da questo strumento solo le righe selezionate gialle posso essere taggate ")
            header_item.setBackground(QColor('lightgrey'))
            header_item.setFlags(header_item.flags() & ~Qt.ItemIsSelectable)  # rendi l'item non selezionabile
            self.new_list_widget.addItem(header_item)
            # Aggiungi le immagini alla QListWidget

            for i in self.record_us_list:
                search_dict = {'id_media': "'" + str(i.id_media) + "'"}
                u = Utility()
                search_dict = u.remove_empty_items_fr_dict(search_dict)
                mediathumb_data = self.DB_MANAGER.query_bool(search_dict, "MEDIA_THUMB")
                thumb_path = str(mediathumb_data[0].filepath)
                # Verifica se l'immagine è già in cache
                if thumb_path not in self.image_cache:
                    # Se non è in cache, carica l'immagine
                    icon = load_icon(get_image_path(thumb_path_str, thumb_path))

                    # Se la cache ha raggiunto il limite, rimuove l'elemento più vecchio
                    if len(self.image_cache) >= self.cache_limit:
                        self.image_cache.popitem(last=False)

                    # Aggiunge l'immagine alla cache
                    self.image_cache[thumb_path] = icon
                else:

                    icon = self.image_cache[thumb_path]

                self.image_cache.move_to_end(thumb_path)

                item = QListWidgetItem(str(i.media_filename))
                item.setData(Qt.UserRole, str(i.media_filename))
                icon = load_icon(get_image_path(thumb_path_str, thumb_path))
                item.setIcon(icon)

                item.setBackground(QColor("yellow"))

                self.new_list_widget.addItem(item)
        else:
            for image in all_images:
                # Crea un nuovo dizionario di ricerca per MEDIATOENTITY
                search_dict = {'id_media': "'" + str(image.id_media) + "'",
                               'entity_type': "'CERAMICA'"}
                search_dict = u.remove_empty_items_fr_dict(search_dict)

                # Recupera l'elenco di 'US' associati all'immagine
                mediatoentity_data = self.DB_MANAGER.query_bool(search_dict, "MEDIATOENTITY")

                # Se l'immagine ha una o più 'US' associate, aggiungila all'elenco
                if mediatoentity_data:
                    us_images.append(image)

            # A questo punto, 'us_images' dovrebbe contenere tutte le immagini non taggate
            # e quelle taggate con almeno un 'US', senza duplicati.

            if filter_text:  # se il filtro è attivo
                filtered_images = [i for i in us_images if filter_text.lower() in i.media_filename.lower()]
            else:
                filtered_images = us_images

            # Calcola gli indici di inizio e fine per la pagina corrente
            start_index = (self.current_page - 1) * self.page_size
            end_index = start_index + self.page_size

            # Ottieni i record delle immagini per la pagina corrente
            self.record_us_list = filtered_images[start_index:end_index]

            # Pulisci la QListWidget prima di aggiungere le nuove immagini
            self.new_list_widget.clear()

            # Aggiungi l'intestazione alla QListWidget
            header_item = QListWidgetItem(
                "Le righe selezionate in giallo indicano immagini non taggate\n Da questo strumento solo le righe selezionate gialle posso essere taggate ")
            header_item.setBackground(QColor('lightgrey'))
            header_item.setFlags(header_item.flags() & ~Qt.ItemIsSelectable)  # rendi l'item non selezionabile
            self.new_list_widget.addItem(header_item)
            # Aggiungi le immagini alla QListWidget

            for i in self.record_us_list:
                search_dict = {'id_media': "'" + str(i.id_media) + "'"}
                u = Utility()
                search_dict = u.remove_empty_items_fr_dict(search_dict)
                mediathumb_data = self.DB_MANAGER.query_bool(search_dict, "MEDIA_THUMB")
                thumb_path = str(mediathumb_data[0].filepath)
                # Verifica se l'immagine è già in cache
                if thumb_path not in self.image_cache:
                    # Se non è in cache, carica l'immagine
                    icon = load_icon(get_image_path(thumb_path_str, thumb_path))

                    # Se la cache ha raggiunto il limite, rimuove l'elemento più vecchio
                    if len(self.image_cache) >= self.cache_limit:
                        self.image_cache.popitem(last=False)

                    # Aggiunge l'immagine alla cache
                    self.image_cache[thumb_path] = icon
                else:
                    # Se è in cache, utilizza l'icona dalla cache
                    icon = self.image_cache[thumb_path]

                    # Aggiorna l'ordine della cache spostando l'elemento utilizzato alla fine
                    self.image_cache.move_to_end(thumb_path)
                # Crea un nuovo dizionario di ricerca per MEDIATOENTITY
                search_dict = {'id_media': "'" + str(i.id_media) + "'",
                               'entity_type': "'CERAMICA'"}
                search_dict = u.remove_empty_items_fr_dict(search_dict)
                # Recupera l'elenco di US associati all'immagine
                mediatoentity_data = self.DB_MANAGER.query_bool(search_dict, "MEDIATOENTITY")
                us_list = [str(g.id_entity) for g in
                           mediatoentity_data]  # Se 'entity_type' è 'US', aggiungi l'id_media a us_images
                # Rimuovi i duplicati dalla lista convertendola in un set e poi di nuovo in una lista
                # us_list = list(set(us_list))

                if us_list:
                    # us_list = [g.id_entity for g in mediatoentity_data if 'US' in g.entity_type]
                    item = QListWidgetItem(str(i.media_filename))
                    item.setData(Qt.UserRole, str(i.media_filename))
                    icon = load_icon(get_image_path(thumb_path_str, thumb_path))
                    item.setIcon(icon)

                    item.setBackground(QColor("white"))

                    # Inizializza una lista vuota per i nomi delle US
                    us_names = []

                    for us_id in us_list:
                        # Crea un nuovo dizionario di ricerca per l'US
                        search_dict_us = {'id_rep': us_id}
                        search_dict_us = u.remove_empty_items_fr_dict(search_dict_us)

                        # Query the US table
                        us_data = self.DB_MANAGER.query_bool(search_dict_us, "POTTERY")

                        # Se l'US esiste, aggiungi il suo nome alla lista
                        if us_data:
                            us_names.extend([str(us.id_number) for us in us_data])

                    # Se ci sono dei nomi US, aggiungi questi all'elemento
                    if us_names:
                        item.setText(item.text() + " - POTTERY: " + ', '.join(us_names))
                    else:
                        pass  # oppure: item.setText(item.text() + " - US: Non trovato")
                # item.setText(item.text() + " - US: Non trovato")

                else:
                    # us_list = [g.id_entity for g in mediatoentity_data if 'US' in g.entity_type]
                    item = QListWidgetItem(str(i.media_filename))
                    item.setData(Qt.UserRole, str(i.media_filename))
                    icon = load_icon(get_image_path(thumb_path_str, thumb_path))
                    item.setIcon(icon)

                    item.setBackground(QColor("yellow"))

                    # Aggiungi l'elemento alla QListWidget
                # self.new_list_widget.clear()
                self.new_list_widget.addItem(item)

        # Calcola il numero totale di pagine

        self.total_pages = math.ceil(len(filtered_images) / self.page_size)

        # Aggiorna l'aspetto delle etichette dei numeri delle pagine
        self.update_page_labels()

    def update_page_labels(self):
        # Disabilita il pulsante "Indietro" se siamo alla prima pagina
        self.prevButton.setEnabled(self.current_page > 1)

        # Disabilita il pulsante "Avanti" se siamo all'ultima pagina
        self.nextButton.setEnabled(self.current_page < self.total_pages)

        # Aggiorna l'aspetto delle etichette dei numeri delle pagine
        for label in self.pageLabels:
            page_number = int(label.text())
            label.setEnabled(page_number != self.current_page)

        # Aggiorna l'etichetta della pagina corrente e del totale delle pagine
        self.current_page_label.setText(f"Pagina corrente: {self.current_page}")
        self.total_pages_label.setText(f"Totale pagine: {self.total_pages}")

    def go_to_previous_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.load_images(self.current_filter_text)

    def go_to_next_page(self):
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.load_images(self.current_filter_text)

    def on_page_label_clicked(self, page, _=None):
        if page != self.current_page:
            self.current_page = page
            self.load_images(self.current_filter_text)

    def filter_items(self):
        # Ottieni il testo corrente nel campo di ricerca
        self.current_filter_text = self.search_field.text().lower()
        self.load_images(self.current_filter_text)

    def on_done_selecting_all(self):

        global item

        def r_list():
            result_list = []

            sito = self.comboBox_sito.currentText()
            area = self.comboBox_area.currentText()
            us = self.lineEdit_us.text()
            nv = self.lineEdit_id_number.text()
            # for sing_tags in tags_list:
            search_dict = {
                'id_number': "'" + str(nv) + "'",
                'sito': "'" + str(sito) + "'",
                'area': "'" + str(area) + "'",
                'us': "'" + str(us) + "'",
            }
            pottery_records = self.DB_MANAGER.query_bool(search_dict, 'POTTERY')

            for pottery in pottery_records:
                result_list.append([pottery.id_rep, 'CERAMICA', 'pottery_table'])
            # QMessageBox.information(self, "Scheda US", str(result_list), QMessageBox.Ok)
            return result_list

        items_selected = self.new_list_widget.selectedItems()
        for item in items_selected:
            for us_data in r_list():
                id_orig_item = item.text()  # return the name of original file
                search_dict = {'filename': "'" + str(id_orig_item) + "'"}
                media_data = self.DB_MANAGER.query_bool(search_dict, 'MEDIA')

                # Check if media_data is not empty
                if media_data:
                    # Check if this image is already in the database
                    search_dict = {'id_media': "'" + str(media_data[0].id_media) + "'"}
                    existing_entry = self.DB_MANAGER.query_bool(search_dict, 'MEDIATOENTITY')

                    # If this image is already in the database, continue with the next item
                    if existing_entry:
                        continue

                    self.insert_mediaToEntity_rec(us_data[0], us_data[1], us_data[2], media_data[0].id_media,
                                                  media_data[0].filepath, media_data[0].filename)
                else:
                    pass
                    #QMessageBox.warning(self, "Attenzione",
                                        #"Immagine già taggata: " + str(id_orig_item))
                    # After tagging the image, update the corresponding QListWidgetItem

        # After tagging, update the iconListWidget
        self.fill_iconListWidget()
        self.update_list_widget_item(item)

    def update_list_widget_item(self,item):
        #items_selected = self.new_list_widg)et.selectedItems(
        search_dict = {'media_name': "'" + str(item.text()) + "'"}
        u = Utility()
        search_dict = u.remove_empty_items_fr_dict(search_dict)
        mediatoentity_data = self.DB_MANAGER.query_bool(search_dict, "MEDIATOENTITY")

        # Update the QListWidgetItem based on whether it matches
        if mediatoentity_data:
            item.setBackground(QColor("white"))

            # Create a new search dictionary for the US
            search_dict_us = {'id_rep': "'" + str(mediatoentity_data[0].id_entity) + "'"}
            search_dict_us = u.remove_empty_items_fr_dict(search_dict_us)

            # Query the US table
            us_data = self.DB_MANAGER.query_bool(search_dict_us, "POTTERY")

            # If the US exists, add its name to the item
            if us_data:
                item.setText(item.text() + " - Pottery: " + str(us_data[0].id_number))
            else:
                item.setText(item.text() + " - Pottery: Not found")

        else:
            item.setBackground(QColor("yellow"))

    def fill_iconListWidget(self):
        #self.iconListWidget.clear()  # pulisci prima il widget
        items_selected = self.new_list_widget.selectedItems()
        for item in items_selected:
            item.text()
        # Prendi i dati dal tuo database o dalla tua fonte dati
        #data = self.DB_MANAGER.query('MEDIA_THUMB')
        search_dict = {'media_filename': "'" + str(item.text()) + "'"}
        u = Utility()
        search_dict = u.remove_empty_items_fr_dict(search_dict)
        data = self.DB_MANAGER.query_bool(search_dict, "MEDIA_THUMB")
        #QMessageBox.information(self, 'ok',str(item.text()))
        conn = Connection()

        thumb_path = conn.thumb_path()
        thumb_path_str = thumb_path['thumb_path']
        # crea un nuovo QListWidgetItem
        if data:
            list_item = QListWidgetItem(data[0].media_filename)  # utilizza il nome del file come testo dell'elemento
            list_item.setData(Qt.UserRole,data[0].media_filename)  # utilizza il nome del file come dati personalizzati dell'elemento

            # crea una QIcon con l'immagine
            icon = load_icon(get_image_path(thumb_path_str, data[0].filepath))  # utilizza il percorso del file per creare l'icona
            #QMessageBox.information(self,'ok',str(thumb_path_str + data[0].filepath))
            # imposta l'icona dell'elemento
            list_item.setIcon(icon)

            # aggiungi l'elemento al QListWidget
            self.iconListWidget.addItem(list_item)
    def loadMediaPreview(self):
        self.iconListWidget.clear()
        conn = Connection()
        thumb_path = conn.thumb_path()
        thumb_path_str = thumb_path['thumb_path']
        # if mode == 0:
        # """ if has geometry column load to map canvas """
        rec_list = self.ID_TABLE + " = " + str(
            eval("self.DATA_LIST[int(self.REC_CORR)]." + self.ID_TABLE))
        search_dict = {
            'id_entity': "'" + str(eval("self.DATA_LIST[int(self.REC_CORR)]." + self.ID_TABLE)) + "'",
            'entity_type': "'CERAMICA'"}
        record_us_list = self.DB_MANAGER.query_bool(search_dict, 'MEDIATOENTITY')
        for i in record_us_list:
            search_dict = {'id_media': "'" + str(i.id_media) + "'"}
            u = Utility()
            search_dict = u.remove_empty_items_fr_dict(search_dict)
            mediathumb_data = self.DB_MANAGER.query_bool(search_dict, "MEDIA_THUMB")
            thumb_path = str(mediathumb_data[0].filepath)
            item = QListWidgetItem(str(i.media_name))
            item.setData(Qt.UserRole, str(i.media_name))
            icon = load_icon(get_image_path(thumb_path_str, thumb_path))
            item.setIcon(icon)
            self.iconListWidget.addItem(item)
        # elif mode == 1:
        # self.iconListWidget.clear()

    def load_and_process_3d_model(self, filepath):
        filename = os.path.basename(filepath)
        filename, filetype = filename.split(".")[0], filename.split(".")[1]
        mediatype = '3d_model'

        # Inserisci il record nel database
        self.insert_record_media(mediatype, filename, filetype, filepath)

        # Genera una thumbnail del modello 3D
        thumbnail_path = self.generate_3d_thumbnail(filepath)

        # Inserisci il record della thumbnail
        media_max_num_id = self.DB_MANAGER.max_num_id('MEDIA', 'id_media')
        self.insert_record_mediathumb(media_max_num_id, mediatype, filename, f"{filename}_thumb.png", 'png',
                                      thumbnail_path, filepath)

        # Aggiungi l'item alla lista
        item = QListWidgetItem(str(filename))
        item.setData(Qt.UserRole, str(media_max_num_id))
        icon = load_icon(thumbnail_path)
        item.setIcon(icon)
        self.iconListWidget.addItem(item)

        self.assignTags_reperti(item)

    def show_3d_model(self, file_path):
        mesh = pv.read(file_path)
        points = []
        measuring = False
        measurement_objects = []

        main_widget = QWidget()
        main_layout = QHBoxLayout()
        main_widget.setLayout(main_layout)

        frame = QFrame()
        layout = QVBoxLayout()

        plotter = QtInteractor(frame)

        debug_widget = QTextEdit()
        debug_widget.setReadOnly(True)

        def add_debug_message(message, important=False):
            timestamp = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
            formatted_message = f"[{timestamp}] {message}"
            if important:
                formatted_message = f"<b>{formatted_message}</b>"
            debug_widget.append(formatted_message)
            debug_widget.ensureCursorVisible()

            max_messages = 1000
            if debug_widget.document().blockCount() > max_messages:
                cursor = debug_widget.textCursor()
                cursor.movePosition(cursor.Start)
                cursor.select(cursor.LineUnderCursor)
                cursor.removeSelectedText()
                cursor.deletePreviousChar()
                cursor.movePosition(cursor.End)
                debug_widget.setTextCursor(cursor)

        def mouse_click_callback(obj, event):
            nonlocal measuring, points
            if event == "LeftButtonPressEvent" and measuring:
                x, y = plotter.interactor.GetEventPosition()
                add_debug_message(f"Evento di clic a posizione schermo: (x: {x}, y: {y})")

                picker = vtk.vtkCellPicker()
                picker.SetTolerance(10)  # Aumenta la tolleranza per migliorare il picking
                picker.Pick(x, y, 0, plotter.renderer)

                if picker.GetCellId() != -1:
                    point = np.array(picker.GetPickPosition())
                    add_debug_message(f"Punto selezionato nello spazio del modello: {point}")

                    closest_point_id = mesh.find_closest_point(point)
                    closest_point = mesh.points[closest_point_id]
                    add_debug_message(f"Punto più vicino sulla superficie della mesh: {closest_point}")

                    on_left_click(closest_point)
                else:
                    add_debug_message("Nessun punto sulla superficie trovato", important=True)

        plotter.interactor.AddObserver(vtk.vtkCommand.LeftButtonPressEvent, mouse_click_callback)

        layout.addWidget(plotter.interactor)
        plotter.clear()

        texture_file = os.path.splitext(file_path)[0] + '.jpg'
        if os.path.exists(texture_file):
            texture = pv.read_texture(texture_file)
            plotter.add_mesh(mesh, texture=texture, show_edges=False)
        else:
            plotter.add_mesh(mesh, show_edges=False)

        instructions_widget = QTextEdit()
        instructions_widget.setReadOnly(True)
        instructions_widget.hide()

        instructions = (
            "Trackball Controls:\n"
            "- Rotate: Left-click and drag\n"
            "- Pan: Right-click and drag\n"
            "- Zoom: Mouse wheel or middle-click and drag\n"
            "- Reset view: 'r'\n"
            "- Start/Stop measuring: 'o'\n"
            "- Show bounding box measures: 'm'\n"
            "- Export image: 'e'\n"
            "- Clear measurements: 'c'\n"
            "\nMain Views:\n"
            "- XY View (top): 'z'\n"
            "- YZ View (front): 'x'\n"
            "- XZ View (side): 'y'\n"
            "- ZY View (back): 'w'\n"
            "- ZX View (opposite side): 'v'\n"
            "- YX View (bottom): 'b'"
        )
        instructions_widget.setText(instructions)

        def toggle_instructions():
            if instructions_widget.isHidden():
                instructions_widget.show()
            else:
                instructions_widget.hide()

        instructions_button = QPushButton("Toggle Instructions")
        instructions_button.clicked.connect(toggle_instructions)
        layout.addWidget(instructions_button)

        frame.setLayout(layout)
        main_layout.addWidget(frame)
        main_layout.addWidget(instructions_widget)

        # main_layout.addWidget(debug_widget)

        def toggle_measure():
            nonlocal measuring, points
            measuring = not measuring
            points.clear()
            if measuring:
                add_debug_message("Misurazione iniziata", important=True)
            else:
                add_debug_message("Misurazione terminata", important=True)

        def on_left_click(picked_point):
            nonlocal points
            if not measuring:
                return

            add_debug_message(f"Punto selezionato: {picked_point}")

            if picked_point is not None:
                points.append(picked_point)
                sphere = pv.Sphere(radius=mesh.length * 0.005,
                                   center=picked_point)  # Aumenta leggermente il raggio della sfera per una miglior visibilità
                sphere_actor = plotter.add_mesh(sphere, color='red')
                measurement_objects.append(sphere_actor)

                add_debug_message(f"Punto aggiunto. Totale punti: {len(points)}")
                if len(points) == 2:
                    add_debug_message("Due punti raccolti. Misurazione in corso...", important=True)
                    measure_distance(points[0], points[1])
                    points.clear()
            else:
                add_debug_message("Nessun punto selezionato", important=True)

        def verify_coordinates(coord1, coord2):
            add_debug_message(f"Verifica delle coordinate:\nPunto1: {coord1}\nPunto2: {coord2}", important=True)

        def measure_distance(point1, point2):
            add_debug_message(f"Misurazione della distanza tra {point1} e {point2}")
            distance = np.linalg.norm(np.array(point1) - np.array(point2))

            line = pv.Line(point1, point2)
            line_actor = plotter.add_mesh(line, color='red', line_width=2)
            measurement_objects.append(line_actor)
            add_debug_message("Linea aggiunta")

            labels = plotter.add_point_labels([point1, point2], ["P1", "P2"], point_size=1, font_size=6)
            measurement_objects.append(labels)
            add_debug_message("Etichette dei punti aggiunte")

            mid_point = (np.array(point1) + np.array(point2)) / 2
            distance_label = plotter.add_point_labels([mid_point], [f"{distance:.2f} cm"], point_size=0, font_size=6)
            measurement_objects.append(distance_label)
            add_debug_message("Etichetta della distanza aggiunta")

            verify_coordinates(point1, mid_point)  # Verifica le coordinate durante la misura

            plotter.render()
            add_debug_message(f"Distanza misurata: {distance:.2f} cm", important=True)

        def clear_measurements():
            nonlocal measurement_objects, points
            for obj in measurement_objects:
                plotter.remove_actor(obj)
            measurement_objects.clear()
            points.clear()
            plotter.render()

        def export_image():
            try:
                options = QFileDialog.Options()
                file_path, _ = QFileDialog.getSaveFileName(self, "Save Image", "",
                                                           "PNG Files (*.png);;All Files (*)", options=options)
                if file_path:
                    camera = plotter.camera_position
                    width_cm, height_cm = 15, 10
                    width_inches, height_inches = width_cm / 2.54, height_cm / 2.54
                    dpi = 300
                    width_pixels, height_pixels = int(width_inches * dpi), int(height_inches * dpi)
                    plotter.screenshot(file_path, transparent_background=False,
                                       window_size=(width_pixels, height_pixels),
                                       return_img=False)
                    plotter.camera_position = camera
                    add_debug_message(f"Immagine salvata come {file_path}", important=True)
                    QMessageBox.information(self, "Success", f"Image saved {file_path} to 300 DPI (15x10 cm)")
            except Exception as e:
                add_debug_message(f'Error: {str(e)}', important=True)
                QMessageBox.warning(self, "Error", f"Error saving image: {str(e)}")

        def get_visible_faces(plotter, mesh):
            camera_position = np.array(plotter.camera_position[0])
            center = np.array(mesh.center)
            direction = camera_position - center
            normals = np.array([
                [1, 0, 0], [-1, 0, 0],
                [0, 1, 0], [0, -1, 0],
                [0, 0, 1], [0, 0, -1]
            ])
            return [i for i, normal in enumerate(normals) if np.dot(direction, normal) > 0]

        def edge_visibility(edge, visible_faces):
            edge_to_faces = {
                (0, 1): [0, 2, 4], (1, 2): [0, 1, 4], (2, 3): [0, 3, 4], (3, 0): [0, 2, 4],
                (4, 5): [1, 2, 5], (5, 6): [1, 3, 5], (6, 7): [1, 3, 5], (7, 4): [1, 2, 5],
                (0, 4): [2, 4, 5], (1, 5): [1, 4, 5], (2, 6): [1, 3, 5], (3, 7): [2, 3, 5]
            }
            return any(face in visible_faces for face in edge_to_faces[edge])

        def calculate_label_position(p1, p2, offset_factor=0.1):
            mid_point = (p1 + p2) / 2
            direction = p2 - p1
            length = np.linalg.norm(direction)
            normalized_direction = direction / length
            perpendicular = np.cross(normalized_direction, [0, 0, 1])
            if np.allclose(perpendicular, 0):
                perpendicular = np.cross(normalized_direction, [0, 1, 0])
            perpendicular = perpendicular / np.linalg.norm(perpendicular)
            return mid_point + perpendicular * (length * offset_factor)

        def create_oriented_label(plotter, position, text, direction, is_vertical=False):
            vtk_text = vtk.vtkBillboardTextActor3D()
            vtk_text.SetPosition(position)
            vtk_text.SetInput(text)
            vtk_text.GetTextProperty().SetFontSize(6)
            vtk_text.GetTextProperty().SetColor(0, 0, 0)  # Testo nero
            vtk_text.GetTextProperty().SetBackgroundColor(1, 1, 1)  # Sfondo bianco
            vtk_text.GetTextProperty().SetBackgroundOpacity(0.8)
            vtk_text.GetTextProperty().SetJustificationToCentered()
            vtk_text.GetTextProperty().SetVerticalJustificationToCentered()

            if is_vertical:
                angle = 90
            else:
                angle = np.degrees(np.arctan2(direction[1], direction[0]))
            vtk_text.SetOrientation(0, 0, angle)

            plotter.add_actor(vtk_text)
            return vtk_text

        self.last_update_time = 0
        self.update_interval = 0.5  # Secondi tra gli aggiornamenti
        bounding_box_visible = False

        def show_measures():
            nonlocal bounding_box_visible, measurement_objects
            if not bounding_box_visible:
                return

            current_time = time.time()
            if current_time - self.last_update_time < self.update_interval:
                return
            self.last_update_time = current_time

            # Rimuovi le misure esistenti
            for obj in measurement_objects:
                plotter.remove_actor(obj)
            measurement_objects = []

            bounds = mesh.bounds
            x_min, x_max, y_min, y_max, z_min, z_max = bounds
            point = np.array([
                [x_min, y_min, z_min], [x_max, y_min, z_min], [x_max, y_max, z_min], [x_min, y_max, z_min],
                [x_min, y_min, z_max], [x_max, y_min, z_max], [x_max, y_max, z_max], [x_min, y_max, z_max]
            ])

            edges = [
                (0, 1),  # Larghezza (X)
                (0, 3),  # Profondità (Y)
                (0, 4)  # Altezza (Z)
            ]

            get_visible_faces(plotter, mesh)

            for i, edge in enumerate(edges):
                # if edge_visibility(edge):
                p1, p2 = point[edge[0]], point[edge[1]]
                distance = np.linalg.norm(p2 - p1) * 100

                label_position = calculate_label_position(p1, p2)

                line = pv.Line(p1, p2)
                line_actor = plotter.add_mesh(line, color='black', line_width=0.8)
                measurement_objects.append(line_actor)

                label = f"{distance:.2f} cm"
                is_vertical = (i == 2)  # L'etichetta verticale è per l'altezza (Z)
                label_actor = create_oriented_label(plotter, label_position, label, p2 - p1, is_vertical)
                measurement_objects.append(label_actor)

            plotter.render()

        def toggle_bounding_box_measures():
            nonlocal bounding_box_visible
            bounding_box_visible = not bounding_box_visible
            if bounding_box_visible:
                show_measures()
                add_debug_message("Misure del bounding box attivate", important=True)
            else:
                for obj in measurement_objects:
                    plotter.remove_actor(obj)
                measurement_objects.clear()
                plotter.render()
                add_debug_message("Misure del bounding box disattivate", important=True)

        def camera_changed(obj, event):
            if bounding_box_visible:
                show_measures()

        plotter.iren.add_observer('InteractionEvent', camera_changed)

        def reset_view():
            plotter.reset_camera()

        def change_view(direction):
            getattr(plotter, f'view_{direction}')()

        plotter.add_key_event("o", toggle_measure)
        plotter.add_key_event("c", clear_measurements)
        plotter.add_key_event('e', export_image)
        plotter.add_key_event('m', toggle_bounding_box_measures)
        plotter.add_key_event('r', reset_view)
        plotter.add_key_event('x', lambda: change_view('yz'))
        plotter.add_key_event('y', lambda: change_view('xz'))
        plotter.add_key_event('z', lambda: change_view('xy'))
        plotter.add_key_event('w', lambda: change_view('zy'))
        plotter.add_key_event('v', lambda: change_view('zx'))
        plotter.add_key_event('b', lambda: change_view('yx'))

        plotter.add_orientation_widget(plotter.enable_trackball_style())
        plotter.enable_trackball_style()
        frame.show()
        return main_widget

    def generate_3d_thumbnail(self, filepath):

        mesh = pv.read(filepath)
        plotter = pv.Plotter(off_screen=True)
        plotter.add_mesh(mesh)
        plotter.camera_position = 'xy'

        # Genera un nome file unico per la thumbnail
        thumbnail_filename = f"{os.path.splitext(os.path.basename(filepath))[0]}_thumb.png"
        thumbnail_path = os.path.join(self.thumb_path, thumbnail_filename)

        plotter.screenshot(thumbnail_path)
        return thumbnail_path

    def process_3d_model(self, media_max_num_id, filepath, filename, thumb_path_str, thumb_resize_str,
                         media_thumb_suffix, media_resize_suffix):
        import pyvista as pv

        # Carica il modello 3D
        mesh = pv.read(filepath)

        # Genera una thumbnail
        plotter = pv.Plotter(off_screen=True)
        plotter.add_mesh(mesh)
        plotter.camera_position = 'xy'
        thumbnail_path = os.path.join(thumb_path_str, f"{media_max_num_id}_{filename}{media_thumb_suffix}")
        plotter.screenshot(thumbnail_path)

        # Copia il file originale nella cartella di resize (non possiamo ridimensionare un modello 3D come un'immagine)
        import shutil
        resize_path = os.path.join(thumb_resize_str, f"{media_max_num_id}_{filename}{media_resize_suffix}")
        shutil.copy(filepath, resize_path)
        # Controlla se esiste una texture JPG con lo stesso nome del modello
        texture_filename = os.path.splitext(filename)[0] + ".jpg"
        texture_filepath = os.path.join(os.path.dirname(filepath), texture_filename)

        if os.path.exists(texture_filepath):
            # Se la texture esiste, copiala nella cartella di resize
            texture_resize_path = os.path.join(thumb_resize_str, f"{media_max_num_id}_{texture_filename}")
            shutil.copy(texture_filepath, texture_resize_path)

        return thumbnail_path, resize_path

    def openWide_image(self):
        items = self.iconListWidget.selectedItems()
        conn = Connection()

        thumb_resize = conn.thumb_resize()
        thumb_resize_str = thumb_resize['thumb_resize']

        def process_file_path(file_path):
            return urllib.parse.unquote(file_path)

        def show_image(file_path):
            dlg = ImageViewer(self)
            dlg.show_image(file_path)
            dlg.exec_()

        def show_video(file_path):
            if self.video_player is None:
                self.video_player = VideoPlayerWindow(self, db_manager=self.DB_MANAGER,
                                                      icon_list_widget=self.iconListWidget,
                                                      main_class=self)
            self.video_player.set_video(file_path)
            self.video_player.show()

        def show_media(file_path, media_type):
            full_path = os.path.join(thumb_resize_str, file_path)
            if media_type == 'video':
                show_video(full_path)
            elif media_type == 'image':
                show_image(full_path)
            elif media_type == '3d_model':
                self.show_3d_model(file_path)
            else:
                QMessageBox.warning(self, "Error", f"Unsupported media type: {media_type}", QMessageBox.Ok)

        def query_media(search_dict, table="MEDIA_THUMB"):
            u = Utility()
            search_dict = u.remove_empty_items_fr_dict(search_dict)
            try:
                return self.DB_MANAGER.query_bool(search_dict, table)
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Database query failed: {str(e)}", QMessageBox.Ok)
                return None

        for item in items:
            id_orig_item = item.text()
            search_dict = {'media_filename': f"'{id_orig_item}'"}
            res = query_media(search_dict)

            if res:

                file_path = process_file_path(os.path.join(thumb_resize_str, str(res[0].path_resize)))

                media_type = res[0].mediatype

                if media_type == '3d_model':
                    widget_3d = self.show_3d_model(file_path)

                    # Crea un nuovo QDialog per contenere il widget 3D
                    dialog = QDialog(self)
                    dialog.setWindowTitle("3D Model Viewer")
                    dialog_layout = QVBoxLayout()
                    dialog_layout.addWidget(widget_3d)
                    dialog.setLayout(dialog_layout)

                    # Imposta le dimensioni del dialog
                    dialog.resize(800, 600)  # Puoi modificare queste dimensioni come preferisci

                    # Mostra il dialog
                    dialog.exec_()
                else:
                    show_media(file_path, media_type)
            else:
                QMessageBox.warning(self, "Error", f"File not found: {id_orig_item}", QMessageBox.Ok)

    def charge_list(self):
        # Get language from QgsSettings
        l = QgsSettings().value("locale/userLocale", QVariant)
        lang = ""
        for key, values in self.LANG.items():
            if values.__contains__(l):
                lang = str(key)
        # Fallback to EN if no match found
        if not lang:
            lang = "EN"
        lang = "'" + lang + "'"

        # Lista sito
        sito_vl = self.UTILITY.tup_2_list_III(self.DB_MANAGER.group_by('site_table', 'sito', 'SITE'))
        try:
            sito_vl.remove('')
        except Exception as e:
            if str(e) == "list.remove(x): x not in list":
                pass
            else:
                QMessageBox.warning(self, "Message", "Update system in site list: " + str(e), QMessageBox.Ok)

        self.comboBox_sito.clear()
        sito_vl.sort()
        self.comboBox_sito.addItems(sito_vl)

        # Lista area from thesaurus (11.13)
        self.comboBox_area.clear()
        search_dict = {
            'lingua': lang,
            'nome_tabella': "'Pottery'",
            'tipologia_sigla': "'11.13'"
        }
        area_vl_thesaurus = self.DB_MANAGER.query_bool(search_dict, 'PYARCHINIT_THESAURUS_SIGLE')
        area_vl = []
        for i in range(len(area_vl_thesaurus)):
            area_vl.append(area_vl_thesaurus[i].sigla_estesa)
        area_vl.sort()
        self.comboBox_area.addItems(area_vl)

        # Load thesaurus values for Pottery comboboxes
        self.charge_thesaurus_combos(lang)

    def charge_thesaurus_combos(self, lang):
        """Load thesaurus values for all Pottery comboboxes"""

        # Helper function to load thesaurus values
        def load_thesaurus(tipologia_sigla, use_sigla=False):
            search_dict = {
                'lingua': lang,
                'nome_tabella': "'Pottery'",
                'tipologia_sigla': "'" + tipologia_sigla + "'"
            }
            thesaurus_data = self.DB_MANAGER.query_bool(search_dict, 'PYARCHINIT_THESAURUS_SIGLE')
            values = []
            for item in thesaurus_data:
                if use_sigla:
                    val = item.sigla.strip() if item.sigla else ''
                else:
                    val = item.sigla_estesa.strip() if item.sigla_estesa else ''
                if val and val not in values:
                    values.append(val)
            values.sort()
            return values

        # 11.1 - Fabric
        self.comboBox_fabric.clear()
        fabric_vl = load_thesaurus('11.1')
        if fabric_vl:
            self.comboBox_fabric.addItems(fabric_vl)

        # 11.2 - Percent
        self.comboBox_percent.clear()
        percent_vl = load_thesaurus('11.2', use_sigla=True)
        if percent_vl:
            self.comboBox_percent.addItems(percent_vl)

        # 11.3 - Material
        self.comboBox_material.clear()
        material_vl = load_thesaurus('11.3')
        if material_vl:
            self.comboBox_material.addItems(material_vl)

        # 11.4 - Form (Open/Closed)
        self.comboBox_form.clear()
        form_vl = load_thesaurus('11.4')
        if form_vl:
            self.comboBox_form.addItems(form_vl)

        # 11.5 - Specific Form/Part
        self.comboBox_specific_form.clear()
        specific_form_vl = load_thesaurus('11.5')
        if specific_form_vl:
            self.comboBox_specific_form.addItems(specific_form_vl)

        # 11.6 - Ware Type
        self.comboBox_ware.clear()
        ware_vl = load_thesaurus('11.6')
        if ware_vl:
            self.comboBox_ware.addItems(ware_vl)

        # 11.7 - Munsell Color
        self.comboBox_munsell.clear()
        munsell_vl = load_thesaurus('11.7')
        if munsell_vl:
            self.comboBox_munsell.addItems(munsell_vl)

        # 11.8 - Surface Treatment
        self.comboBox_surf_trat.clear()
        surf_trat_vl = load_thesaurus('11.8')
        if surf_trat_vl:
            self.comboBox_surf_trat.addItems(surf_trat_vl)

        # 11.9 - External Decoration
        self.comboBox_exdeco.clear()
        exdeco_vl = load_thesaurus('11.9')
        if exdeco_vl:
            self.comboBox_exdeco.addItems(exdeco_vl)

        # 11.10 - Internal Decoration
        self.comboBox_intdeco.clear()
        intdeco_vl = load_thesaurus('11.10')
        if intdeco_vl:
            self.comboBox_intdeco.addItems(intdeco_vl)

        # 11.11 - Wheel Made
        self.comboBox_wheel_made.clear()
        wheel_made_vl = load_thesaurus('11.11')
        if wheel_made_vl:
            self.comboBox_wheel_made.addItems(wheel_made_vl)

        # 11.12 - Specific Shape
        self.comboBox_specific_shape.clear()
        specific_shape_vl = load_thesaurus('11.12')
        if specific_shape_vl:
            self.comboBox_specific_shape.addItems(specific_shape_vl)

        # 11.14 - Decoration Type
        self.comboBox_decoration_type.clear()
        decoration_type_vl = load_thesaurus('11.14')
        if decoration_type_vl:
            self.comboBox_decoration_type.addItems(decoration_type_vl)

        # 11.15 - Decoration Motif
        self.comboBox_decoration_motif.clear()
        decoration_motif_vl = load_thesaurus('11.15')
        if decoration_motif_vl:
            self.comboBox_decoration_motif.addItems(decoration_motif_vl)

        # 11.16 - Decoration Position
        self.comboBox_decoration_position.clear()
        decoration_position_vl = load_thesaurus('11.16')
        if decoration_position_vl:
            self.comboBox_decoration_position.addItems(decoration_position_vl)

        # Load datazione from periodizzazione_table
        self.load_datazione_list()

    def load_datazione_list(self):
        """Load datazione_estesa values from periodizzazione_table for current sito"""
        try:
            self.comboBox_datazione.clear()
            sito = str(self.comboBox_sito.currentText())
            if sito:
                search_dict = {'sito': "'" + sito + "'"}
                datazione_list_vl = self.DB_MANAGER.query_bool(search_dict, 'PERIODIZZAZIONE')
                datazione_list = []
                for item in datazione_list_vl:
                    if item.datazione_estesa and str(item.datazione_estesa).strip():
                        datazione_list.append(str(item.datazione_estesa))
                if datazione_list:
                    datazione_list = list(set(datazione_list))  # Remove duplicates
                    datazione_list.sort()
                    self.comboBox_datazione.addItems(datazione_list)
        except Exception as e:
            pass  # Silently fail if periodizzazione_table doesn't exist

    def on_toolButtonPreview_toggled(self):
        if self.L=='it':
            if self.toolButtonPreview.isChecked():
                QMessageBox.warning(self, "Messaggio",
                                    "Modalita' Preview US attivata. Le piante delle US saranno visualizzate nella sezione Piante",
                                    QMessageBox.Ok)
                self.loadMapPreview()
            else:
                self.loadMapPreview(1)
        elif self.L=='de':
            if self.toolButtonPreview.isChecked():
                QMessageBox.warning(self, "Message",
                                    "Modalität' Preview der aktivierten SE. Die Plana der SE werden in der Auswahl der Plana visualisiert",
                                    QMessageBox.Ok)
                self.loadMapPreview()
            else:
                self.loadMapPreview(1)
        else:
            if self.toolButtonPreview.isChecked():
                QMessageBox.warning(self, "Message",
                                    "Preview SU mode enabled. US plants will be displayed in the Plants section",
                                    QMessageBox.Ok)
                self.loadMapPreview()
            else:
                self.loadMapPreview(1)

    def on_pushButton_sort_pressed(self):
        if self.check_record_state() == 1:
            pass
        else:
            dlg = SortPanelMain(self)
            dlg.insertItems(self.SORT_ITEMS)
            dlg.exec_()

            items,order_type = dlg.ITEMS, dlg.TYPE_ORDER

            self.SORT_ITEMS_CONVERTED = []
            for i in items:
                #QMessageBox.warning(self, "Messaggio",i, QMessageBox.Ok)
                self.SORT_ITEMS_CONVERTED.append(self.CONVERSION_DICT[str(i)]) #apportare la modifica nellle altre schede

            self.SORT_MODE = order_type
            self.empty_fields()

            id_list = []
            for i in self.DATA_LIST:
                id_list.append(eval("i." + self.ID_TABLE))
            self.DATA_LIST = []

            temp_data_list = self.DB_MANAGER.query_sort(id_list, self.SORT_ITEMS_CONVERTED, self.SORT_MODE, self.MAPPER_TABLE_CLASS, self.ID_TABLE)

            for i in temp_data_list:
                self.DATA_LIST.append(i)
            self.BROWSE_STATUS = 'b'
            self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
            if type(self.REC_CORR) == "<type 'str'>":
                corr = 0
            else:
                corr = self.REC_CORR

            self.REC_TOT, self.REC_CORR = len(self.DATA_LIST), 0
            self.DATA_LIST_REC_TEMP = self.DATA_LIST_REC_CORR = self.DATA_LIST[0]
            self.SORT_STATUS = "o"
            self.label_sort.setText(self.SORTED_ITEMS[self.SORT_STATUS])
            self.set_rec_counter(len(self.DATA_LIST), self.REC_CORR+1)
            self.fill_fields()




    def insert_new_row(self, table_name):
        """insert new row into a table based on table_name"""
        cmd = table_name+".insertRow(0)"
        eval(cmd)


    def remove_row(self, table_name):
        """insert new row into a table based on table_name"""

        table_row_count_cmd = ("%s.rowCount()") % (table_name)
        table_row_count = eval(table_row_count_cmd)
        rowSelected_cmd = ("%s.selectedIndexes()") % (table_name)
        rowSelected = eval(rowSelected_cmd)
        try:
            rowIndex = (rowSelected[1].row())
            cmd = ("%s.removeRow(%d)") % (table_name, rowIndex)
            eval(cmd)
        except:
            QMessageBox.warning(self, "Messaggio", "Devi selezionare una riga",  QMessageBox.Ok)



    def on_pushButton_new_rec_pressed(self):
        conn = Connection()

        sito_set = conn.sito_set()
        sito_set_str = sito_set['sito_set']
        if bool(self.DATA_LIST):
            if self.data_error_check() == 1:
                pass
        if self.BROWSE_STATUS != "n":
            if bool(self.comboBox_sito.currentText()) and self.comboBox_sito.currentText() == sito_set_str:
                self.BROWSE_STATUS = "n"
                self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                self.empty_fields_nosite()

                #self.setComboBoxEditable(["self.lineEdit_id_number"], 1)
                #self.setComboBoxEditable(["self.comboBox_unita_tipo"], 1)
                self.setComboBoxEnable(["self.comboBox_sito"], False)
                self.setComboBoxEnable(["self.lineEdit_id_number"], True)
                #self.setComboBoxEnable(["self.lineEdit_us"], True)
                #self.setComboBoxEnable(["self.comboBox_unita_tipo"], True)
                self.SORT_STATUS = "n"
                self.label_sort.setText(self.SORTED_ITEMS[self.SORT_STATUS])
                self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                self.set_rec_counter('', '')
                self.label_sort.setText(self.SORTED_ITEMS["n"])

            else:
                self.BROWSE_STATUS = "n"
                self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                self.empty_fields()
                self.setComboBoxEditable(["self.comboBox_sito"], 1)
                #self.setComboBoxEditable(["self.lineEdit_id_number"], 1)
                #self.setComboBoxEditable(["self.comboBox_unita_tipo"], 1)
                #self.setComboBoxEnable(["self.comboBox_sito"], "True")
                #self.setComboBoxEnable(["self.comboBox_area"], "True")
                #self.setComboBoxEnable(["self.lineEdit_us"], "True")
                #self.setComboBoxEnable(["self.comboBox_unita_tipo"], "True")
                self.SORT_STATUS = "n"
                self.label_sort.setText(self.SORTED_ITEMS[self.SORT_STATUS])

                self.set_rec_counter('', '')
                self.label_sort.setText(self.SORTED_ITEMS["n"])

            self.enable_button(0)

    def on_pushButton_save_pressed(self):

        # self.checkBox_query.setChecked(False)
        # if self.checkBox_query.isChecked():
        #     self.model_a.database().close()
        if self.BROWSE_STATUS == "b":
            
                    # Check for version conflicts before updating
                    if hasattr(self, 'current_record_version') and self.current_record_version:
                        conflict = self.concurrency_manager.check_version_conflict(
                            'pottery_table',
                            self.editing_record_id,
                            self.current_record_version,
                            self.DB_MANAGER
                        )

                        if conflict and conflict['has_conflict']:
                            # Handle the conflict
                            record_data = self.fill_record()
                            if self.concurrency_manager.handle_conflict(
                                'pottery_table',
                                record_data,
                                conflict
                            ):
                                # User chose to reload - refresh the form
                                self.charge_records()
                                self.fill_fields(self.REC_CORR)
                                return
                            # Otherwise continue with save (user chose to overwrite)

                    if self.data_error_check() == 0:
                        if self.records_equal_check() == 1:
                            if self.L == 'it':
                                self.update_if(QMessageBox.warning(self, 'Errore',
                                                                   "Il record e' stato modificato. Vuoi salvare le modifiche?",
                                                                   QMessageBox.Ok | QMessageBox.Cancel))
                            elif self.L == 'de':
                                self.update_if(QMessageBox.warning(self, 'Error',
                                                                   "Der Record wurde geändert. Möchtest du die Änderungen speichern?",
                                                                   QMessageBox.Ok | QMessageBox.Cancel))
                            else:
                                self.update_if(QMessageBox.warning(self, 'Error',
                                                                   "The record has been changed. Do you want to save the changes?",
                                                                   QMessageBox.Ok | QMessageBox.Cancel))
                            self.empty_fields()
                            self.SORT_STATUS = "n"
                            self.label_sort.setText(self.SORTED_ITEMS[self.SORT_STATUS])
                            self.enable_button(1)

                            # Unlock the record after successful save
                            if hasattr(self, 'concurrency_manager') and hasattr(self, 'editing_record_id') and self.editing_record_id:
                                self.concurrency_manager.unlock_record('pottery_table', self.editing_record_id, self.DB_MANAGER)

                            self.fill_fields(self.REC_CORR)
                        else:
                            if self.L == 'it':
                                QMessageBox.warning(self, "ATTENZIONE", "Non è stata realizzata alcuna modifica.",
                                                    QMessageBox.Ok)
                            elif self.L == 'de':
                                QMessageBox.warning(self, "ACHTUNG", "Keine Änderung vorgenommen", QMessageBox.Ok)
                            else:
                                QMessageBox.warning(self, "Warning", "No changes have been made", QMessageBox.Ok)
        else:
            if self.data_error_check() == 0:
                test_insert = self.insert_new_rec()
                if test_insert == 1:
                    self.empty_fields()
                    self.SORT_STATUS = "n"
                    self.label_sort.setText(self.SORTED_ITEMS[self.SORT_STATUS])
                    self.charge_records()
                    self.charge_list()
                    self.set_sito()
                    self.BROWSE_STATUS = "b"
                    self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                    self.REC_TOT, self.REC_CORR = len(self.DATA_LIST), len(self.DATA_LIST) - 1
                    self.set_rec_counter(self.REC_TOT, self.REC_CORR + 1)
                    self.setComboBoxEditable(["self.comboBox_sito"], 1)
                    #self.setComboBoxEditable(["self.comboBox_area"], 1)
                    self.setComboBoxEnable(["self.lineEdit_id_number"], "False")
                    self.setComboBoxEnable(["self.comboBox_sito"], "False")
                    #self.setComboBoxEnable(["self.comboBox_area"], "False")
                    #self.setComboBoxEnable(["self.lineEdit_us"], "False")
                    self.fill_fields(self.REC_CORR)
                    self.enable_button(1)
            else:
                if self.L == 'it':
                    QMessageBox.warning(self, "ATTENZIONE", "Problema nell'inserimento dati", QMessageBox.Ok)
                elif self.L == 'de':
                    QMessageBox.warning(self, "ACHTUNG", "Problem der Dateneingabe", QMessageBox.Ok)
                else:
                    QMessageBox.warning(self, "Warning", "Problem with data entry", QMessageBox.Ok)


    def insert_new_rec(self):

        try:
            if self.lineEdit_us.text() == "":
                us = None
            else:
                us = str(self.lineEdit_us.text())

            if self.lineEdit_box.text() == "":
                box = None
            else:
                box = int(self.lineEdit_box.text())

            if self.lineEdit_anno.text() == "":
                anno = None
            else:
                anno = int(self.lineEdit_anno.text())

            if self.lineEdit_diametro_max.text() == "":
                diametro_max = None
            else:
                diametro_max = float(self.lineEdit_diametro_max.text())



            if self.lineEdit_qty.text() == "":
                qty = 0
            else:
                qty = int(self.lineEdit_qty.text())

            if self.lineEdit_diametro_rim.text() == "":
                diametro_rim = None
            else:
                diametro_rim = float(self.lineEdit_diametro_rim.text())

            if self.lineEdit_diametro_bottom.text() == "":
                diametro_bottom = None
            else:
                diametro_bottom = float(self.lineEdit_diametro_bottom.text())

            if self.lineEdit_diametro_height.text() == "":
                diametro_height = None
            else:
                diametro_height = float(self.lineEdit_diametro_height.text())

            if self.lineEdit_diametro_preserved.text() == "":
                diametro_preserved = None
            else:
                diametro_preserved = float(self.lineEdit_diametro_preserved.text())

            if self.lineEdit_bag.text() == "":
                bag = None
            else:
                bag = int(self.lineEdit_bag.text())
            data = self.DB_MANAGER.insert_pottery_values(
            self.DB_MANAGER.max_num_id(self.MAPPER_TABLE_CLASS, self.ID_TABLE)+1,
                    int(self.lineEdit_id_number.text()),
                    str(self.comboBox_sito.currentText()), 				#1 - Sito
                    str(self.comboBox_area.currentText()), 				#2 - Area
                    us,									#3 - US
                    box,
                    str(self.lineEdit_photo.text()),		#6 - descrizione
                    str(self.lineEdit_drawing.text()),#7 - interpretazione
                    anno,
                    str(self.comboBox_fabric.currentText()),								#14 - anno scavo
                    str(self.comboBox_percent.currentText()), 			#15 - metodo
                    str(self.comboBox_material.currentText()),			#9 - fase iniziale
                    str(self.comboBox_form.currentText()), 			#10 - periodo finale iniziale
                    str(self.comboBox_specific_form.currentText()), 			#11 - fase finale
                    str(self.comboBox_ware.currentText()),			#12 - scavato
                    str(self.comboBox_munsell.currentText()),							#13 - attivita
                    str(self.comboBox_surf_trat.currentText()),	#22 - conservazione												#18 - rapporti
                    str(self.comboBox_exdeco.currentText()),				#19 - data schedatura
                    str(self.comboBox_intdeco.currentText()),		#20 - schedatore
                    str(self.comboBox_wheel_made.currentText()),		#21 - formazione				#23 - colore
                    str(self.textEdit_descrip_ex_deco.toPlainText()),
                    str(self.textEdit_descrip_in_deco.toPlainText()),#24 - consistenza
                    str(self.textEdit_note.toPlainText()),
                    diametro_max,
                    qty,
                    diametro_rim,
                    diametro_bottom,
                    diametro_height,
                    diametro_preserved,
                    str(self.comboBox_specific_shape.currentText()),
                    bag,
                    str(self.comboBox_sector.currentText()),
                    str(self.comboBox_decoration_type.currentText()),
                    str(self.comboBox_decoration_motif.currentText()),
                    str(self.comboBox_decoration_position.currentText()),
                    str(self.comboBox_datazione.currentText()),
                    )							#25 - struttura
                                                    #28 - documentazione
            try:
                self.DB_MANAGER.insert_data_session(data)
                return 1
            except Exception as e:
                e_str = str(e)
                if e_str.__contains__("IntegrityError"):
                    msg = self.ID_TABLE + u" already present in database"
                    QMessageBox.warning(self, "Warning", "Error"+ str(msg),  QMessageBox.Ok)
                else:
                    msg = e
                    QMessageBox.warning(self, "Error", "Insert error 1 \n"+ str(msg),  QMessageBox.Ok)
                return 0

        except Exception as e:
            QMessageBox.warning(self, "Error", "Insert error 3 \n"+str(e),  QMessageBox.Ok)
            return 0
    #rif biblio
    # def on_pushButton_insert_row_rif_biblio_pressed(self):
    #     self.insert_new_row('self.tableWidget_rif_biblio')
    # def on_pushButton_remove_row_rif_biblio_pressed(self):
    #     self.remove_row('self.tableWidget_rif_biblio')
    def data_error_check(self):
        test = 0
        EC = Error_check()

        if EC.data_is_empty(str(self.lineEdit_id_number.text())) == 0:
            QMessageBox.warning(self, "Warning", "Site field. \n This field cannot be empty",  QMessageBox.Ok)
            test = 1

        if EC.data_is_empty(str(self.comboBox_sito.currentText())) == 0:
            QMessageBox.warning(self, "Warning", "Site field. \n This field cannot be empty",  QMessageBox.Ok)
            test = 1

        # if EC.data_is_empty(str(self.comboBox_area.currentText())) == 0:
        #     QMessageBox.warning(self, "Warning", "Area field. \n This field cannot be empty",  QMessageBox.Ok)
        #     test = 1
        #
        # if EC.data_is_empty(str(self.lineEdit_us.text())) == 0:
        #     QMessageBox.warning(self, "Warning", "US field. \n >This field cannot be empty",  QMessageBox.Ok)
        #     test = 1

        # area = self.comboBox_area.currentText()
        # us = self.lineEdit_us.text()
        # if us != "":
        #     if EC.data_is_int(us) == 0:
        #         QMessageBox.warning(self, "Warning", "US field. \n The value has to be numeric",  QMessageBox.Ok)
        #         test = 1

        return test

    def check_record_state(self):
        ec = self.data_error_check()
        if ec == 1:
            return 1  # ci sono errori di immissione
        elif self.records_equal_check() == 1 and ec == 0:
            if self.L == 'it':
                self.update_if(
                    QMessageBox.warning(self, 'Errore', "Il record e' stato modificato. Vuoi salvare le modifiche?",
                                        QMessageBox.Ok | QMessageBox.Cancel))
            elif self.L == 'de':
                self.update_if(
                    QMessageBox.warning(self, 'Errore',
                                        "Der Record wurde geändert. Möchtest du die Änderungen speichern?",
                                        QMessageBox.Ok | QMessageBox.Cancel))
            else:
                self.update_if(
                    QMessageBox.warning(self, "Error", "The record has been changed. You want to save the changes?",
                                        QMessageBox.Ok | QMessageBox.Cancel))
            return 0
            # records surf functions



    def on_pushButton_view_all_pressed(self):

        self.empty_fields()
        self.charge_records()
        self.fill_fields()
        self.BROWSE_STATUS = "b"
        self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
        if type(self.REC_CORR) == "<class 'str'>":
            corr = 0
        else:
            corr = self.REC_CORR
        self.set_rec_counter(len(self.DATA_LIST), self.REC_CORR + 1)
        self.REC_TOT, self.REC_CORR = len(self.DATA_LIST), 0
        self.DATA_LIST_REC_TEMP = self.DATA_LIST_REC_CORR = self.DATA_LIST[0]
        self.SORT_STATUS = "n"
        self.label_sort.setText(self.SORTED_ITEMS[self.SORT_STATUS])
    #records surf functions
    def on_pushButton_first_rec_pressed(self):
        if self.check_record_state() == 1:
            pass
        else:

            self.empty_fields()
            self.REC_TOT, self.REC_CORR = len(self.DATA_LIST), 0
            self.fill_fields(0)
            self.set_rec_counter(self.REC_TOT, self.REC_CORR+1)


    def on_pushButton_last_rec_pressed(self):
        if self.check_record_state() == 1:
            pass
        else:
            try:
                self.empty_fields()
                self.REC_TOT, self.REC_CORR = len(self.DATA_LIST), len(self.DATA_LIST) - 1
                self.fill_fields(self.REC_CORR)
                self.set_rec_counter(self.REC_TOT, self.REC_CORR + 1)
            except:
                pass

    def on_pushButton_prev_rec_pressed(self):
        if self.check_record_state() == 1:
            pass
        else:
            self.REC_CORR = self.REC_CORR-1
            if self.REC_CORR == -1:
                self.REC_CORR = 0
                QMessageBox.warning(self, "Error", "You are on the first record!",  QMessageBox.Ok)
            else:

                self.empty_fields()
                self.fill_fields(self.REC_CORR)
                self.set_rec_counter(self.REC_TOT, self.REC_CORR+1)


    def on_pushButton_next_rec_pressed(self):
        if self.check_record_state() == 1:
            pass
        else:
            self.REC_CORR = self.REC_CORR+1
            if self.REC_CORR >= self.REC_TOT:
                self.REC_CORR = self.REC_CORR-1
                QMessageBox.warning(self, "Error", "You are on the last record!",  QMessageBox.Ok)
            else:

                self.empty_fields()
                self.fill_fields(self.REC_CORR)
                self.set_rec_counter(self.REC_TOT, self.REC_CORR+1)



    def on_pushButton_delete_pressed(self):
        if self.L == 'it':
            msg = QMessageBox.warning(self, "Attenzione!!!",
                                      "Vuoi veramente eliminare il record? \n L'azione è irreversibile",
                                      QMessageBox.Ok | QMessageBox.Cancel)
            if msg == QMessageBox.Cancel:
                QMessageBox.warning(self, "Messagio!!!", "Azione Annullata!")
            else:
                try:
                    id_to_delete = eval("self.DATA_LIST[self.REC_CORR]." + self.ID_TABLE)
                    self.DB_MANAGER.delete_one_record(self.TABLE_NAME, self.ID_TABLE, id_to_delete)
                    self.charge_records()  # charge records from DB
                    QMessageBox.warning(self, "Messaggio!!!", "Record eliminato!")
                except Exception as e:
                    QMessageBox.warning(self, "Messaggio!!!", "Tipo di errore: " + str(e))
                if not bool(self.DATA_LIST):
                    QMessageBox.warning(self, "Attenzione", "Il database è vuoto!", QMessageBox.Ok)
                    self.DATA_LIST = []
                    self.DATA_LIST_REC_CORR = []
                    self.DATA_LIST_REC_TEMP = []
                    self.REC_CORR = 0
                    self.REC_TOT = 0
                    self.empty_fields()
                    self.set_rec_counter(0, 0)
                    # check if DB is empty
                if bool(self.DATA_LIST):
                    self.REC_TOT, self.REC_CORR = len(self.DATA_LIST), 0
                    self.DATA_LIST_REC_TEMP = self.DATA_LIST_REC_CORR = self.DATA_LIST[0]
                    self.BROWSE_STATUS = "b"
                    self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                    self.set_rec_counter(len(self.DATA_LIST), self.REC_CORR + 1)
                    self.charge_list()
                    self.fill_fields()
                    self.set_sito()
        elif self.L == 'de':
            msg = QMessageBox.warning(self, "Achtung!!!",
                                      "Willst du wirklich diesen Eintrag löschen? \n Der Vorgang ist unumkehrbar",
                                      QMessageBox.Ok | QMessageBox.Cancel)
            if msg == QMessageBox.Cancel:
                QMessageBox.warning(self, "Message!!!", "Aktion annulliert!")
            else:
                try:
                    id_to_delete = eval("self.DATA_LIST[self.REC_CORR]." + self.ID_TABLE)
                    self.DB_MANAGER.delete_one_record(self.TABLE_NAME, self.ID_TABLE, id_to_delete)
                    self.charge_records()  # charge records from DB
                    QMessageBox.warning(self, "Message!!!", "Record gelöscht!")
                except Exception as e:
                    QMessageBox.warning(self, "Messagge!!!", "Errortyp: " + str(e))
                if not bool(self.DATA_LIST):
                    QMessageBox.warning(self, "Attenzione", "Die Datenbank ist leer!", QMessageBox.Ok)
                    self.DATA_LIST = []
                    self.DATA_LIST_REC_CORR = []
                    self.DATA_LIST_REC_TEMP = []
                    self.REC_CORR = 0
                    self.REC_TOT = 0
                    self.empty_fields()
                    self.set_rec_counter(0, 0)
                    # check if DB is empty
                if bool(self.DATA_LIST):
                    self.REC_TOT, self.REC_CORR = len(self.DATA_LIST), 0
                    self.DATA_LIST_REC_TEMP = self.DATA_LIST_REC_CORR = self.DATA_LIST[0]
                    self.BROWSE_STATUS = "b"
                    self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                    self.set_rec_counter(len(self.DATA_LIST), self.REC_CORR + 1)
                    self.charge_list()
                    self.fill_fields()
                    self.set_sito()
        else:
            msg = QMessageBox.warning(self, "Warning!!!",
                                      "Do you really want to break the record? \n Action is irreversible.",
                                      QMessageBox.Ok | QMessageBox.Cancel)
            if msg == QMessageBox.Cancel:
                QMessageBox.warning(self, "Message!!!", "Action deleted!")
            else:
                try:
                    id_to_delete = eval("self.DATA_LIST[self.REC_CORR]." + self.ID_TABLE)
                    self.DB_MANAGER.delete_one_record(self.TABLE_NAME, self.ID_TABLE, id_to_delete)
                    self.charge_records()  # charge records from DB
                    QMessageBox.warning(self, "Message!!!", "Record deleted!")
                except Exception as e:
                    QMessageBox.warning(self, "Message", "error type: " + str(e))
                if not bool(self.DATA_LIST):
                    QMessageBox.warning(self, "Warning", "the db is empty!", QMessageBox.Ok)
                    self.DATA_LIST = []
                    self.DATA_LIST_REC_CORR = []
                    self.DATA_LIST_REC_TEMP = []
                    self.REC_CORR = 0
                    self.REC_TOT = 0
                    self.empty_fields()
                    self.set_rec_counter(0, 0)
                    # check if DB is empty
                if bool(self.DATA_LIST):
                    self.REC_TOT, self.REC_CORR = len(self.DATA_LIST), 0
                    self.DATA_LIST_REC_TEMP = self.DATA_LIST_REC_CORR = self.DATA_LIST[0]
                    self.BROWSE_STATUS = "b"
                    self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                    self.set_rec_counter(len(self.DATA_LIST), self.REC_CORR + 1)
                    self.charge_list()
                    self.fill_fields()
                    self.set_sito()
            self.SORT_STATUS = "n"
            self.label_sort.setText(self.SORTED_ITEMS[self.SORT_STATUS])


    def on_pushButton_new_search_pressed(self):
        if self.BROWSE_STATUS != "f" and self.check_record_state() == 1:
            pass
        else:
            self.enable_button_search(0)
            conn = Connection()

            sito_set = conn.sito_set()
            sito_set_str = sito_set['sito_set']
            if self.BROWSE_STATUS != "f":
                if bool(self.comboBox_sito.currentText()) and self.comboBox_sito.currentText() == sito_set_str:

                    self.BROWSE_STATUS = "f"
                    #self.setComboBoxEditable(["self.comboBox_sito"],0)
                    #self.setComboBoxEditable(["self.comboBox_area"],1)
                    self.setComboBoxEnable(["self.lineEdit_id_number"],"True")
                    self.setComboBoxEnable(["self.comboBox_sito"],"False")
                    #self.setComboBoxEnable(["self.comboBox_area"],"True")
                    #self.setComboBoxEnable(["self.lineEdit_us"],"True")
                    #self.setComboBoxEnable(["self.lineEdit_qty"],"True")
                    #self.setTableEnable(["self.tableWidget_rif_biblio"], "False")
                    ###
                    self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                    self.set_rec_counter('','')
                    self.label_sort.setText(self.SORTED_ITEMS["n"])
                    self.charge_list()
                    self.empty_fields_nosite()
                else:
                    self.BROWSE_STATUS = "f"
                    #self.setComboBoxEditable(["self.comboBox_sito"], 0)
                    #self.setComboBoxEditable(["self.comboBox_area"], 1)
                    self.setComboBoxEnable(["self.lineEdit_id_number"], "True")
                    self.setComboBoxEnable(["self.comboBox_sito"], "False")
                    #self.setComboBoxEnable(["self.comboBox_area"], "True")
                    #self.setComboBoxEnable(["self.lineEdit_us"], "True")
                    self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                    self.set_rec_counter('', '')
                    self.label_sort.setText(self.SORTED_ITEMS["n"])
                    self.charge_list()
                    self.empty_fields()

    def on_pushButton_search_go_pressed(self):
        check_for_buttons = 0
        if self.BROWSE_STATUS != "f":
            if self.L == 'it':
                QMessageBox.warning(self, "ATTENZIONE",
                                    "Per eseguire una nuova ricerca clicca sul pulsante 'new search' ",
                                    QMessageBox.Ok)
            elif self.L == 'de':
                QMessageBox.warning(self, "ACHTUNG", "Um eine neue Abfrage zu starten drücke  'new search' ",
                                    QMessageBox.Ok)
            else:
                QMessageBox.warning(self, "WARNING", "To perform a new search click on the 'new search' button ",
                                    QMessageBox.Ok)
        else:
            if self.lineEdit_id_number.text() != "":
                id_number = int(self.lineEdit_id_number.text())
            else:
                id_number = ""
            if self.lineEdit_us.text() != "":
                us = str(self.lineEdit_us.text())
            else:
                us = ""
            if self.lineEdit_box.text() != "":
                box = int(self.lineEdit_box.text())
            else:
                box = ""
            if self.lineEdit_anno.text() != "":
                anno = int(self.lineEdit_anno.text())
            else:
                anno = ""



            if self.lineEdit_diametro_max.text() != "":
                diametro_max = float(self.lineEdit_diametro_max.text())
            else:
                diametro_max = None



            if self.lineEdit_qty.text() != "":
                qty = int(self.lineEdit_qty.text())
            else:
                qty = ""

            if self.lineEdit_diametro_rim.text() != "":
                diametro_rim = float(self.lineEdit_diametro_rim.text())
            else:
                diametro_rim = None

            if self.lineEdit_diametro_bottom.text() != "":
                diametro_bottom = float(self.lineEdit_diametro_bottom.text())
            else:
                diametro_bottom = None

            if self.lineEdit_diametro_height.text() != "":
                diametro_height = float(self.lineEdit_diametro_height.text())
            else:
                diametro_height = None

            if self.lineEdit_diametro_preserved.text() != "":
                diametro_preserved = float(self.lineEdit_diametro_preserved.text())
            else:
                diametro_preserved = None

            if self.lineEdit_bag.text() != "":
                bag = int(self.lineEdit_bag.text())
            else:
                bag = ""
            search_dict = {
            self.TABLE_FIELDS[0]  : id_number,									#1 - Sito
            self.TABLE_FIELDS[1]  : "'"+str(self.comboBox_sito.currentText())+"'",								#2 - Area
            self.TABLE_FIELDS[2]  : "'"+str(self.comboBox_area.currentText())+"'",																									#3 - US
            self.TABLE_FIELDS[3]  : us,								#4 - Definizione stratigrafica
            self.TABLE_FIELDS[4]  : box,							#5 - Definizione intepretata
            self.TABLE_FIELDS[5]  : "'"+str(self.lineEdit_photo.text())+"'",											#6 - descrizione
            self.TABLE_FIELDS[6]  : "'"+str(self.lineEdit_drawing.text())+"'",										#7 - interpretazione
            self.TABLE_FIELDS[7]  : anno,								#8 - periodo iniziale
            self.TABLE_FIELDS[8]  : "'"+str(self.comboBox_fabric.currentText())+"'",								#9 - fase iniziale
            self.TABLE_FIELDS[9]  : "'"+str(self.comboBox_percent.currentText())+"'",	 							#10 - periodo finale iniziale
            self.TABLE_FIELDS[10] : "'"+str(self.comboBox_material.currentText())+"'", 								#11 - fase finale
            self.TABLE_FIELDS[11] : "'"+str(self.comboBox_form.currentText())+"'",								#12 - scavato
            self.TABLE_FIELDS[12] : "'"+str(self.comboBox_specific_form.currentText())+"'",												#13 - attivita
            self.TABLE_FIELDS[13] : "'"+str(self.comboBox_ware.currentText())+"'",													#14 - anno scavo
            self.TABLE_FIELDS[14] : "'"+str(self.comboBox_munsell.currentText())+"'", 								#15 - metodo
            self.TABLE_FIELDS[15] : "'"+str(self.comboBox_surf_trat.currentText())+"'",
            self.TABLE_FIELDS[16] : "'"+str(self.comboBox_exdeco.currentText())+"'",
            self.TABLE_FIELDS[17] : "'"+str(self.comboBox_intdeco.currentText())+"'",
            self.TABLE_FIELDS[18] : "'"+str(self.comboBox_wheel_made.currentText())+"'",#16 - data schedatura
            self.TABLE_FIELDS[19] : "'"+str(self.textEdit_descrip_ex_deco.toPlainText())+ "'",
            self.TABLE_FIELDS[20] : "'"+str(self.textEdit_descrip_in_deco.toPlainText())+ "'",
            self.TABLE_FIELDS[21] : "'"+str(self.textEdit_note.toPlainText())+ "'",				#19 - conservazione
            self.TABLE_FIELDS[22] : diametro_max,
            self.TABLE_FIELDS[23] : qty,
            self.TABLE_FIELDS[24] : diametro_rim, 								#15 - metodo
            self.TABLE_FIELDS[25] : diametro_bottom,
            self.TABLE_FIELDS[26] : diametro_height,
            self.TABLE_FIELDS[27] : diametro_preserved,
            self.TABLE_FIELDS[28] : "'"+str(self.comboBox_specific_shape.currentText())+"'",
            self.TABLE_FIELDS[29]: bag,
            self.TABLE_FIELDS[30]: "'" + str(self.comboBox_sector.currentText()) + "'",
            self.TABLE_FIELDS[31]: "'" + str(self.comboBox_decoration_type.currentText()) + "'",
            self.TABLE_FIELDS[32]: "'" + str(self.comboBox_decoration_motif.currentText()) + "'",
            self.TABLE_FIELDS[33]: "'" + str(self.comboBox_decoration_position.currentText()) + "'",
            self.TABLE_FIELDS[34]: "'" + str(self.comboBox_datazione.currentText()) + "'",
            }
            u = Utility()
            search_dict = u.remove_empty_items_fr_dict(search_dict)
            if not bool(search_dict):
                if self.L == 'it':
                    QMessageBox.warning(self, "ATTENZIONE", "Non è stata impostata nessuna ricerca!!!", QMessageBox.Ok)
                elif self.L == 'de':
                    QMessageBox.warning(self, "ACHTUNG", "Keine Abfrage definiert!!!", QMessageBox.Ok)
                else:
                    QMessageBox.warning(self, " WARNING", "No search has been set!!!", QMessageBox.Ok)
            else:
                res = self.DB_MANAGER.query_bool(search_dict, self.MAPPER_TABLE_CLASS)
                if not bool(res):
                    if self.L == 'it':
                        QMessageBox.warning(self, "ATTENZIONE", "Non è stato trovato nessun record!", QMessageBox.Ok)
                    elif self.L == 'de':
                        QMessageBox.warning(self, "ACHTUNG", "Keinen Record gefunden!", QMessageBox.Ok)
                    else:
                        QMessageBox.warning(self, "WARNING", "No record found!", QMessageBox.Ok)
                    self.set_rec_counter(len(self.DATA_LIST), self.REC_CORR + 1)
                    self.DATA_LIST_REC_TEMP = self.DATA_LIST_REC_CORR = self.DATA_LIST[0]
                    self.BROWSE_STATUS = "b"
                    self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                    self.setComboBoxEnable(["self.comboBox_sito"], "False")
                    #self.setComboBoxEnable(["self.comboBox_area"], "True")

                    #self.setComboBoxEnable(["self.lineEdit_us"], "True")


                    self.fill_fields(self.REC_CORR)
                else:
                    self.DATA_LIST = []
                    for i in res:
                        self.DATA_LIST.append(i)
                    self.REC_TOT, self.REC_CORR = len(self.DATA_LIST), 0
                    self.DATA_LIST_REC_TEMP = self.DATA_LIST_REC_CORR = self.DATA_LIST[0]
                    self.fill_fields()
                    self.BROWSE_STATUS = "b"
                    self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                    self.set_rec_counter(len(self.DATA_LIST), self.REC_CORR + 1)
                    if self.L == 'it':
                        if self.REC_TOT == 1:
                            strings = ("E' stato trovato", self.REC_TOT, "record")

                        else:
                            strings = ("Sono stati trovati", self.REC_TOT, "records")

                    elif self.L == 'de':
                        if self.REC_TOT == 1:
                            strings = ("Es wurde gefunden", self.REC_TOT, "record")

                        else:
                            strings = ("Sie wurden gefunden", self.REC_TOT, "records")

                    else:
                        if self.REC_TOT == 1:
                            strings = ("It has been found", self.REC_TOT, "record")

                        else:
                            strings = ("They have been found", self.REC_TOT, "records")


                    self.setComboBoxEnable(["self.comboBox_sito"], "False")
                    #self.setComboBoxEnable(["self.comboBox_area"], "True")
                    #self.setComboBoxEnable(["self.lineEdit_us"], "True")


                    QMessageBox.warning(self, "Message", "%s %d %s" % strings, QMessageBox.Ok)
        self.enable_button_search(1)

    def update_if(self, msg):
        rec_corr = self.REC_CORR
        if msg == QMessageBox.Ok:
            test = self.update_record()
            if test == 1:
                id_list = []
                for i in self.DATA_LIST:
                    id_list.append(eval("i." + self.ID_TABLE))
                self.DATA_LIST = []
                if self.SORT_STATUS == "n":
                    temp_data_list = self.DB_MANAGER.query_sort(id_list, [self.ID_TABLE], 'asc',
                                                                self.MAPPER_TABLE_CLASS,
                                                                self.ID_TABLE)
                else:
                    temp_data_list = self.DB_MANAGER.query_sort(id_list, self.SORT_ITEMS_CONVERTED, self.SORT_MODE,
                                                                self.MAPPER_TABLE_CLASS, self.ID_TABLE)
                for i in temp_data_list:
                    self.DATA_LIST.append(i)
                self.BROWSE_STATUS = "b"
                self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
                if type(self.REC_CORR) == "<type 'str'>":
                    corr = 0
                else:
                    corr = self.REC_CORR
                return 1
            elif test == 0:
                return 0

    def update_record(self):
        try:
            self.DB_MANAGER.update(self.MAPPER_TABLE_CLASS,
                                   self.ID_TABLE,
                                   [eval("int(self.DATA_LIST[self.REC_CORR]." + self.ID_TABLE + ")")],
                                   self.TABLE_FIELDS,
                                   self.rec_toupdate())
            return 1
        except Exception as e:
            str(e)
            save_file = '{}{}{}'.format(self.HOME, os.sep, "pyarchinit_Report_folder")
            file_ = os.path.join(save_file, 'error_encodig_data_recover.txt')
            with open(file_, "a") as fh:
                try:
                    raise ValueError(str(e))
                except ValueError as s:
                    print(s, file=fh)
            if self.L == 'it':
                QMessageBox.warning(self, "Messaggio",
                                    "Problema di encoding: sono stati inseriti accenti o caratteri non accettati dal database. Verrà fatta una copia dell'errore con i dati che puoi recuperare nella cartella pyarchinit_Report _Folder",
                                    QMessageBox.Ok)


            elif self.L == 'de':
                QMessageBox.warning(self, "Message",
                                    "Encoding problem: accents or characters not accepted by the database were entered. A copy of the error will be made with the data you can retrieve in the pyarchinit_Report _Folder",
                                    QMessageBox.Ok)
            else:
                QMessageBox.warning(self, "Message",
                                    "Kodierungsproblem: Es wurden Akzente oder Zeichen eingegeben, die von der Datenbank nicht akzeptiert werden. Es wird eine Kopie des Fehlers mit den Daten erstellt, die Sie im pyarchinit_Report _Ordner abrufen können",
                                    QMessageBox.Ok)

            return 0

    def rec_toupdate(self):
        rec_to_update = self.UTILITY.pos_none_in_list(self.DATA_LIST_REC_TEMP)
        return rec_to_update
        # custom functions

    def charge_records(self):
        self.DATA_LIST = []
        if self.DB_SERVER == 'sqlite':
            for i in self.DB_MANAGER.query(self.MAPPER_TABLE_CLASS):
                self.DATA_LIST.append(i)
        else:
            id_list = []
            for i in self.DB_MANAGER.query(self.MAPPER_TABLE_CLASS):
                id_list.append(eval("i." + self.ID_TABLE))
            temp_data_list = self.DB_MANAGER.query_sort(id_list, [self.ID_TABLE], 'asc', self.MAPPER_TABLE_CLASS,
                                                        self.ID_TABLE)
            for i in temp_data_list:
                self.DATA_LIST.append(i)


    def datestrfdate(self):
        now = date.today()
        today = now.strftime("%d-%m-%Y")
        return today

    def yearstrfdate(self):
        now = date.today()
        year = now.strftime("%Y")
        return year

    def table2dict(self, n):
        self.tablename = n
        row = eval(self.tablename+".rowCount()")
        col = eval(self.tablename+".columnCount()")
        lista=[]
        for r in range(row):
            sub_list = []
            for c in range(col):
                value = eval(self.tablename+".item(r,c)")
                if value != None:
                    sub_list.append(str(value.text()))

            if bool(sub_list) == True:
                lista.append(sub_list)

        return lista


    def tableInsertData(self, t, d):
        """Set the value into alls Grid"""
        self.table_name = t
        self.data_list = eval(d)
        self.data_list.sort()

        #column table count
        table_col_count_cmd = ("%s.columnCount()") % (self.table_name)
        table_col_count = eval(table_col_count_cmd)

        #clear table
        table_clear_cmd = ("%s.clearContents()") % (self.table_name)
        eval(table_clear_cmd)

        for i in range(table_col_count):
            table_rem_row_cmd = ("%s.removeRow(%d)") % (self.table_name, i)
            eval(table_rem_row_cmd)

        for i in range(len(self.data_list)):
            self.insert_new_row(self.table_name)

        for row in range(len(self.data_list)):
            cmd = ('%s.insertRow(%s)') % (self.table_name, row)
            eval(cmd)
            for col in range(len(self.data_list[row])):
                #item = self.comboBox_sito.setEditText(self.data_list[0][col]
                item = QTableWidgetItem(str(self.data_list[row][col]))
                exec_str = ('%s.setItem(%d,%d,item)') % (self.table_name,row,col)
                eval(exec_str)





    def empty_fields(self):
        #rif_biblio_row_count = self.tableWidget_rif_biblio.rowCount()
        self.lineEdit_id_number.clear()
        self.comboBox_sito.setEditText("")  								#1 - Sito
        self.comboBox_area.setEditText("") 								#2 - Area
        self.lineEdit_us.clear()													#3 - US
        self.lineEdit_box.clear()						#4 - Definizione stratigrafica					#5 - Definizione intepretata
        self.lineEdit_photo.clear()#9 - fase iniziale
        self.lineEdit_drawing.clear()									#7 - interpretazione
        self.lineEdit_anno.clear()
        self.comboBox_fabric.setEditText("")
        self.comboBox_percent.setEditText("")
        self.comboBox_material.setEditText("")									#8 - periodo iniziale
        self.comboBox_form.setEditText("")
        self.comboBox_specific_form.setEditText("")
        self.comboBox_ware.setEditText("")								#10 - periodo finale iniziale
        self.comboBox_munsell.setEditText("")								#11 - fase finale
        self.comboBox_surf_trat.setEditText("")								#12 - scavato
        self.comboBox_exdeco.setEditText("")	#20 - data schedatura											#13 - attivita
        self.comboBox_intdeco.setEditText("")						#21 - schedatore
        self.comboBox_wheel_made.setEditText("")				#22 - formazione
        self.textEdit_descrip_ex_deco.clear()#6 - descrizione
        self.textEdit_descrip_in_deco.clear()
        self.textEdit_note.clear()
        self.lineEdit_diametro_max.clear()

        self.lineEdit_qty.clear()
        self.lineEdit_diametro_rim.clear()
        self.lineEdit_diametro_bottom.clear()
        self.lineEdit_diametro_height.clear()
        self.lineEdit_diametro_preserved.clear()
        self.comboBox_specific_shape.setEditText("")
        self.lineEdit_bag.clear()
        self.comboBox_sector.setEditText("")
        self.comboBox_decoration_type.setEditText("")
        self.comboBox_decoration_motif.setEditText("")
        self.comboBox_decoration_position.setEditText("")
        self.comboBox_datazione.setEditText("")
        #for i in range(rif_biblio_row_count):
            #self.tableWidget_rif_biblio.removeRow(0)
        #self.insert_new_row("self.tableWidget_rif_biblio")
    def empty_fields_nosite(self):
        #rif_biblio_row_count = self.tableWidget_rif_biblio.rowCount()
        self.lineEdit_id_number.clear()
         								#1 - Sito
        self.comboBox_area.setEditText("") 								#2 - Area
        self.lineEdit_us.clear()													#3 - US
        self.lineEdit_box.clear()						#4 - Definizione stratigrafica					#5 - Definizione intepretata
        self.lineEdit_photo.clear()#9 - fase iniziale
        self.lineEdit_drawing.clear()									#7 - interpretazione
        self.lineEdit_anno.clear()
        self.comboBox_fabric.setEditText("")
        self.comboBox_percent.setEditText("")
        self.comboBox_material.setEditText("")									#8 - periodo iniziale
        self.comboBox_form.setEditText("")
        self.comboBox_specific_form.setEditText("")
        self.comboBox_ware.setEditText("")								#10 - periodo finale iniziale
        self.comboBox_munsell.setEditText("")								#11 - fase finale
        self.comboBox_surf_trat.setEditText("")								#12 - scavato
        self.comboBox_exdeco.setEditText("")	#20 - data schedatura											#13 - attivita
        self.comboBox_intdeco.setEditText("")						#21 - schedatore
        self.comboBox_wheel_made.setEditText("")				#22 - formazione
        self.textEdit_descrip_ex_deco.clear()#6 - descrizione
        self.textEdit_descrip_in_deco.clear()
        self.textEdit_note.clear()
        self.lineEdit_diametro_max.clear()

        self.lineEdit_qty.clear()
        self.lineEdit_diametro_rim.clear()
        self.lineEdit_diametro_bottom.clear()
        self.lineEdit_diametro_height.clear()
        self.lineEdit_diametro_preserved.clear()
        self.comboBox_specific_shape.setEditText("")
        self.lineEdit_bag.clear()
        self.comboBox_sector.setEditText("")
        self.comboBox_decoration_type.setEditText("")
        self.comboBox_decoration_motif.setEditText("")
        self.comboBox_decoration_position.setEditText("")
        self.comboBox_datazione.setEditText("")
        #for i in range(rif_biblio_row_count):
            #self.tableWidget_rif_biblio.removeRow(0)
        #self.insert_new_row("self.tableWidget_rif_biblio")

    def fill_fields(self, n=0):
        self.rec_num = n
        #QMessageBox.warning(self, "check fill fields", str(self.rec_num),  QMessageBox.Ok)
        try:
            self.lineEdit_id_number.setText(str(self.DATA_LIST[self.rec_num].id_number))	#3 - US
            str(self.comboBox_sito.setEditText(self.DATA_LIST[self.rec_num].sito))													#1 - Sito
            str(self.comboBox_area.setEditText(self.DATA_LIST[self.rec_num].area))												#2 - Area
            if not self.DATA_LIST[self.rec_num].us:
                self.lineEdit_us.setText("")
            else:
                self.lineEdit_us.setText(str(self.DATA_LIST[self.rec_num].us))
            if not self.DATA_LIST[self.rec_num].box:
                self.lineEdit_box.setText("")
            else:
                self.lineEdit_box.setText(str(self.DATA_LIST[self.rec_num].box))
            str(self.lineEdit_photo.setText(self.DATA_LIST[self.rec_num].photo))
            str(self.lineEdit_drawing.setText(self.DATA_LIST[self.rec_num].drawing))
            if not self.DATA_LIST[self.rec_num].anno:
                self.lineEdit_anno.setText("")
            else:
                self.lineEdit_anno.setText(str(self.DATA_LIST[self.rec_num].anno))
            str(self.comboBox_fabric.setEditText(self.DATA_LIST[self.rec_num].fabric))
            str(self.comboBox_percent.setEditText(self.DATA_LIST[self.rec_num].percent))#5 - Definizione intepretata
            str(self.comboBox_material.setEditText(self.DATA_LIST[self.rec_num].material))
            str(self.comboBox_form.setEditText(self.DATA_LIST[self.rec_num].form))#5 - Definizione intepretata
            str(self.comboBox_specific_form.setEditText(self.DATA_LIST[self.rec_num].specific_form))#5 - Definizione intepretata
            str(self.comboBox_ware.setEditText(self.DATA_LIST[self.rec_num].ware))#5 - Definizione intepretata
            str(self.comboBox_munsell.setEditText(self.DATA_LIST[self.rec_num].munsell))#5 - Definizione intepretata
            str(self.comboBox_surf_trat.setEditText(self.DATA_LIST[self.rec_num].surf_trat))#5 - Definizione intepretata
            str(self.comboBox_exdeco.setEditText(self.DATA_LIST[self.rec_num].exdeco))#5 - Definizione intepretata
            str(self.comboBox_intdeco.setEditText(self.DATA_LIST[self.rec_num].intdeco))#5 - Definizione intepretata
            str(self.comboBox_wheel_made.setEditText(self.DATA_LIST[self.rec_num].wheel_made))#5 - Definizione intepretata
            str(self.textEdit_descrip_in_deco.setText(self.DATA_LIST[self.rec_num].descrip_in_deco))								#6 - descrizione
            str(self.textEdit_descrip_ex_deco.setText(self.DATA_LIST[self.rec_num].descrip_ex_deco))
            str(self.textEdit_note.setText(self.DATA_LIST[self.rec_num].note))
            #7 - interpretazione


            #self.tableInsertData("self.tableWidget_rif_biblio", self.DATA_LIST[self.rec_num].rif_biblio)
            if not self.DATA_LIST[self.rec_num].diametro_max:
                self.lineEdit_diametro_max.setText("")
            else:
                self.lineEdit_diametro_max.setText(str(self.DATA_LIST[self.rec_num].diametro_max))
            if not self.DATA_LIST[self.rec_num].qty:
                self.lineEdit_qty.setText("")
            else:
                self.lineEdit_qty.setText(str(self.DATA_LIST[self.rec_num].qty))

            if not self.DATA_LIST[self.rec_num].diametro_rim:
                self.lineEdit_diametro_rim.setText("")
            else:
                self.lineEdit_diametro_rim.setText(str(self.DATA_LIST[self.rec_num].diametro_rim))

            if not self.DATA_LIST[self.rec_num].diametro_bottom:
                self.lineEdit_diametro_bottom.setText("")
            else:
                self.lineEdit_diametro_bottom.setText(str(self.DATA_LIST[self.rec_num].diametro_bottom))

            if not self.DATA_LIST[self.rec_num].diametro_height:
                self.lineEdit_diametro_height.setText("")
            else:
                self.lineEdit_diametro_height.setText(str(self.DATA_LIST[self.rec_num].diametro_height))

            if not self.DATA_LIST[self.rec_num].diametro_preserved:
                self.lineEdit_diametro_preserved.setText("")
            else:
                self.lineEdit_diametro_preserved.setText(str(self.DATA_LIST[self.rec_num].diametro_preserved))


            str(self.comboBox_specific_shape.setEditText(self.DATA_LIST[self.rec_num].specific_shape))

            if not self.DATA_LIST[self.rec_num].bag:
                self.lineEdit_bag.setText("")
            else:
                self.lineEdit_bag.setText(str(self.DATA_LIST[self.rec_num].bag))

            str(self.comboBox_sector.setEditText(self.DATA_LIST[self.rec_num].sector))

            # New decoration fields

            str(self.comboBox_decoration_type.setEditText(self.DATA_LIST[self.rec_num].decoration_type))



            str(self.comboBox_decoration_motif.setEditText(self.DATA_LIST[self.rec_num].decoration_motif))



            str(self.comboBox_decoration_position.setEditText(self.DATA_LIST[self.rec_num].decoration_position))

            # Datazione field
            if self.DATA_LIST[self.rec_num].datazione:
                self.comboBox_datazione.setEditText(str(self.DATA_LIST[self.rec_num].datazione))
            else:
                self.comboBox_datazione.setEditText("")

            if self.toolButtonPreviewMedia.isChecked() == False:
                self.loadMediaPreview()
        except Exception as e:
            QMessageBox.warning(self, "Error Fill Fields", str(e),  QMessageBox.Ok)

        # Track version number and record ID for concurrency
        if hasattr(self, 'concurrency_manager'):
            try:
                if n < len(self.DATA_LIST):
                    current_record = self.DATA_LIST[n]
                    if hasattr(current_record, 'version_number'):
                        self.current_record_version = current_record.version_number
                    if hasattr(current_record, 'id_rep'):
                        self.editing_record_id = getattr(current_record, 'id_rep')

                    # Update lock indicator
                    if hasattr(current_record, 'editing_by'):
                        self.lock_indicator.update_lock_status(
                            current_record.editing_by,
                            current_record.editing_since if hasattr(current_record, 'editing_since') else None
                        )

                    # Set soft lock for this record
                    if self.editing_record_id and hasattr(self, 'concurrency_manager'):
                        self.concurrency_manager.lock_record(
                            'pottery_table',
                            self.editing_record_id,
                            self.DB_MANAGER
                        )
            except Exception as e:
                QgsMessageLog.logMessage(f"Error setting version tracking: {str(e)}", "PyArchInit", Qgis.Warning)

    
    def set_rec_counter(self, t, c):
        self.rec_tot = t
        self.rec_corr = c
        self.label_rec_tot.setText(str(self.rec_tot))
        self.label_rec_corrente.setText(str(self.rec_corr))

    def set_LIST_REC_TEMP(self):
        if self.lineEdit_us.text() == "":
            us = None
        else:
            us = int(self.lineEdit_us.text())

        if self.lineEdit_box.text() == "":
            box =None
        else:
            box = int(self.lineEdit_box.text())

        if self.lineEdit_anno.text() == "":
            anno = None
        else:
            anno = int(self.lineEdit_anno.text())



        if self.lineEdit_diametro_max.text() == "":
            diametro_max = None
        else:
            diametro_max = float(self.lineEdit_diametro_max.text())


        if self.lineEdit_qty.text() == "":
            qty = 0
        else:
            qty = int(self.lineEdit_qty.text())

        if self.lineEdit_diametro_rim.text() == "":
            diametro_rim = None
        else:
            diametro_rim = float(self.lineEdit_diametro_rim.text())

        if self.lineEdit_diametro_bottom.text() == "":
            diametro_bottom = None
        else:
            diametro_bottom = float(self.lineEdit_diametro_bottom.text())

        if self.lineEdit_diametro_height.text() == "":
            diametro_height = None
        else:
            diametro_height = float(self.lineEdit_diametro_height.text())

        if self.lineEdit_diametro_preserved.text() == "":
            diametro_preserved = None
        else:
            diametro_preserved = float(self.lineEdit_diametro_preserved.text())

        if self.lineEdit_bag.text() == "":
            bag = None
        else:
            bag = int(self.lineEdit_bag.text())


        self.DATA_LIST_REC_TEMP = [
        str(self.lineEdit_id_number.text()),	#3 - US
        str(self.comboBox_sito.currentText()), 						#1 - Sito
        str(self.comboBox_area.currentText()), 						#2 - Area
        str(us),
        str(box),
        str(self.lineEdit_photo.text()),
        str(self.lineEdit_drawing.text()),
        str(anno),	#3 - US
        str(self.comboBox_fabric.currentText()),
        str(self.comboBox_percent.currentText()),
        str(self.comboBox_material.currentText()),#4 - Definizione stratigrafica
        str(self.comboBox_form.currentText()),
        str(self.comboBox_specific_form.currentText()),
        str(self.comboBox_ware.currentText()),
        str(self.comboBox_munsell.currentText()),
        str(self.comboBox_surf_trat.currentText()),
        str(self.comboBox_exdeco.currentText()),
        str(self.comboBox_intdeco.currentText()),
        str(self.comboBox_wheel_made.currentText()),
        str(self.textEdit_descrip_ex_deco.toPlainText()),		#6 - descrizione
        str(self.textEdit_descrip_in_deco.toPlainText()),
        str(self.textEdit_note.toPlainText()),
        str(diametro_max),

        str(qty),
        str(diametro_rim),
        str(diametro_bottom),
        str(diametro_height),
        str(diametro_preserved),
        str(self.comboBox_specific_shape.currentText()),
        str(bag),
        str(self.comboBox_sector.currentText()),
        str(self.comboBox_decoration_type.currentText()),
        str(self.comboBox_decoration_motif.currentText()),
        str(self.comboBox_decoration_position.currentText()),
        str(self.comboBox_datazione.currentText()),
        ]

    def set_LIST_REC_CORR(self):
        self.DATA_LIST_REC_CORR = []
        for i in self.TABLE_FIELDS:
            # Use getattr with default None to handle missing attributes (backward compatibility)
            value = getattr(self.DATA_LIST[self.REC_CORR], i, None)
            self.DATA_LIST_REC_CORR.append(str(value) if value is not None else "")

    def records_equal_check(self):
        self.set_LIST_REC_TEMP()
        self.set_LIST_REC_CORR()
        #QMessageBox.warning(self, "Error", str(self.DATA_LIST_REC_CORR) + str(self.DATA_LIST_REC_TEMP),  QMessageBox.Ok)
        if self.DATA_LIST_REC_CORR == self.DATA_LIST_REC_TEMP:
            return 0
        else:
            return 1

    def setComboBoxEditable(self, f, n):
        field_names = f
        value = n

        for fn in field_names:
            cmd = ('%s%s%d%s') % (fn, '.setEditable(', n, ')')
            eval(cmd)

    def setComboBoxEnable(self, f, v):
        field_names = f
        value = v

        for fn in field_names:
            cmd = ('%s%s%s%s') % (fn, '.setEnabled(', v, ')')
            eval(cmd)

    def setTableEnable(self, t, v):
        tab_names = t
        value = v

        for tn in tab_names:
            cmd = ('%s%s%s%s') % (tn, '.setEnabled(', v, ')')
            eval(cmd)

    def testing(self, name_file, message):
        f = open(str(name_file), 'w')
        f.write(str(message))
        f.close()

    def on_pushButton_open_dir_pressed(self):
        HOME = os.environ['PYARCHINIT_HOME']
        path = '{}{}{}'.format(HOME, os.sep, "pyarchinit_PDF_folder")

        if platform.system() == "Windows":
            os.startfile(path)
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])

    def populate_photo_drawing_from_media(self):
        """
        Auto-populate photo and drawing fields for all Pottery records from media associations.
        Photos = images NOT starting with 'D_'
        Drawings = images starting with 'D_'
        """
        try:
            from sqlalchemy import text

            # Query to get all pottery records with their associated images
            query = """
                SELECT p.id_rep, mt.filename
                FROM pottery_table p
                INNER JOIN media_to_entity_table mte ON p.id_rep = mte.id_entity
                    AND mte.entity_type = 'CERAMICA'
                INNER JOIN media_thumb_table mt ON mte.id_media = mt.id_media
                ORDER BY p.id_rep, mt.filename
            """

            session = self.DB_MANAGER.Session()
            try:
                results = session.execute(text(query)).fetchall()
            finally:
                session.close()

            if not results:
                QMessageBox.warning(self, "Info",
                    "No media associations found for Pottery records.\n"
                    "Make sure images are tagged with entity_type='CERAMICA' in media_to_entity_table.")
                return

            # Group images by pottery record
            pottery_images = {}
            for row in results:
                id_rep, filename = row
                if id_rep not in pottery_images:
                    pottery_images[id_rep] = {'photos': [], 'drawings': []}

                if filename:
                    # Check if it's a drawing (starts with D_) or photo
                    if filename.startswith('D_'):
                        pottery_images[id_rep]['drawings'].append(filename)
                    else:
                        pottery_images[id_rep]['photos'].append(filename)

            # Update each pottery record
            updated_count = 0
            session = self.DB_MANAGER.Session()
            try:
                for id_rep, images in pottery_images.items():
                    photo_str = '; '.join(images['photos']) if images['photos'] else ''
                    drawing_str = '; '.join(images['drawings']) if images['drawings'] else ''

                    update_query = text("""
                        UPDATE pottery_table
                        SET photo = :photo, drawing = :drawing
                        WHERE id_rep = :id_rep
                    """)
                    session.execute(update_query, {
                        'photo': photo_str,
                        'drawing': drawing_str,
                        'id_rep': id_rep
                    })
                    updated_count += 1

                session.commit()
            except Exception as e:
                session.rollback()
                raise e
            finally:
                session.close()

            # Count how many have photos vs drawings
            photo_count = sum(1 for v in pottery_images.values() if v['photos'])
            drawing_count = sum(1 for v in pottery_images.values() if v['drawings'])

            QMessageBox.information(self, "Auto-Population Complete",
                f"Updated {updated_count} Pottery records:\n\n"
                f"Records with photos: {photo_count}\n"
                f"Records with drawings: {drawing_count}\n\n"
                "Click 'View All' to refresh the data.")

            # Refresh the current view
            self.charge_records()
            self.fill_fields()

        except Exception as e:
            QMessageBox.warning(self, "Error",
                f"Failed to auto-populate photo/drawing fields:\n{str(e)}")

    # =====================================================
    # STATISTICS TAB METHODS
    # =====================================================

    def setup_statistics_tab(self):
        """Setup the Statistics tab with all necessary widgets - RESPONSIVE LAYOUT"""
        try:
            # Find the Statistics tab (tab_8)
            stats_tab = self.tab_8

            # Clear existing layout if any
            if stats_tab.layout():
                old_layout = stats_tab.layout()
                while old_layout.count():
                    item = old_layout.takeAt(0)
                    if item.widget():
                        item.widget().setParent(None)
                QWidget().setLayout(old_layout)

            # Create main layout for the statistics tab
            main_layout = QVBoxLayout(stats_tab)
            main_layout.setContentsMargins(5, 5, 5, 5)
            main_layout.setSpacing(5)

            # Create a splitter for flexible layout - THIS IS KEY FOR RESPONSIVE DESIGN
            splitter = QSplitter(Qt.Horizontal)
            splitter.setChildrenCollapsible(False)

            # ========== LEFT PANEL - Tables and Controls ==========
            left_widget = QWidget()
            left_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            left_layout = QVBoxLayout(left_widget)
            left_layout.setContentsMargins(5, 5, 5, 5)
            left_layout.setSpacing(5)

            # Header with controls - wrap in a scroll area for small windows
            controls_widget = QWidget()
            controls_layout = QHBoxLayout(controls_widget)
            controls_layout.setContentsMargins(0, 0, 0, 0)

            # Refresh button
            self.pushButton_refresh_stats = QPushButton("Aggiorna Statistiche")
            self.pushButton_refresh_stats.setMaximumWidth(150)
            self.pushButton_refresh_stats.clicked.connect(self.calculate_statistics)
            controls_layout.addWidget(self.pushButton_refresh_stats)

            # Analysis type combo
            self.comboBox_stats_type = QComboBox()
            self.comboBox_stats_type.addItems([
                "── DISTRIBUZIONI ──",
                "Distribuzione per Forma (Shape)",
                "Distribuzione per Forma Specifica",
                "Distribuzione per Shape Specifica",
                "Distribuzione per Ware",
                "Distribuzione per Fabric",
                "Distribuzione per Area",
                "Distribuzione per US",
                "Distribuzione per Materiale",
                "Distribuzione per Trattamento Superficie",
                "Distribuzione per Decorazione Esterna",
                "Distribuzione per Decorazione Interna",
                "Distribuzione per Tipo Decorazione",
                "Distribuzione per Motivo Decorazione",
                "Distribuzione per Posizione Decorazione",
                "Distribuzione per Datazione",
                "── CROSSTAB FORME ──",
                "Crosstab Forma × US",
                "Crosstab Forma × Area",
                "Crosstab Forma Specifica × US",
                "Crosstab Forma Specifica × Area",
                "Crosstab Ware × US",
                "Crosstab Ware × Area",
                "── CROSSTAB DECORAZIONI ──",
                "Crosstab Deco Esterna × US",
                "Crosstab Deco Interna × US",
                "Crosstab Tipo Deco × US",
                "Crosstab Motivo Deco × US",
                "Crosstab Posizione Deco × US",
                "Crosstab Deco Esterna × Interna",
                "Crosstab Tipo Deco × Motivo",
                "Crosstab Deco Esterna × Motivo",
                "Crosstab Deco Interna × Motivo"
            ])
            self.comboBox_stats_type.setCurrentIndex(1)  # Skip separator

            # Chart type combo
            self.comboBox_chart_type = QComboBox()
            self.comboBox_chart_type.addItems([
                "Barre Verticali",
                "Barre Orizzontali",
                "Torta",
                "Linea",
                "Area",
                "Donut"
            ])
            controls_layout.addWidget(QLabel("Tipo Analisi:"))
            controls_layout.addWidget(self.comboBox_stats_type)
            self.comboBox_stats_type.currentIndexChanged.connect(self.on_stats_combo_changed)
            controls_layout.addWidget(QLabel("Tipo Grafico:"))
            controls_layout.addWidget(self.comboBox_chart_type)
            self.comboBox_chart_type.currentIndexChanged.connect(self.on_chart_type_changed)
            controls_layout.addStretch()

            left_layout.addWidget(controls_widget)

            # Create a vertical splitter for left panel sections (tables can resize)
            left_splitter = QSplitter(Qt.Vertical)
            left_splitter.setChildrenCollapsible(False)

            # Summary table - RESPONSIVE
            summary_group = QGroupBox("Riepilogo Generale")
            summary_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            summary_layout = QVBoxLayout(summary_group)
            summary_layout.setContentsMargins(5, 5, 5, 5)

            self.tableWidget_stats = QTableWidget()
            self.tableWidget_stats.setColumnCount(3)
            self.tableWidget_stats.setHorizontalHeaderLabels(["Categoria", "Conteggio", "Percentuale"])
            self.tableWidget_stats.horizontalHeader().setStretchLastSection(True)
            self.tableWidget_stats.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            self.tableWidget_stats.setAlternatingRowColors(True)
            self.tableWidget_stats.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self.tableWidget_stats.setMinimumHeight(100)
            summary_layout.addWidget(self.tableWidget_stats)

            left_splitter.addWidget(summary_group)

            # Measurements statistics table - RESPONSIVE
            measures_group = QGroupBox("Statistiche Misure")
            measures_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
            measures_layout = QVBoxLayout(measures_group)
            measures_layout.setContentsMargins(5, 5, 5, 5)

            self.tableWidget_measures = QTableWidget()
            self.tableWidget_measures.setColumnCount(5)
            self.tableWidget_measures.setHorizontalHeaderLabels(["Misura", "Min", "Max", "Media", "Mediana"])
            self.tableWidget_measures.horizontalHeader().setStretchLastSection(True)
            self.tableWidget_measures.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            self.tableWidget_measures.setAlternatingRowColors(True)
            self.tableWidget_measures.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
            self.tableWidget_measures.setMinimumHeight(80)
            self.tableWidget_measures.setMaximumHeight(200)
            measures_layout.addWidget(self.tableWidget_measures)

            left_splitter.addWidget(measures_group)

            # AI Report section - RESPONSIVE
            ai_group = QGroupBox("Report AI Descrittivo")
            ai_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            ai_layout = QVBoxLayout(ai_group)
            ai_layout.setContentsMargins(5, 5, 5, 5)

            ai_buttons_layout = QHBoxLayout()
            self.pushButton_generate_ai_report = QPushButton("Genera Report AI")
            self.pushButton_generate_ai_report.clicked.connect(self.generate_ai_report)
            ai_buttons_layout.addWidget(self.pushButton_generate_ai_report)

            self.pushButton_export_stats_pdf = QPushButton("Esporta PDF")
            self.pushButton_export_stats_pdf.clicked.connect(self.export_statistics_pdf)
            ai_buttons_layout.addWidget(self.pushButton_export_stats_pdf)
            ai_buttons_layout.addStretch()

            ai_layout.addLayout(ai_buttons_layout)

            self.textEdit_ai_report = QTextEdit()
            self.textEdit_ai_report.setPlaceholderText("Il report AI verrà visualizzato qui dopo la generazione...")
            self.textEdit_ai_report.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self.textEdit_ai_report.setMinimumHeight(80)
            ai_layout.addWidget(self.textEdit_ai_report)

            left_splitter.addWidget(ai_group)

            # Set proportions for left splitter
            left_splitter.setStretchFactor(0, 3)  # Summary table gets more space
            left_splitter.setStretchFactor(1, 1)  # Measures table smaller
            left_splitter.setStretchFactor(2, 2)  # AI report medium

            left_layout.addWidget(left_splitter)

            splitter.addWidget(left_widget)

            # ========== RIGHT PANEL - Chart (existing Mplwidget) ==========
            # Create a container for the chart that can resize
            right_widget = QWidget()
            right_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            right_layout = QVBoxLayout(right_widget)
            right_layout.setContentsMargins(5, 5, 5, 5)

            # Move existing Mplwidget to our layout
            if hasattr(self, 'widget') and self.widget is not None:
                self.widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                self.widget.setMinimumSize(300, 200)
                right_layout.addWidget(self.widget)

            splitter.addWidget(right_widget)

            # Set splitter proportions and stretch factors
            splitter.setStretchFactor(0, 1)  # Left panel
            splitter.setStretchFactor(1, 1)  # Right panel (chart)
            splitter.setSizes([450, 550])

            # Add splitter to main layout
            main_layout.addWidget(splitter)

            # Keep reference to pushButtonQuant if it exists
            if hasattr(self, 'pushButtonQuant'):
                self.pushButtonQuant.setParent(None)
                controls_layout.addWidget(self.pushButtonQuant)

            # Store splitter reference for programmatic resizing
            self.stats_splitter = splitter
            self.stats_left_splitter = left_splitter

            # Initialize statistics data storage
            self.stats_data = {}

        except Exception as e:
            print(f"Error setting up statistics tab: {e}")
            import traceback
            traceback.print_exc()

    def calculate_statistics(self):
        """Calculate all statistics for the current site"""
        try:
            sito = self.comboBox_sito.currentText()
            if not sito or sito == "Inserisci un valore":
                QMessageBox.warning(self, "Attenzione", "Seleziona prima un sito.", QMessageBox.Ok)
                return

            # Query all pottery records for the site
            search_dict = {'sito': f"'{sito}'"}
            records = self.DB_MANAGER.query_bool(search_dict, self.MAPPER_TABLE_CLASS)

            if not records:
                QMessageBox.information(self, "Info", "Nessun record trovato per questo sito.", QMessageBox.Ok)
                return

            # Store records for later use
            self.stats_records = records
            self.stats_data = {
                'sito': sito,
                'total': len(records),
                'records': records
            }

            # Update the statistics table based on current selection
            self.on_stats_combo_changed()

            # Calculate measurement statistics
            self.generate_measurement_stats()

        except Exception as e:
            QMessageBox.warning(self, "Errore", f"Errore nel calcolo statistiche: {e}", QMessageBox.Ok)

    def on_stats_combo_changed(self):
        """Handle statistics type combo box change"""
        if not hasattr(self, 'stats_records') or not self.stats_records:
            return

        stat_type = self.comboBox_stats_type.currentText()

        # Skip separator items
        if stat_type.startswith("──"):
            return

        # Map combo text to field name
        field_mapping = {
            "Distribuzione per Forma (Shape)": "form",
            "Distribuzione per Forma Specifica": "specific_form",
            "Distribuzione per Shape Specifica": "specific_shape",
            "Distribuzione per Ware": "ware",
            "Distribuzione per Fabric": "fabric",
            "Distribuzione per Area": "area",
            "Distribuzione per US": "us",
            "Distribuzione per Materiale": "material",
            "Distribuzione per Trattamento Superficie": "surf_trat",
            "Distribuzione per Decorazione Esterna": "exdeco",
            "Distribuzione per Decorazione Interna": "intdeco",
            "Distribuzione per Tipo Decorazione": "decoration_type",
            "Distribuzione per Motivo Decorazione": "decoration_motif",
            "Distribuzione per Posizione Decorazione": "decoration_position",
            "Distribuzione per Datazione": "datazione"
        }

        # Check if it's a crosstab analysis
        crosstab_mapping = {
            # Forme × Provenienza
            "Crosstab Forma × US": ("form", "us"),
            "Crosstab Forma × Area": ("form", "area"),
            "Crosstab Forma Specifica × US": ("specific_form", "us"),
            "Crosstab Forma Specifica × Area": ("specific_form", "area"),
            "Crosstab Ware × US": ("ware", "us"),
            "Crosstab Ware × Area": ("ware", "area"),
            # Decorazioni × US
            "Crosstab Deco Esterna × US": ("exdeco", "us"),
            "Crosstab Deco Interna × US": ("intdeco", "us"),
            "Crosstab Tipo Deco × US": ("decoration_type", "us"),
            "Crosstab Motivo Deco × US": ("decoration_motif", "us"),
            "Crosstab Posizione Deco × US": ("decoration_position", "us"),
            # Decorazioni × Decorazioni
            "Crosstab Deco Esterna × Interna": ("exdeco", "intdeco"),
            "Crosstab Tipo Deco × Motivo": ("decoration_type", "decoration_motif"),
            "Crosstab Deco Esterna × Motivo": ("exdeco", "decoration_motif"),
            "Crosstab Deco Interna × Motivo": ("intdeco", "decoration_motif")
        }

        if stat_type in crosstab_mapping:
            field1, field2 = crosstab_mapping[stat_type]
            self.generate_crosstab_stats(field1, field2)
        else:
            field = field_mapping.get(stat_type, "form")
            self.generate_category_stats(field)

    def on_chart_type_changed(self):
        """Handle chart type combo box change - refresh current chart"""
        # Check if we have crosstab data
        if hasattr(self, 'current_crosstab_data') and self.current_crosstab_data:
            self.update_crosstab_chart_display()
        # Check if we have regular chart data
        elif hasattr(self, 'current_chart_data') and self.current_chart_data:
            self.update_chart_display()

    def generate_category_stats(self, field):
        """Generate statistics for a specific category field"""
        if not hasattr(self, 'stats_records') or not self.stats_records:
            return

        try:
            # Count occurrences
            counts = {}
            total = len(self.stats_records)

            for record in self.stats_records:
                value = getattr(record, field, None)
                if value is None or str(value).strip() == '' or str(value) == 'None':
                    value = 'N/D'
                else:
                    value = str(value)

                counts[value] = counts.get(value, 0) + 1

            # Sort by count descending
            sorted_counts = sorted(counts.items(), key=lambda x: x[1], reverse=True)

            # Calculate additional statistics
            num_categories = len(sorted_counts)
            values_list = [cnt for _, cnt in sorted_counts]
            mean_count = sum(values_list) / len(values_list) if values_list else 0
            max_count = max(values_list) if values_list else 0
            min_count = min(values_list) if values_list else 0

            # Expand table with more columns
            self.tableWidget_stats.setColumnCount(5)
            self.tableWidget_stats.setHorizontalHeaderLabels([
                "Categoria", "Conteggio", "Percentuale", "% Cumulativa", "Rank"
            ])

            # Update table
            self.tableWidget_stats.setRowCount(len(sorted_counts) + 2)  # +2 for total and summary

            # Add total row first
            self.tableWidget_stats.setItem(0, 0, QTableWidgetItem("TOTALE"))
            self.tableWidget_stats.setItem(0, 1, QTableWidgetItem(str(total)))
            self.tableWidget_stats.setItem(0, 2, QTableWidgetItem("100%"))
            self.tableWidget_stats.setItem(0, 3, QTableWidgetItem("-"))
            self.tableWidget_stats.setItem(0, 4, QTableWidgetItem(f"{num_categories} categorie"))

            # Make total row bold
            for col in range(5):
                item = self.tableWidget_stats.item(0, col)
                if item:
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)

            # Add category rows with cumulative percentage
            cumulative = 0
            for i, (category, count) in enumerate(sorted_counts, 1):
                percentage = (count / total) * 100 if total > 0 else 0
                cumulative += percentage
                self.tableWidget_stats.setItem(i, 0, QTableWidgetItem(category))
                self.tableWidget_stats.setItem(i, 1, QTableWidgetItem(str(count)))
                self.tableWidget_stats.setItem(i, 2, QTableWidgetItem(f"{percentage:.1f}%"))
                self.tableWidget_stats.setItem(i, 3, QTableWidgetItem(f"{cumulative:.1f}%"))
                self.tableWidget_stats.setItem(i, 4, QTableWidgetItem(f"#{i}"))

            # Add summary row
            summary_row = len(sorted_counts) + 1
            self.tableWidget_stats.setItem(summary_row, 0, QTableWidgetItem("STATISTICHE"))
            self.tableWidget_stats.setItem(summary_row, 1, QTableWidgetItem(f"Media: {mean_count:.1f}"))
            self.tableWidget_stats.setItem(summary_row, 2, QTableWidgetItem(f"Max: {max_count}"))
            self.tableWidget_stats.setItem(summary_row, 3, QTableWidgetItem(f"Min: {min_count}"))
            self.tableWidget_stats.setItem(summary_row, 4, QTableWidgetItem(""))

            # Make summary row italic
            for col in range(5):
                item = self.tableWidget_stats.item(summary_row, col)
                if item:
                    font = item.font()
                    font.setItalic(True)
                    item.setFont(font)

            self.tableWidget_stats.resizeColumnsToContents()

            # Store chart data for dynamic updates
            chart_data = [(cat, cnt) for cat, cnt in sorted_counts[:15]]  # Top 15 for readability
            self.current_chart_data = chart_data
            self.current_chart_title = f"Distribuzione per {field.replace('_', ' ').title()}"
            self.current_crosstab_data = None  # Reset crosstab data

            # Update chart using selected chart type
            if chart_data:
                self.update_chart_display()

            # Store for AI report
            self.stats_data[field] = sorted_counts

        except Exception as e:
            print(f"Error generating category stats: {e}")
            import traceback
            traceback.print_exc()

    def update_chart_display(self):
        """Update chart based on selected chart type"""
        if not hasattr(self, 'current_chart_data') or not self.current_chart_data:
            return

        chart_type = self.comboBox_chart_type.currentText() if hasattr(self, 'comboBox_chart_type') else "Barre Verticali"

        chart_type_map = {
            "Barre Verticali": "vbar",
            "Barre Orizzontali": "hbar",
            "Torta": "pie",
            "Linea": "line",
            "Area": "area",
            "Donut": "donut"
        }

        chart_code = chart_type_map.get(chart_type, "vbar")
        self.plot_advanced_chart(self.current_chart_data, self.current_chart_title, chart_code)

    def plot_advanced_chart(self, d, title, chart_type='vbar'):
        """Generate various chart types"""
        if not d or not hasattr(self, 'widget') or self.widget is None:
            return

        try:
            data_diz = {item[0]: item[1] for item in d}
            labels = list(data_diz.keys())
            values = list(data_diz.values())

            self.widget.canvas.ax.clear()

            colors = ['#3498db', '#e74c3c', '#2ecc71', '#f39c12', '#9b59b6',
                     '#1abc9c', '#e67e22', '#34495e', '#16a085', '#c0392b',
                     '#27ae60', '#8e44ad', '#2980b9', '#d35400', '#7f8c8d']

            if chart_type == 'vbar':
                x = list(range(len(values)))
                bars = self.widget.canvas.ax.bar(x, height=values, width=0.6, alpha=0.8, color=colors[:len(values)])
                self.widget.canvas.ax.set_ylabel('Quantità')
                # Add value labels
                for bar, val in zip(bars, values):
                    height = bar.get_height()
                    self.widget.canvas.ax.text(bar.get_x() + bar.get_width()/2., height,
                                              f'{val}', ha='center', va='bottom', fontsize=8)
                # Rotate labels
                for i, label in enumerate(labels):
                    self.widget.canvas.ax.text(i, -max(values)*0.05, label[:12],
                                              ha='center', va='top', fontsize=7, rotation=45)
                self.widget.canvas.ax.set_xticks([])

            elif chart_type == 'hbar':
                y = list(range(len(values)))
                bars = self.widget.canvas.ax.barh(y, values, height=0.6, alpha=0.8, color=colors[:len(values)])
                self.widget.canvas.ax.set_yticks(y)
                self.widget.canvas.ax.set_yticklabels([l[:15] for l in labels], fontsize=8)
                self.widget.canvas.ax.set_xlabel('Quantità')
                # Add value labels
                for bar, val in zip(bars, values):
                    width = bar.get_width()
                    self.widget.canvas.ax.text(width, bar.get_y() + bar.get_height()/2.,
                                              f' {val}', ha='left', va='center', fontsize=8)

            elif chart_type == 'pie':
                wedges, texts, autotexts = self.widget.canvas.ax.pie(
                    values, labels=[l[:10] for l in labels], autopct='%1.1f%%',
                    startangle=90, colors=colors[:len(values)], textprops={'fontsize': 8}
                )
                self.widget.canvas.ax.axis('equal')

            elif chart_type == 'donut':
                wedges, texts, autotexts = self.widget.canvas.ax.pie(
                    values, labels=[l[:10] for l in labels], autopct='%1.1f%%',
                    startangle=90, colors=colors[:len(values)], textprops={'fontsize': 8},
                    wedgeprops=dict(width=0.5)
                )
                self.widget.canvas.ax.axis('equal')

            elif chart_type == 'line':
                x = list(range(len(values)))
                self.widget.canvas.ax.plot(x, values, 'o-', linewidth=2, markersize=8, color='#3498db')
                self.widget.canvas.ax.fill_between(x, values, alpha=0.3, color='#3498db')
                self.widget.canvas.ax.set_xticks(x)
                self.widget.canvas.ax.set_xticklabels([l[:10] for l in labels], rotation=45, ha='right', fontsize=7)
                self.widget.canvas.ax.set_ylabel('Quantità')
                self.widget.canvas.ax.grid(True, alpha=0.3)

            elif chart_type == 'area':
                x = list(range(len(values)))
                self.widget.canvas.ax.fill_between(x, values, alpha=0.6, color='#3498db')
                self.widget.canvas.ax.plot(x, values, 'o-', linewidth=2, markersize=6, color='#2980b9')
                self.widget.canvas.ax.set_xticks(x)
                self.widget.canvas.ax.set_xticklabels([l[:10] for l in labels], rotation=45, ha='right', fontsize=7)
                self.widget.canvas.ax.set_ylabel('Quantità')
                self.widget.canvas.ax.grid(True, alpha=0.3)

            self.widget.canvas.ax.set_title(title, fontsize=10, fontweight='bold')
            self.widget.canvas.figure.tight_layout()
            self.widget.canvas.draw()

        except Exception as e:
            print(f"Error plotting advanced chart: {e}")
            import traceback
            traceback.print_exc()

    def generate_measurement_stats(self):
        """Generate statistics for measurement fields"""
        if not hasattr(self, 'stats_records') or not self.stats_records:
            return

        try:
            measurements = {
                'Diametro Max': 'diametro_max',
                'Diametro Orlo': 'diametro_rim',
                'Diametro Fondo': 'diametro_bottom',
                'Altezza Totale': 'diametro_height',
                'Altezza Conservata': 'diametro_preserved'
            }

            self.tableWidget_measures.setRowCount(len(measurements))
            self.stats_data['measurements'] = {}

            for row, (label, field) in enumerate(measurements.items()):
                values = []
                for record in self.stats_records:
                    value = getattr(record, field, None)
                    if value is not None and str(value) not in ['', 'None', 'NULL']:
                        try:
                            values.append(float(value))
                        except (ValueError, TypeError):
                            pass

                self.tableWidget_measures.setItem(row, 0, QTableWidgetItem(label))

                if values:
                    min_val = min(values)
                    max_val = max(values)
                    mean_val = sum(values) / len(values)
                    sorted_vals = sorted(values)
                    median_val = sorted_vals[len(sorted_vals) // 2]

                    self.tableWidget_measures.setItem(row, 1, QTableWidgetItem(f"{min_val:.2f}"))
                    self.tableWidget_measures.setItem(row, 2, QTableWidgetItem(f"{max_val:.2f}"))
                    self.tableWidget_measures.setItem(row, 3, QTableWidgetItem(f"{mean_val:.2f}"))
                    self.tableWidget_measures.setItem(row, 4, QTableWidgetItem(f"{median_val:.2f}"))

                    self.stats_data['measurements'][label] = {
                        'min': min_val, 'max': max_val,
                        'mean': mean_val, 'median': median_val,
                        'count': len(values)
                    }
                else:
                    for col in range(1, 5):
                        self.tableWidget_measures.setItem(row, col, QTableWidgetItem("N/D"))

            self.tableWidget_measures.resizeColumnsToContents()

        except Exception as e:
            print(f"Error generating measurement stats: {e}")

    def generate_crosstab_stats(self, field1, field2):
        """Generate cross-tabulation statistics for two fields"""
        if not hasattr(self, 'stats_records') or not self.stats_records:
            return

        try:
            total = len(self.stats_records)

            # Build crosstab data
            crosstab = {}
            field1_values = set()
            field2_values = set()

            for record in self.stats_records:
                val1 = getattr(record, field1, None)
                val2 = getattr(record, field2, None)

                if val1 is None or str(val1).strip() == '' or str(val1) == 'None':
                    val1 = 'N/D'
                else:
                    val1 = str(val1)

                if val2 is None or str(val2).strip() == '' or str(val2) == 'None':
                    val2 = 'N/D'
                else:
                    val2 = str(val2)

                field1_values.add(val1)
                field2_values.add(val2)

                key = (val1, val2)
                crosstab[key] = crosstab.get(key, 0) + 1

            # Sort values by frequency
            field1_counts = {}
            field2_counts = {}
            for (v1, v2), cnt in crosstab.items():
                field1_counts[v1] = field1_counts.get(v1, 0) + cnt
                field2_counts[v2] = field2_counts.get(v2, 0) + cnt

            sorted_field1 = sorted(field1_counts.keys(), key=lambda x: field1_counts[x], reverse=True)[:15]
            sorted_field2 = sorted(field2_counts.keys(), key=lambda x: field2_counts[x], reverse=True)[:15]

            # Update table as a crosstab matrix
            self.tableWidget_stats.setColumnCount(len(sorted_field2) + 2)  # +2 for row label and total
            headers = [field1.replace('_', ' ').title()] + [f"{v2[:10]}" for v2 in sorted_field2] + ["Totale"]
            self.tableWidget_stats.setHorizontalHeaderLabels(headers)

            self.tableWidget_stats.setRowCount(len(sorted_field1) + 1)  # +1 for column totals

            # Fill data
            for row_idx, val1 in enumerate(sorted_field1):
                self.tableWidget_stats.setItem(row_idx, 0, QTableWidgetItem(val1[:15]))
                row_total = 0
                for col_idx, val2 in enumerate(sorted_field2):
                    count = crosstab.get((val1, val2), 0)
                    row_total += count
                    item = QTableWidgetItem(str(count) if count > 0 else "-")
                    if count > 0:
                        # Color intensity based on value
                        intensity = min(255, 200 + int(55 * (count / max(crosstab.values()))))
                        item.setBackground(QColor(200, 230, intensity))
                    self.tableWidget_stats.setItem(row_idx, col_idx + 1, item)
                # Row total
                total_item = QTableWidgetItem(str(row_total))
                total_item.setFont(self._get_bold_font())
                self.tableWidget_stats.setItem(row_idx, len(sorted_field2) + 1, total_item)

            # Column totals row
            total_row = len(sorted_field1)
            self.tableWidget_stats.setItem(total_row, 0, QTableWidgetItem("Totale"))
            self.tableWidget_stats.item(total_row, 0).setFont(self._get_bold_font())
            grand_total = 0
            for col_idx, val2 in enumerate(sorted_field2):
                col_total = sum(crosstab.get((v1, val2), 0) for v1 in sorted_field1)
                grand_total += col_total
                item = QTableWidgetItem(str(col_total))
                item.setFont(self._get_bold_font())
                self.tableWidget_stats.setItem(total_row, col_idx + 1, item)

            # Grand total
            item = QTableWidgetItem(str(grand_total))
            item.setFont(self._get_bold_font())
            self.tableWidget_stats.setItem(total_row, len(sorted_field2) + 1, item)

            self.tableWidget_stats.resizeColumnsToContents()

            # Plot heatmap-style chart
            self.plot_crosstab_chart(crosstab, sorted_field1, sorted_field2, field1, field2)

            # Store for export
            self.stats_data['crosstab'] = {
                'field1': field1,
                'field2': field2,
                'data': crosstab,
                'rows': sorted_field1,
                'cols': sorted_field2
            }

        except Exception as e:
            print(f"Error generating crosstab stats: {e}")
            import traceback
            traceback.print_exc()

    def _get_bold_font(self):
        """Helper to get bold font"""
        from qgis.PyQt.QtGui import QFont
        font = QFont()
        font.setBold(True)
        return font

    def plot_crosstab_chart(self, crosstab, rows, cols, field1, field2):
        """Plot a stacked bar chart for crosstab data"""
        if not hasattr(self, 'widget') or self.widget is None:
            return

        try:
            import numpy as np

            self.widget.canvas.ax.clear()

            # Prepare data matrix
            data_matrix = []
            for col in cols[:10]:  # Limit columns
                col_data = [crosstab.get((row, col), 0) for row in rows[:10]]
                data_matrix.append(col_data)

            data_matrix = np.array(data_matrix)
            x = np.arange(len(rows[:10]))
            width = 0.6

            colors = ['#3498db', '#e74c3c', '#2ecc71', '#f39c12', '#9b59b6',
                     '#1abc9c', '#e67e22', '#34495e', '#16a085', '#c0392b']

            # Plot stacked bars
            bottom = np.zeros(len(rows[:10]))
            for i, col in enumerate(cols[:10]):
                values = data_matrix[i]
                self.widget.canvas.ax.bar(x, values, width, bottom=bottom,
                                         label=col[:12], color=colors[i % len(colors)], alpha=0.8)
                bottom += values

            self.widget.canvas.ax.set_ylabel('Quantità')
            self.widget.canvas.ax.set_xticks(x)
            self.widget.canvas.ax.set_xticklabels([r[:10] for r in rows[:10]], rotation=45, ha='right', fontsize=7)
            self.widget.canvas.ax.set_title(f'Crosstab: {field1.title()} × {field2.title()}', fontsize=10, fontweight='bold')

            # Legend
            self.widget.canvas.ax.legend(loc='upper right', fontsize=7, ncol=2)

            self.widget.canvas.figure.tight_layout()
            self.widget.canvas.draw()

            # Store current chart info for dynamic updates
            self.current_chart_data = None  # Reset regular chart data
            self.current_crosstab_data = {
                'crosstab': crosstab,
                'rows': rows,
                'cols': cols,
                'field1': field1,
                'field2': field2
            }
            self.current_chart_title = f'Crosstab: {field1.title()} × {field2.title()}'

        except Exception as e:
            print(f"Error plotting crosstab chart: {e}")
            import traceback
            traceback.print_exc()

    def update_crosstab_chart_display(self):
        """Update crosstab chart based on selected chart type"""
        if not hasattr(self, 'current_crosstab_data') or not self.current_crosstab_data:
            return

        data = self.current_crosstab_data
        chart_type = self.comboBox_chart_type.currentText() if hasattr(self, 'comboBox_chart_type') else "Barre Verticali"

        chart_type_map = {
            "Barre Verticali": "stacked",
            "Barre Orizzontali": "hstacked",
            "Torta": "pie_top",
            "Linea": "line_multi",
            "Area": "area_stacked",
            "Donut": "heatmap"
        }

        chart_code = chart_type_map.get(chart_type, "stacked")
        self.plot_crosstab_advanced(data['crosstab'], data['rows'], data['cols'],
                                    data['field1'], data['field2'], chart_code)

    def plot_crosstab_advanced(self, crosstab, rows, cols, field1, field2, chart_type='stacked'):
        """Plot crosstab with various chart types"""
        if not hasattr(self, 'widget') or self.widget is None:
            return

        try:
            import numpy as np

            self.widget.canvas.ax.clear()

            colors = ['#3498db', '#e74c3c', '#2ecc71', '#f39c12', '#9b59b6',
                     '#1abc9c', '#e67e22', '#34495e', '#16a085', '#c0392b']

            rows_limited = rows[:10]
            cols_limited = cols[:10]

            # Prepare data matrix
            data_matrix = []
            for col in cols_limited:
                col_data = [crosstab.get((row, col), 0) for row in rows_limited]
                data_matrix.append(col_data)

            data_matrix = np.array(data_matrix)
            x = np.arange(len(rows_limited))
            width = 0.6

            if chart_type == 'stacked':
                # Stacked vertical bars
                bottom = np.zeros(len(rows_limited))
                for i, col in enumerate(cols_limited):
                    values = data_matrix[i]
                    self.widget.canvas.ax.bar(x, values, width, bottom=bottom,
                                             label=col[:12], color=colors[i % len(colors)], alpha=0.8)
                    bottom += values
                self.widget.canvas.ax.set_ylabel('Quantità')
                self.widget.canvas.ax.set_xticks(x)
                self.widget.canvas.ax.set_xticklabels([r[:10] for r in rows_limited], rotation=45, ha='right', fontsize=7)
                self.widget.canvas.ax.legend(loc='upper right', fontsize=6, ncol=2)

            elif chart_type == 'hstacked':
                # Stacked horizontal bars
                y = np.arange(len(rows_limited))
                left = np.zeros(len(rows_limited))
                for i, col in enumerate(cols_limited):
                    values = data_matrix[i]
                    self.widget.canvas.ax.barh(y, values, height=0.6, left=left,
                                              label=col[:12], color=colors[i % len(colors)], alpha=0.8)
                    left += values
                self.widget.canvas.ax.set_xlabel('Quantità')
                self.widget.canvas.ax.set_yticks(y)
                self.widget.canvas.ax.set_yticklabels([r[:12] for r in rows_limited], fontsize=7)
                self.widget.canvas.ax.legend(loc='upper right', fontsize=6, ncol=2)

            elif chart_type == 'pie_top':
                # Pie chart of row totals
                row_totals = [sum(crosstab.get((row, col), 0) for col in cols_limited) for row in rows_limited]
                self.widget.canvas.ax.pie(row_totals, labels=[r[:10] for r in rows_limited],
                                         autopct='%1.1f%%', colors=colors[:len(rows_limited)],
                                         textprops={'fontsize': 7})
                self.widget.canvas.ax.axis('equal')

            elif chart_type == 'line_multi':
                # Multi-line chart
                for i, col in enumerate(cols_limited[:5]):  # Limit to 5 lines
                    values = data_matrix[i]
                    self.widget.canvas.ax.plot(x, values, 'o-', linewidth=2, markersize=6,
                                              label=col[:12], color=colors[i % len(colors)])
                self.widget.canvas.ax.set_ylabel('Quantità')
                self.widget.canvas.ax.set_xticks(x)
                self.widget.canvas.ax.set_xticklabels([r[:10] for r in rows_limited], rotation=45, ha='right', fontsize=7)
                self.widget.canvas.ax.legend(loc='upper right', fontsize=6)
                self.widget.canvas.ax.grid(True, alpha=0.3)

            elif chart_type == 'area_stacked':
                # Stacked area chart
                cumulative = np.zeros(len(rows_limited))
                for i, col in enumerate(cols_limited):
                    values = data_matrix[i]
                    self.widget.canvas.ax.fill_between(x, cumulative, cumulative + values,
                                                       label=col[:12], color=colors[i % len(colors)], alpha=0.7)
                    cumulative += values
                self.widget.canvas.ax.set_ylabel('Quantità')
                self.widget.canvas.ax.set_xticks(x)
                self.widget.canvas.ax.set_xticklabels([r[:10] for r in rows_limited], rotation=45, ha='right', fontsize=7)
                self.widget.canvas.ax.legend(loc='upper right', fontsize=6, ncol=2)

            elif chart_type == 'heatmap':
                # Heatmap style
                im = self.widget.canvas.ax.imshow(data_matrix, cmap='YlOrRd', aspect='auto')
                self.widget.canvas.ax.set_xticks(np.arange(len(rows_limited)))
                self.widget.canvas.ax.set_yticks(np.arange(len(cols_limited)))
                self.widget.canvas.ax.set_xticklabels([r[:8] for r in rows_limited], rotation=45, ha='right', fontsize=7)
                self.widget.canvas.ax.set_yticklabels([c[:10] for c in cols_limited], fontsize=7)
                # Add text annotations
                for i in range(len(cols_limited)):
                    for j in range(len(rows_limited)):
                        val = data_matrix[i, j]
                        if val > 0:
                            self.widget.canvas.ax.text(j, i, str(int(val)), ha='center', va='center',
                                                       fontsize=6, color='black' if val < data_matrix.max()/2 else 'white')
                self.widget.canvas.figure.colorbar(im, ax=self.widget.canvas.ax, shrink=0.8)

            self.widget.canvas.ax.set_title(f'Crosstab: {field1.title()} × {field2.title()}', fontsize=10, fontweight='bold')
            self.widget.canvas.figure.tight_layout()
            self.widget.canvas.draw()

        except Exception as e:
            print(f"Error plotting crosstab advanced: {e}")
            import traceback
            traceback.print_exc()

    def torta_chart(self, d, t, yl):
        """Generate a pie chart"""
        self.data_list = d
        self.title = t
        self.ylabel = yl

        if isinstance(self.data_list, list) and len(self.data_list) > 0:
            data_diz = {item[0]: item[1] for item in self.data_list}
            labels, values = zip(*data_diz.items())

            self.widget.canvas.ax.clear()
            self.widget.canvas.ax.pie(values, labels=labels, autopct='%1.1f%%')
            self.widget.canvas.ax.axis('equal')
            self.widget.canvas.ax.set_title(self.title)
            self.widget.canvas.draw()

    def calculate_provenance_stats(self):
        """Calculate provenance statistics (area, US, sector distributions)"""
        if not hasattr(self, 'stats_records') or not self.stats_records:
            return

        try:
            # Calculate stats for area, us, sector
            for field in ['area', 'us', 'sector']:
                counts = {}
                for record in self.stats_records:
                    value = getattr(record, field, None)
                    if value is None or str(value).strip() == '' or str(value) == 'None':
                        value = 'N/D'
                    else:
                        value = str(value)
                    counts[value] = counts.get(value, 0) + 1

                sorted_counts = sorted(counts.items(), key=lambda x: x[1], reverse=True)
                self.stats_data[field] = sorted_counts

        except Exception as e:
            print(f"Error calculating provenance stats: {e}")

    def build_statistics_prompt(self):
        """Build a prompt for AI report generation with language support"""
        if not hasattr(self, 'stats_data') or not self.stats_data:
            return None

        sito = self.stats_data.get('sito', 'Sconosciuto')
        total = self.stats_data.get('total', 0)

        # Calculate provenance stats if not already done
        self.calculate_provenance_stats()

        # Build data section (language-neutral)
        data_section = f"""
SITE: {sito}
TOTAL POTTERY FRAGMENTS: {total}

"""

        # Add category distributions
        categories_to_report = ['form', 'specific_form', 'specific_shape', 'ware', 'fabric', 'material']
        category_labels = {
            'form': 'FORM/SHAPE',
            'specific_form': 'SPECIFIC FORM',
            'specific_shape': 'SPECIFIC SHAPE',
            'ware': 'WARE TYPE',
            'fabric': 'FABRIC',
            'material': 'MATERIAL'
        }

        for cat in categories_to_report:
            if cat in self.stats_data:
                data = self.stats_data[cat]
                data_section += f"\n{category_labels.get(cat, cat.upper())} DISTRIBUTION:\n"
                for item, count in data[:10]:
                    percentage = (count / total) * 100 if total > 0 else 0
                    data_section += f"  - {item}: {count} ({percentage:.1f}%)\n"

        # Add provenance distributions
        data_section += "\n--- PROVENANCE DATA ---\n"

        provenance_fields = ['area', 'us', 'sector']
        provenance_labels = {
            'area': 'AREA',
            'us': 'STRATIGRAPHIC UNIT (US)',
            'sector': 'SECTOR'
        }

        for field in provenance_fields:
            if field in self.stats_data:
                data = self.stats_data[field]
                data_section += f"\n{provenance_labels.get(field, field.upper())} DISTRIBUTION:\n"
                for item, count in data[:15]:  # Show more for provenance
                    percentage = (count / total) * 100 if total > 0 else 0
                    data_section += f"  - {item}: {count} ({percentage:.1f}%)\n"

        # Add measurement statistics
        if 'measurements' in self.stats_data:
            data_section += "\nMEASUREMENT STATISTICS:\n"
            for label, stats in self.stats_data['measurements'].items():
                if isinstance(stats, dict):
                    data_section += f"  - {label}: min={stats['min']:.2f}, max={stats['max']:.2f}, mean={stats['mean']:.2f}, median={stats['median']:.2f} (n={stats['count']})\n"

        # Language-specific instructions
        if self.L == 'it':
            instructions = """
ISTRUZIONI:
Genera un report descrittivo in ITALIANO (circa 400-600 parole) che:
1. Riassuma le caratteristiche principali del complesso ceramico
2. Evidenzi le categorie predominanti (forme, classi ceramiche, impasti)
3. Analizzi la distribuzione spaziale per area, US e settore, identificando concentrazioni significative
4. Discuta le dimensioni medie e la variabilità dei reperti
5. Fornisca considerazioni cronologiche e tipologiche se possibile dai dati
6. Identifichi eventuali pattern significativi nella distribuzione spaziale
7. Proponga interpretazioni archeologiche preliminari basate sulla provenienza

Il tono deve essere professionale e scientifico, adatto a una pubblicazione archeologica.
Usa paragrafi ben strutturati con titoli per ogni sezione.
"""
        elif self.L == 'de':
            instructions = """
ANWEISUNGEN:
Erstellen Sie einen beschreibenden Bericht auf DEUTSCH (ca. 400-600 Wörter), der:
1. Die Hauptmerkmale des Keramikkomplexes zusammenfasst
2. Die vorherrschenden Kategorien hervorhebt (Formen, Keramikklassen, Magerung)
3. Die räumliche Verteilung nach Areal, SE und Sektor analysiert und signifikante Konzentrationen identifiziert
4. Die durchschnittlichen Abmessungen und die Variabilität der Funde diskutiert
5. Chronologische und typologische Überlegungen liefert, wenn möglich aus den Daten
6. Signifikante Muster in der räumlichen Verteilung identifiziert
7. Vorläufige archäologische Interpretationen basierend auf der Herkunft vorschlägt

Der Ton sollte professionell und wissenschaftlich sein, geeignet für eine archäologische Publikation.
Verwenden Sie gut strukturierte Absätze mit Überschriften für jeden Abschnitt.
"""
        else:  # English default
            instructions = """
INSTRUCTIONS:
Generate a descriptive report in ENGLISH (approximately 400-600 words) that:
1. Summarizes the main characteristics of the pottery assemblage
2. Highlights the predominant categories (forms, ware types, fabrics)
3. Analyzes the spatial distribution by area, SU and sector, identifying significant concentrations
4. Discusses the average dimensions and variability of the finds
5. Provides chronological and typological considerations if possible from the data
6. Identifies any significant patterns in the spatial distribution
7. Proposes preliminary archaeological interpretations based on provenance

The tone should be professional and scientific, suitable for an archaeological publication.
Use well-structured paragraphs with headings for each section.
"""

        prompt = f"""Analyze the following pottery data from the archaeological site:

{data_section}

{instructions}
"""

        return prompt

    def get_openai_api_key(self):
        """Get OpenAI API key from file or prompt user to enter it"""
        HOME = os.environ.get('PYARCHINIT_HOME', os.path.join(os.path.expanduser('~'), 'pyarchinit'))
        BIN = os.path.join(HOME, 'bin')

        # Ensure bin directory exists
        if not os.path.exists(BIN):
            os.makedirs(BIN, exist_ok=True)

        path_key = os.path.join(BIN, 'gpt_api_key.txt')
        api_key = ""

        if os.path.exists(path_key):
            # Read API key from file
            with open(path_key, 'r') as f:
                api_key = f.read().strip()

            if api_key:
                return api_key
            else:
                # File exists but is empty, ask user
                api_key, ok = QInputDialog.getText(
                    self, 'OpenAI API Key',
                    'Il file gpt_api_key.txt è vuoto.\nInserisci la tua OpenAI API key:'
                )
                if ok and api_key:
                    with open(path_key, 'w') as f:
                        f.write(api_key)
                    return api_key
        else:
            # File doesn't exist, ask user
            api_key, ok = QInputDialog.getText(
                self, 'OpenAI API Key',
                'API key non trovata.\nInserisci la tua OpenAI API key:\n\n'
                '(Verrà salvata in ~/pyarchinit/bin/gpt_api_key.txt)'
            )
            if ok and api_key:
                with open(path_key, 'w') as f:
                    f.write(api_key)
                return api_key

        return None

    def generate_ai_report(self):
        """Generate AI-based descriptive report with streaming"""
        if not hasattr(self, 'stats_data') or not self.stats_data:
            if self.L == 'it':
                msg = "Calcola prima le statistiche cliccando su 'Aggiorna Statistiche'."
            elif self.L == 'de':
                msg = "Berechnen Sie zuerst die Statistiken, indem Sie auf 'Statistiken aktualisieren' klicken."
            else:
                msg = "Calculate statistics first by clicking 'Refresh Statistics'."
            QMessageBox.warning(self, "Warning", msg, QMessageBox.Ok)
            return

        try:
            # Get API key from file or prompt user
            api_key = self.get_openai_api_key()

            if not api_key:
                if self.L == 'it':
                    msg = "È necessaria una API key di OpenAI per generare il report."
                elif self.L == 'de':
                    msg = "Für die Berichterstellung ist ein OpenAI API-Schlüssel erforderlich."
                else:
                    msg = "An OpenAI API key is required to generate the report."
                QMessageBox.warning(self, "API Key", msg, QMessageBox.Ok)
                return

            # Model to use
            model = 'gpt-4o-mini'

            # Build prompt
            prompt = self.build_statistics_prompt()
            if not prompt:
                QMessageBox.warning(self, "Error", "Cannot build prompt.", QMessageBox.Ok)
                return

            # Show progress message
            if self.L == 'it':
                progress_msg = "Generazione report in corso..."
            elif self.L == 'de':
                progress_msg = "Bericht wird erstellt..."
            else:
                progress_msg = "Generating report..."

            self.textEdit_ai_report.setText(progress_msg)
            QApplication.processEvents()

            # Check internet connection
            from ..modules.utility.report_generator import ReportGenerator
            if not ReportGenerator.is_connected():
                if self.L == 'it':
                    msg = "Connessione internet non disponibile."
                elif self.L == 'de':
                    msg = "Keine Internetverbindung verfügbar."
                else:
                    msg = "No internet connection available."
                QMessageBox.warning(self, "Connection", msg, QMessageBox.Ok)
                self.textEdit_ai_report.setText("")
                return

            # Generate report with streaming
            try:
                from openai import OpenAI
                client = OpenAI(api_key=api_key)

                response = client.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": prompt}],
                    stream=True,
                )

                # Stream the response
                self.textEdit_ai_report.clear()
                full_report = ""

                for chunk in response:
                    if chunk.choices[0].delta.content is not None:
                        content = chunk.choices[0].delta.content
                        full_report += content
                        self.textEdit_ai_report.setText(full_report)
                        QApplication.processEvents()

                if full_report:
                    self.stats_data['ai_report'] = full_report
                else:
                    if self.L == 'it':
                        self.textEdit_ai_report.setText("Errore nella generazione del report.")
                    elif self.L == 'de':
                        self.textEdit_ai_report.setText("Fehler bei der Berichterstellung.")
                    else:
                        self.textEdit_ai_report.setText("Error generating report.")

            except ImportError:
                # Fallback to non-streaming if openai not available
                generator = ReportGenerator()
                report = generator.generate_report_with_openai(prompt, model, api_key)
                if report:
                    self.textEdit_ai_report.setText(report)
                    self.stats_data['ai_report'] = report

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error generating AI report: {e}", QMessageBox.Ok)
            self.textEdit_ai_report.setText("")

    def generate_chart_image(self, data, title, chart_type='bar'):
        """Generate a chart image and return the path"""
        import matplotlib
        matplotlib.use('Agg')  # Non-interactive backend
        import matplotlib.pyplot as plt
        from io import BytesIO

        try:
            fig, ax = plt.subplots(figsize=(8, 4))

            if not data:
                return None

            labels = [str(item[0])[:20] for item in data[:10]]  # Truncate labels
            values = [item[1] for item in data[:10]]

            if chart_type == 'bar':
                bars = ax.bar(range(len(values)), values, color='steelblue', alpha=0.7)
                ax.set_xticks(range(len(values)))
                ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=8)
                ax.set_ylabel('Count')
            elif chart_type == 'pie':
                ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
                ax.axis('equal')

            ax.set_title(title, fontsize=10, fontweight='bold')
            plt.tight_layout()

            # Save to bytes
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            img_buffer.seek(0)
            plt.close(fig)

            return img_buffer

        except Exception as e:
            print(f"Error generating chart: {e}")
            return None

    def export_statistics_pdf(self):
        """Export statistics to PDF with charts and multilingual support"""
        if not hasattr(self, 'stats_data') or not self.stats_data:
            if self.L == 'it':
                msg = "Calcola prima le statistiche cliccando su 'Aggiorna Statistiche'."
            elif self.L == 'de':
                msg = "Berechnen Sie zuerst die Statistiken."
            else:
                msg = "Calculate statistics first by clicking 'Refresh Statistics'."
            QMessageBox.warning(self, "Warning", msg, QMessageBox.Ok)
            return

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
            from reportlab.lib.units import cm

            sito = self.stats_data.get('sito', 'Sconosciuto')
            total = self.stats_data.get('total', 0)

            # Language-specific labels
            if self.L == 'it':
                labels = {
                    'title': f"Report Statistico Ceramica - {sito}",
                    'summary': "Riepilogo Generale",
                    'total': f"Totale reperti ceramici: {total}",
                    'form_dist': "Distribuzione per Forma (Shape)",
                    'specific_form_dist': "Distribuzione per Forma Specifica",
                    'specific_shape_dist': "Distribuzione per Shape Specifica",
                    'ware_dist': "Distribuzione per Classe Ceramica",
                    'provenance': "Distribuzione per Provenienza",
                    'area_dist': "Distribuzione per Area",
                    'us_dist': "Distribuzione per US",
                    'sector_dist': "Distribuzione per Settore",
                    'measures': "Statistiche Misure",
                    'crosstab': "Analisi Incrociata",
                    'ai_report': "Report Descrittivo AI",
                    'category': "Categoria",
                    'count': "Conteggio",
                    'percentage': "Percentuale",
                    'measure': "Misura",
                    'min': "Min",
                    'max': "Max",
                    'mean': "Media",
                    'median': "Mediana",
                    'save_title': "Salva Report PDF",
                    'success': "Successo",
                    'saved_msg': f"Report PDF salvato in:\n"
                }
            elif self.L == 'de':
                labels = {
                    'title': f"Statistischer Keramikbericht - {sito}",
                    'summary': "Allgemeine Zusammenfassung",
                    'total': f"Gesamtzahl der Keramikfunde: {total}",
                    'form_dist': "Verteilung nach Form (Shape)",
                    'specific_form_dist': "Verteilung nach Spezifische Form",
                    'specific_shape_dist': "Verteilung nach Spezifische Shape",
                    'ware_dist': "Verteilung nach Keramikklasse",
                    'provenance': "Verteilung nach Herkunft",
                    'area_dist': "Verteilung nach Areal",
                    'us_dist': "Verteilung nach SE",
                    'sector_dist': "Verteilung nach Sektor",
                    'measures': "Maßstatistiken",
                    'crosstab': "Kreuzanalyse",
                    'ai_report': "KI-Beschreibender Bericht",
                    'category': "Kategorie",
                    'count': "Anzahl",
                    'percentage': "Prozent",
                    'measure': "Maß",
                    'min': "Min",
                    'max': "Max",
                    'mean': "Mittel",
                    'median': "Median",
                    'save_title': "PDF-Bericht speichern",
                    'success': "Erfolg",
                    'saved_msg': f"PDF-Bericht gespeichert in:\n"
                }
            else:
                labels = {
                    'title': f"Pottery Statistical Report - {sito}",
                    'summary': "General Summary",
                    'total': f"Total pottery finds: {total}",
                    'form_dist': "Distribution by Form (Shape)",
                    'specific_form_dist': "Distribution by Specific Form",
                    'specific_shape_dist': "Distribution by Specific Shape",
                    'ware_dist': "Distribution by Ware Type",
                    'provenance': "Distribution by Provenance",
                    'area_dist': "Distribution by Area",
                    'us_dist': "Distribution by SU",
                    'sector_dist': "Distribution by Sector",
                    'measures': "Measurement Statistics",
                    'crosstab': "Cross-Tabulation Analysis",
                    'ai_report': "AI Descriptive Report",
                    'category': "Category",
                    'count': "Count",
                    'percentage': "Percentage",
                    'measure': "Measure",
                    'min': "Min",
                    'max': "Max",
                    'mean': "Mean",
                    'median': "Median",
                    'save_title': "Save PDF Report",
                    'success': "Success",
                    'saved_msg': f"PDF report saved to:\n"
                }

            # Get save path
            filename, _ = QFileDialog.getSaveFileName(
                self, labels['save_title'],
                os.path.join(self.PDFFOLDER, f"Pottery_Statistics_{sito}.pdf"),
                "PDF Files (*.pdf)"
            )

            if not filename:
                return

            # Create PDF
            doc = SimpleDocTemplate(filename, pagesize=A4,
                                   rightMargin=2*cm, leftMargin=2*cm,
                                   topMargin=2*cm, bottomMargin=2*cm)

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=20)
            heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=12, spaceAfter=10)
            normal_style = styles['Normal']

            elements = []

            # Title
            elements.append(Paragraph(labels['title'], title_style))
            elements.append(Spacer(1, 12))

            # General info
            elements.append(Paragraph(labels['summary'], heading_style))
            elements.append(Paragraph(labels['total'], normal_style))
            elements.append(Spacer(1, 12))

            # Form distribution with chart
            if 'form' in self.stats_data and self.stats_data['form']:
                elements.append(Paragraph(labels['form_dist'], heading_style))

                # Add chart
                chart_buffer = self.generate_chart_image(self.stats_data['form'], labels['form_dist'], 'bar')
                if chart_buffer:
                    img = Image(chart_buffer, width=14*cm, height=7*cm)
                    elements.append(img)
                    elements.append(Spacer(1, 6))

                # Add table
                table_data = [[labels['category'], labels['count'], labels['percentage']]]
                for item, count in self.stats_data['form'][:15]:
                    percentage = (count / total) * 100 if total > 0 else 0
                    table_data.append([str(item), str(count), f"{percentage:.1f}%"])

                table = Table(table_data, colWidths=[8*cm, 3*cm, 3*cm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 12))

            # Specific Form distribution
            if 'specific_form' in self.stats_data and self.stats_data['specific_form']:
                elements.append(Paragraph(labels['specific_form_dist'], heading_style))
                chart_buffer = self.generate_chart_image(self.stats_data['specific_form'], labels['specific_form_dist'], 'bar')
                if chart_buffer:
                    img = Image(chart_buffer, width=14*cm, height=6*cm)
                    elements.append(img)
                    elements.append(Spacer(1, 6))

                table_data = [[labels['category'], labels['count'], labels['percentage']]]
                for item, count in self.stats_data['specific_form'][:15]:
                    percentage = (count / total) * 100 if total > 0 else 0
                    table_data.append([str(item), str(count), f"{percentage:.1f}%"])

                table = Table(table_data, colWidths=[8*cm, 3*cm, 3*cm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 12))

            # Specific Shape distribution
            if 'specific_shape' in self.stats_data and self.stats_data['specific_shape']:
                elements.append(Paragraph(labels['specific_shape_dist'], heading_style))
                chart_buffer = self.generate_chart_image(self.stats_data['specific_shape'], labels['specific_shape_dist'], 'pie')
                if chart_buffer:
                    img = Image(chart_buffer, width=12*cm, height=8*cm)
                    elements.append(img)
                    elements.append(Spacer(1, 12))

            # Ware distribution with chart
            if 'ware' in self.stats_data and self.stats_data['ware']:
                elements.append(Paragraph(labels['ware_dist'], heading_style))

                chart_buffer = self.generate_chart_image(self.stats_data['ware'], labels['ware_dist'], 'pie')
                if chart_buffer:
                    img = Image(chart_buffer, width=12*cm, height=8*cm)
                    elements.append(img)
                    elements.append(Spacer(1, 12))

            # Provenance section
            elements.append(PageBreak())
            elements.append(Paragraph(labels['provenance'], heading_style))

            # Area distribution
            if 'area' in self.stats_data and self.stats_data['area']:
                elements.append(Paragraph(labels['area_dist'], heading_style))
                chart_buffer = self.generate_chart_image(self.stats_data['area'], labels['area_dist'], 'bar')
                if chart_buffer:
                    img = Image(chart_buffer, width=14*cm, height=6*cm)
                    elements.append(img)
                    elements.append(Spacer(1, 6))

            # US distribution
            if 'us' in self.stats_data and self.stats_data['us']:
                elements.append(Paragraph(labels['us_dist'], heading_style))
                chart_buffer = self.generate_chart_image(self.stats_data['us'], labels['us_dist'], 'bar')
                if chart_buffer:
                    img = Image(chart_buffer, width=14*cm, height=6*cm)
                    elements.append(img)
                    elements.append(Spacer(1, 6))

            # Sector distribution
            if 'sector' in self.stats_data and self.stats_data['sector']:
                elements.append(Paragraph(labels['sector_dist'], heading_style))
                chart_buffer = self.generate_chart_image(self.stats_data['sector'], labels['sector_dist'], 'bar')
                if chart_buffer:
                    img = Image(chart_buffer, width=14*cm, height=6*cm)
                    elements.append(img)
                    elements.append(Spacer(1, 12))

            # Measurement statistics
            if 'measurements' in self.stats_data:
                elements.append(Paragraph(labels['measures'], heading_style))
                table_data = [[labels['measure'], labels['min'], labels['max'], labels['mean'], labels['median']]]
                for label_m, stats in self.stats_data['measurements'].items():
                    if isinstance(stats, dict):
                        table_data.append([
                            label_m,
                            f"{stats['min']:.2f}",
                            f"{stats['max']:.2f}",
                            f"{stats['mean']:.2f}",
                            f"{stats['median']:.2f}"
                        ])

                table = Table(table_data, colWidths=[4*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 12))

            # AI Report if available
            if 'ai_report' in self.stats_data and self.stats_data['ai_report']:
                elements.append(PageBreak())
                elements.append(Paragraph(labels['ai_report'], heading_style))
                report_text = self.stats_data['ai_report']
                for para in report_text.split('\n\n'):
                    if para.strip():
                        # Clean markdown-style headers
                        clean_para = para.strip()
                        if clean_para.startswith('##'):
                            clean_para = clean_para.replace('##', '').strip()
                            elements.append(Paragraph(f"<b>{clean_para}</b>", normal_style))
                        elif clean_para.startswith('#'):
                            clean_para = clean_para.replace('#', '').strip()
                            elements.append(Paragraph(f"<b>{clean_para}</b>", normal_style))
                        else:
                            elements.append(Paragraph(clean_para, normal_style))
                        elements.append(Spacer(1, 6))

            # Build PDF
            doc.build(elements)

            QMessageBox.information(self, labels['success'],
                labels['saved_msg'] + filename,
                QMessageBox.Ok)

            # Open the file
            if platform.system() == "Windows":
                os.startfile(filename)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", filename])
            else:
                subprocess.Popen(["xdg-open", filename])

        except Exception as e:
            QMessageBox.warning(self, "Errore", f"Errore nell'esportazione PDF: {e}", QMessageBox.Ok)

    # ==================== POTTERY VISUAL SIMILARITY SEARCH ====================

    def setup_similarity_search_ui(self):
        """Setup UI controls for visual similarity search"""
        if not HAS_SIMILARITY_SEARCH:
            return

        # Initialize search engine
        self.similarity_engine = None
        self.similarity_worker = None

        # Find or create a group box for similarity search
        # This creates UI programmatically since we don't have .ui file changes

        # Create similarity search group box
        self.similarity_group = QGroupBox("Visual Similarity Search")
        similarity_layout = QVBoxLayout()

        # Model selector
        model_layout = QHBoxLayout()
        model_layout.addWidget(QLabel("Model:"))
        self.combo_similarity_model = QComboBox()
        self.combo_similarity_model.addItems([
            'CLIP (Local)',
            'DINOv2 (Local)',
            'OpenAI (Cloud)',
            'KhutmML-CLIP (Fine-tuned)'
        ])
        self.combo_similarity_model.setToolTip("Select embedding model for similarity search")
        model_layout.addWidget(self.combo_similarity_model)
        similarity_layout.addLayout(model_layout)

        # Search type selector
        type_layout = QHBoxLayout()
        type_layout.addWidget(QLabel("Search Type:"))
        self.combo_similarity_type = QComboBox()
        self.combo_similarity_type.addItems(['General', 'Decoration', 'Shape'])
        self.combo_similarity_type.setToolTip("Type of similarity to search for")
        type_layout.addWidget(self.combo_similarity_type)
        similarity_layout.addLayout(type_layout)

        # Custom prompt for semantic search (OpenAI only) - hidden by default
        self.custom_prompt_widget = QWidget()
        prompt_layout = QVBoxLayout(self.custom_prompt_widget)
        prompt_layout.setContentsMargins(0, 0, 0, 0)
        prompt_layout.setSpacing(5)

        # Search mode radio buttons
        mode_layout = QHBoxLayout()
        self.radio_search_global = QRadioButton("Global Search")
        self.radio_search_global.setToolTip(
            "Cerca in TUTTO il database ceramiche che corrispondono alla descrizione.\n"
            "Non richiede un'immagine di partenza."
        )
        self.radio_search_combined = QRadioButton("Combined Search")
        self.radio_search_combined.setToolTip(
            "Analizza l'immagine corrente con il prompt personalizzato,\n"
            "poi cerca ceramiche simili per quelle caratteristiche specifiche."
        )
        self.radio_search_combined.setChecked(True)  # Default to combined
        mode_layout.addWidget(QLabel("Mode:"))
        mode_layout.addWidget(self.radio_search_combined)
        mode_layout.addWidget(self.radio_search_global)
        mode_layout.addStretch()
        prompt_layout.addLayout(mode_layout)

        # Prompt input
        prompt_input_layout = QHBoxLayout()
        self.label_custom_prompt = QLabel("Prompt:")
        prompt_input_layout.addWidget(self.label_custom_prompt)
        self.lineEdit_custom_prompt = QLineEdit()
        self.lineEdit_custom_prompt.setPlaceholderText("Es: decorazione a bande, forma globulare, texture ruvida...")
        self.lineEdit_custom_prompt.setToolTip(
            "Combined: Descrivi su cosa focalizzare l'analisi dell'immagine corrente.\n"
            "Global: Descrivi le caratteristiche da cercare in tutto il database."
        )
        prompt_input_layout.addWidget(self.lineEdit_custom_prompt)
        prompt_layout.addLayout(prompt_input_layout)

        similarity_layout.addWidget(self.custom_prompt_widget)
        self.custom_prompt_widget.setVisible(False)  # Hidden by default, shown only for OpenAI

        # Connect model selector to show/hide custom prompt
        self.combo_similarity_model.currentIndexChanged.connect(self.on_similarity_model_changed)

        # Threshold slider
        threshold_layout = QHBoxLayout()
        threshold_layout.addWidget(QLabel("Threshold:"))
        self.slider_similarity_threshold = QSlider(Qt.Horizontal)
        self.slider_similarity_threshold.setMinimum(10)
        self.slider_similarity_threshold.setMaximum(100)
        self.slider_similarity_threshold.setValue(70)
        self.slider_similarity_threshold.setTickPosition(QSlider.TicksBelow)
        self.slider_similarity_threshold.setTickInterval(10)
        self.slider_similarity_threshold.valueChanged.connect(self.on_threshold_changed)
        threshold_layout.addWidget(self.slider_similarity_threshold)
        self.label_threshold_value = QLabel("70%")
        threshold_layout.addWidget(self.label_threshold_value)
        similarity_layout.addLayout(threshold_layout)

        # Filter options
        filter_layout = QHBoxLayout()
        self.chk_only_decorated = QCheckBox("Only decorated pottery")
        self.chk_only_decorated.setToolTip("Filter results to show only pottery with decoration (exdeco or intdeco field not empty)")
        self.chk_only_decorated.setChecked(True)  # Default ON for decoration searches
        filter_layout.addWidget(self.chk_only_decorated)
        similarity_layout.addLayout(filter_layout)

        # Advanced preprocessing options - Row 1
        preproc_layout = QHBoxLayout()
        self.chk_auto_crop = QCheckBox("Auto-crop detail")
        self.chk_auto_crop.setToolTip("Auto-crop to region with most decoration detail")
        self.chk_auto_crop.setChecked(False)
        preproc_layout.addWidget(self.chk_auto_crop)

        self.chk_edge_preproc = QCheckBox("Edge-enhance")
        self.chk_edge_preproc.setToolTip("Use edge-detection preprocessing (better for line decorations)")
        self.chk_edge_preproc.setChecked(False)
        preproc_layout.addWidget(self.chk_edge_preproc)
        similarity_layout.addLayout(preproc_layout)

        # Advanced preprocessing options - Row 2 (Segmentation)
        segment_layout = QHBoxLayout()
        self.chk_segment_decoration = QCheckBox("Isolate decoration")
        self.chk_segment_decoration.setToolTip("Segment and isolate decorated areas (mask plain clay)")
        self.chk_segment_decoration.setChecked(False)
        segment_layout.addWidget(self.chk_segment_decoration)

        self.chk_remove_background = QCheckBox("Remove background")
        self.chk_remove_background.setToolTip("Remove photo background from pottery (useful for studio photos)")
        self.chk_remove_background.setChecked(False)
        segment_layout.addWidget(self.chk_remove_background)
        similarity_layout.addLayout(segment_layout)

        # Buttons - Row 1 (Search)
        buttons_layout = QHBoxLayout()
        self.btn_find_similar = QPushButton("Find Similar")
        self.btn_find_similar.setToolTip("Find pottery visually similar to current record")
        self.btn_find_similar.clicked.connect(self.on_find_similar_clicked)
        buttons_layout.addWidget(self.btn_find_similar)

        self.btn_compare_external = QPushButton("Compare External Image")
        self.btn_compare_external.setToolTip(
            "Compare an external image (not in database) against the pottery index.\n"
            "Useful for identifying unknown pottery fragments by visual similarity."
        )
        self.btn_compare_external.clicked.connect(self.on_compare_external_clicked)
        buttons_layout.addWidget(self.btn_compare_external)
        similarity_layout.addLayout(buttons_layout)

        # Buttons - Row 2 (Index management)
        index_buttons_layout = QHBoxLayout()
        self.btn_build_index = QPushButton("Build Index")
        self.btn_build_index.setToolTip("Build/rebuild similarity index for selected model (from scratch)")
        self.btn_build_index.clicked.connect(self.on_build_index_clicked)
        index_buttons_layout.addWidget(self.btn_build_index)

        self.btn_update_index = QPushButton("Update Index")
        self.btn_update_index.setToolTip("Update existing indexes: add new, update modified, remove deleted images")
        self.btn_update_index.clicked.connect(self.on_update_index_clicked)
        index_buttons_layout.addWidget(self.btn_update_index)
        similarity_layout.addLayout(index_buttons_layout)

        # Buttons - Row 2 (Import/Export)
        io_layout = QHBoxLayout()
        self.btn_export_index = QPushButton("Export Indexes")
        self.btn_export_index.setToolTip("Export all indexes to ZIP for sharing with other PCs")
        self.btn_export_index.clicked.connect(self.on_export_indexes_clicked)
        io_layout.addWidget(self.btn_export_index)

        self.btn_import_index = QPushButton("Import Indexes")
        self.btn_import_index.setToolTip("Import indexes from ZIP (from another PC with same database)")
        self.btn_import_index.clicked.connect(self.on_import_indexes_clicked)
        io_layout.addWidget(self.btn_import_index)
        similarity_layout.addLayout(io_layout)

        # Buttons - Row 4 (Training KhutmML)
        training_layout = QHBoxLayout()
        self.btn_train_khutm = QPushButton("Train KhutmML")
        self.btn_train_khutm.setToolTip(
            "Fine-tune the KhutmML-CLIP model on your pottery dataset.\n"
            "This creates a specialized model for your archaeological collection,\n"
            "improving similarity search accuracy."
        )
        self.btn_train_khutm.clicked.connect(self.on_train_khutm_clicked)
        training_layout.addWidget(self.btn_train_khutm)

        self.btn_prepare_dataset = QPushButton("Prepare Dataset")
        self.btn_prepare_dataset.setToolTip(
            "Prepare a training dataset from existing pottery images.\n"
            "Organizes images by type and creates positive/negative pairs."
        )
        self.btn_prepare_dataset.clicked.connect(self.on_prepare_dataset_clicked)
        training_layout.addWidget(self.btn_prepare_dataset)

        self.btn_export_khutm_model = QPushButton("Export Model")
        self.btn_export_khutm_model.setToolTip(
            "Export the trained KhutmML-CLIP model to a ZIP file.\n"
            "Use this to backup or share your trained model."
        )
        self.btn_export_khutm_model.clicked.connect(self.on_export_khutm_model_clicked)
        training_layout.addWidget(self.btn_export_khutm_model)

        self.btn_import_khutm_model = QPushButton("Import Model")
        self.btn_import_khutm_model.setToolTip(
            "Import a KhutmML-CLIP model from a ZIP file.\n"
            "Replaces the current trained model with the imported one."
        )
        self.btn_import_khutm_model.clicked.connect(self.on_import_khutm_model_clicked)
        training_layout.addWidget(self.btn_import_khutm_model)
        similarity_layout.addLayout(training_layout)

        # Auto-update checkbox
        auto_update_layout = QHBoxLayout()
        self.checkbox_auto_update = QCheckBox("Auto-update index when images are added/removed")
        self.checkbox_auto_update.setToolTip(
            "Automatically update CLIP embedding index when pottery images are added or removed.\n"
            "This keeps the similarity search index up-to-date without manual rebuild."
        )
        # Load saved state from settings
        s = QSettings()
        auto_update_enabled = s.value('pyArchInit/pottery_similarity_auto_update', True, type=bool)
        self.checkbox_auto_update.setChecked(auto_update_enabled)
        self.checkbox_auto_update.stateChanged.connect(self.on_auto_update_changed)
        auto_update_layout.addWidget(self.checkbox_auto_update)
        similarity_layout.addLayout(auto_update_layout)

        # Drop zone for external images
        self.drop_zone_frame = QFrame()
        self.drop_zone_frame.setFrameStyle(QFrame.StyledPanel | QFrame.Sunken)
        self.drop_zone_frame.setAcceptDrops(True)
        self.drop_zone_frame.setMinimumHeight(60)
        self.drop_zone_frame.setStyleSheet("""
            QFrame {
                border: 2px dashed #aaa;
                border-radius: 5px;
                background-color: #f5f5f5;
            }
            QFrame:hover {
                border-color: #666;
                background-color: #e8e8e8;
            }
        """)
        drop_zone_layout = QVBoxLayout(self.drop_zone_frame)
        self.drop_zone_label = QLabel("Drop image here to compare\nor click 'Compare External Image'")
        self.drop_zone_label.setAlignment(Qt.AlignCenter)
        self.drop_zone_label.setStyleSheet("color: #666; border: none; background: transparent;")
        drop_zone_layout.addWidget(self.drop_zone_label)

        # Install event filter for drag & drop
        self.drop_zone_frame.installEventFilter(self)

        similarity_layout.addWidget(self.drop_zone_frame)

        # Status label
        self.label_similarity_status = QLabel("Ready")
        self.label_similarity_status.setStyleSheet("color: gray; font-style: italic;")
        similarity_layout.addWidget(self.label_similarity_status)

        # Progress bar (hidden by default)
        self.progress_similarity = QProgressBar()
        self.progress_similarity.setVisible(False)
        similarity_layout.addWidget(self.progress_similarity)

        self.similarity_group.setLayout(similarity_layout)

        # Try to add to existing tab widget or create dock
        # First look for an existing tab widget
        tab_widget = self.findChild(QTabWidget)
        if tab_widget:
            # Create a new widget for the tab
            similarity_tab = QWidget()
            tab_layout = QVBoxLayout()
            tab_layout.addWidget(self.similarity_group)
            tab_layout.addStretch()
            similarity_tab.setLayout(tab_layout)
            tab_widget.addTab(similarity_tab, "Similarity")
        else:
            # Try to find a suitable place in the existing layout
            # For now, we'll add it to the main layout if possible
            pass

    def setup_embedding_auto_updater(self):
        """Setup the automatic embedding index updater"""
        if not HAS_SIMILARITY_SEARCH:
            self.embedding_updater = None
            return

        try:
            # Get the singleton embedding updater and set up DB manager
            self.embedding_updater = get_embedding_updater(self.DB_MANAGER)

            # Read auto-update setting from QGIS settings
            s = QSettings()
            auto_update_enabled = s.value('pyArchInit/pottery_similarity_auto_update', True, type=bool)
            self.embedding_updater.set_enabled(auto_update_enabled)

            # Connect signals for status updates
            if hasattr(self.embedding_updater, 'embedding_added'):
                self.embedding_updater.embedding_added.connect(self.on_embedding_added)
            if hasattr(self.embedding_updater, 'embedding_removed'):
                self.embedding_updater.embedding_removed.connect(self.on_embedding_removed)
            if hasattr(self.embedding_updater, 'embedding_error'):
                self.embedding_updater.embedding_error.connect(self.on_embedding_error)

            print(f"[Pottery] Embedding auto-updater initialized (enabled={auto_update_enabled})")

        except Exception as e:
            print(f"[Pottery] Error setting up embedding auto-updater: {e}")
            self.embedding_updater = None

    def on_embedding_added(self, pottery_id, media_id, model):
        """Callback when embedding is added to index"""
        print(f"[Pottery] Embedding added: pottery={pottery_id}, media={media_id}, model={model}")
        # Update status in similarity panel if available
        if hasattr(self, 'label_similarity_status'):
            self.label_similarity_status.setText(f"Index updated (+1 {model})")

    def on_embedding_removed(self, pottery_id, media_id, model):
        """Callback when embedding is removed from index"""
        print(f"[Pottery] Embedding removed: pottery={pottery_id}, media={media_id}, model={model}")
        if hasattr(self, 'label_similarity_status'):
            self.label_similarity_status.setText(f"Index updated (-1 {model})")

    def on_embedding_error(self, error_msg):
        """Callback when embedding operation fails"""
        print(f"[Pottery] Embedding error: {error_msg}")

    def on_auto_update_changed(self, state):
        """Handle checkbox state change for auto-update setting"""
        enabled = state == Qt.Checked
        # Save to settings
        s = QSettings()
        s.setValue('pyArchInit/pottery_similarity_auto_update', enabled)

        # Update the embedding updater
        if hasattr(self, 'embedding_updater') and self.embedding_updater is not None:
            self.embedding_updater.set_enabled(enabled)

        # Update status label
        if hasattr(self, 'label_similarity_status'):
            status = "Auto-update enabled" if enabled else "Auto-update disabled"
            self.label_similarity_status.setText(status)

        print(f"[Pottery] Auto-update index {'enabled' if enabled else 'disabled'}")

    def eventFilter(self, obj, event):
        """Handle drag & drop events for external image comparison"""
        if obj == getattr(self, 'drop_zone_frame', None):
            if event.type() == QEvent.DragEnter:
                if event.mimeData().hasUrls():
                    # Check if any URL is an image
                    for url in event.mimeData().urls():
                        file_path = url.toLocalFile()
                        if file_path.lower().endswith(('.jpg', '.jpeg', '.png', '.tif', '.tiff', '.bmp')):
                            event.acceptProposedAction()
                            # Visual feedback
                            self.drop_zone_frame.setStyleSheet("""
                                QFrame {
                                    border: 2px dashed #4CAF50;
                                    border-radius: 5px;
                                    background-color: #e8f5e9;
                                }
                            """)
                            return True
                return False

            elif event.type() == QEvent.DragLeave:
                # Reset visual feedback
                self.drop_zone_frame.setStyleSheet("""
                    QFrame {
                        border: 2px dashed #aaa;
                        border-radius: 5px;
                        background-color: #f5f5f5;
                    }
                    QFrame:hover {
                        border-color: #666;
                        background-color: #e8e8e8;
                    }
                """)
                return True

            elif event.type() == QEvent.Drop:
                # Reset visual feedback
                self.drop_zone_frame.setStyleSheet("""
                    QFrame {
                        border: 2px dashed #aaa;
                        border-radius: 5px;
                        background-color: #f5f5f5;
                    }
                    QFrame:hover {
                        border-color: #666;
                        background-color: #e8e8e8;
                    }
                """)

                # Get dropped file
                if event.mimeData().hasUrls():
                    for url in event.mimeData().urls():
                        file_path = url.toLocalFile()
                        if file_path.lower().endswith(('.jpg', '.jpeg', '.png', '.tif', '.tiff', '.bmp')):
                            self._compare_external_image(file_path)
                            return True
                return False

        return super().eventFilter(obj, event)

    def on_compare_external_clicked(self):
        """Handle compare external image button click - opens file dialog"""
        # Open file dialog
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Image to Compare",
            "",
            "Images (*.jpg *.jpeg *.png *.tif *.tiff *.bmp);;All Files (*)"
        )

        if file_path:
            self._compare_external_image(file_path)

    def _compare_external_image(self, image_path: str):
        """
        Compare an external image against the pottery index.

        Args:
            image_path: Full path to the external image file
        """
        if not os.path.exists(image_path):
            QMessageBox.warning(self, "Error", f"Image file not found:\n{image_path}")
            return

        if not HAS_SIMILARITY_SEARCH:
            QMessageBox.warning(self, "Error", "Similarity search module not available")
            return

        # Initialize engine if needed
        if self.similarity_engine is None:
            self.similarity_engine = PotterySimilaritySearchEngine(self.DB_MANAGER)

        # Get search parameters
        model_name = self.get_similarity_model_name()
        search_type = self.get_similarity_search_type()
        threshold = self.slider_similarity_threshold.value() / 100.0

        # Get preprocessing options
        auto_crop = getattr(self, 'chk_auto_crop', None) and self.chk_auto_crop.isChecked()
        edge_preproc = getattr(self, 'chk_edge_preproc', None) and self.chk_edge_preproc.isChecked()
        segment_decoration = getattr(self, 'chk_segment_decoration', None) and self.chk_segment_decoration.isChecked()
        remove_background = getattr(self, 'chk_remove_background', None) and self.chk_remove_background.isChecked()

        # Update status
        self.label_similarity_status.setText(f"Comparing external image with {model_name}...")
        self.drop_zone_label.setText(f"Analyzing: {os.path.basename(image_path)}")

        # Store external image path for results dialog
        self._external_image_path = image_path

        # Create and start worker
        self.similarity_worker = PotterySimilarityWorker(
            self.similarity_engine,
            'search',
            image_path=image_path,
            model_name=model_name,
            search_type=search_type,
            threshold=threshold,
            auto_crop=auto_crop,
            edge_preprocessing=edge_preproc,
            segment_decoration=segment_decoration,
            remove_background=remove_background,
            exclude_pottery_id=None  # No exclusion for external images
        )

        self.similarity_worker.search_complete.connect(self.on_external_search_complete)
        self.similarity_worker.search_complete_with_meta.connect(self.on_external_search_complete_with_meta)
        self.similarity_worker.error_occurred.connect(self.on_similarity_error)
        self.similarity_worker.start()

    def on_external_search_complete(self, results):
        """Handle completion of external image search"""
        self.on_external_search_complete_with_meta(results, {})

    def on_external_search_complete_with_meta(self, results, meta):
        """Handle completion of external image search with metadata"""
        # Reset drop zone
        self.drop_zone_label.setText("Drop image here to compare\nor click 'Compare External Image'")

        threshold = self.slider_similarity_threshold.value()
        model_name = self.get_similarity_model_name()

        if not results:
            self.label_similarity_status.setText("No similar pottery found")
            # Show info about top scores if available
            top_scores = meta.get('top_scores', [])
            if top_scores:
                msg = f"No matches found above {threshold}% threshold.\n\n"
                msg += f"Best available scores: {', '.join([f'{s:.1f}%' for s in top_scores[:3]])}\n\n"
                msg += "Try lowering the threshold slider to see more results."
                QMessageBox.information(self, "No Matches", msg)
            else:
                QMessageBox.information(self, "No Matches",
                    f"No similar pottery found above {threshold}% threshold.\n"
                    "The index may be empty - try building the index first.")
            return

        self.label_similarity_status.setText(f"Found {len(results)} similar items")

        # Show results dialog with external image info
        self.show_external_results_dialog(results)

    def show_external_results_dialog(self, results):
        """
        Show similarity search results for external image with datazione info.

        Args:
            results: List of match results
        """
        dialog = QDialog(self)
        dialog.setWindowTitle("External Image Comparison Results")
        dialog.setMinimumSize(900, 700)

        layout = QVBoxLayout(dialog)

        # Header with external image info
        header_layout = QHBoxLayout()

        # Show external image thumbnail if available
        external_path = getattr(self, '_external_image_path', None)
        if external_path and os.path.exists(external_path):
            thumb_label = QLabel()
            pixmap = QPixmap(external_path)
            if not pixmap.isNull():
                thumb_label.setPixmap(pixmap.scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                header_layout.addWidget(thumb_label)

        header_text = QLabel(f"<b>External Image:</b> {os.path.basename(external_path) if external_path else 'Unknown'}<br>"
                            f"<b>Found:</b> {len(results)} similar pottery")
        header_text.setTextFormat(Qt.RichText)
        header_layout.addWidget(header_text)
        header_layout.addStretch()
        layout.addLayout(header_layout)

        # Separator
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        layout.addWidget(line)

        # Results table
        table = QTableWidget()
        table.setColumnCount(9)
        table.setHorizontalHeaderLabels([
            "Similarity", "ID", "Sito", "Area", "US", "Form", "Decoration", "Datazione", "Image"
        ])
        table.setRowCount(len(results))
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setAlternatingRowColors(True)

        # Column widths
        table.setColumnWidth(0, 80)   # Similarity
        table.setColumnWidth(1, 60)   # ID
        table.setColumnWidth(2, 100)  # Sito
        table.setColumnWidth(3, 50)   # Area
        table.setColumnWidth(4, 50)   # US
        table.setColumnWidth(5, 100)  # Form
        table.setColumnWidth(6, 120)  # Decoration
        table.setColumnWidth(7, 150)  # Datazione
        table.setColumnWidth(8, 80)   # Image

        for row, result in enumerate(results):
            pottery_data = result.get('pottery_data', {})

            # Similarity percentage
            sim_percent = result.get('similarity_percent', result.get('similarity', 0) * 100)
            sim_item = QTableWidgetItem(f"{sim_percent:.1f}%")
            sim_item.setTextAlignment(Qt.AlignCenter)
            # Color code similarity
            if sim_percent >= 80:
                sim_item.setBackground(QColor(200, 255, 200))  # Green
            elif sim_percent >= 60:
                sim_item.setBackground(QColor(255, 255, 200))  # Yellow
            else:
                sim_item.setBackground(QColor(255, 220, 220))  # Red
            table.setItem(row, 0, sim_item)

            # ID
            id_number = pottery_data.get('id_number', str(result.get('pottery_id', '')))
            table.setItem(row, 1, QTableWidgetItem(str(id_number)))

            # Sito
            table.setItem(row, 2, QTableWidgetItem(str(pottery_data.get('sito', ''))))

            # Area
            table.setItem(row, 3, QTableWidgetItem(str(pottery_data.get('area', ''))))

            # US
            table.setItem(row, 4, QTableWidgetItem(str(pottery_data.get('us', ''))))

            # Form
            form = pottery_data.get('form', '')
            specific_form = pottery_data.get('specific_form', '')
            form_text = f"{form} - {specific_form}" if form and specific_form else (form or specific_form)
            table.setItem(row, 5, QTableWidgetItem(str(form_text)))

            # Decoration
            exdeco = pottery_data.get('exdeco', '')
            intdeco = pottery_data.get('intdeco', '')
            deco_text = f"Ext: {exdeco}" if exdeco else ""
            if intdeco:
                deco_text += f" Int: {intdeco}" if deco_text else f"Int: {intdeco}"
            table.setItem(row, 6, QTableWidgetItem(str(deco_text)))

            # Datazione - from pottery record
            datazione = ''
            pottery_id = result.get('pottery_id')
            if pottery_id and self.DB_MANAGER:
                try:
                    pottery_rec = self.DB_MANAGER.get_pottery_by_id_rep(pottery_id)
                    if pottery_rec and hasattr(pottery_rec, 'datazione'):
                        datazione = pottery_rec.datazione or ''
                except:
                    pass
            datazione_item = QTableWidgetItem(str(datazione))
            datazione_item.setToolTip("Chronological dating from periodizzazione")
            if datazione:
                datazione_item.setBackground(QColor(230, 230, 255))  # Light blue for dated items
            table.setItem(row, 7, datazione_item)

            # Image button
            img_btn = QPushButton("View")
            img_btn.setProperty('pottery_id', result.get('pottery_id'))
            img_btn.setProperty('image_path', result.get('image_path', ''))
            img_btn.clicked.connect(lambda checked, r=result: self._show_result_image(r))
            table.setCellWidget(row, 8, img_btn)

        layout.addWidget(table)

        # Summary statistics
        stats_layout = QHBoxLayout()

        # Count by datazione
        datazioni = {}
        for result in results:
            pottery_id = result.get('pottery_id')
            if pottery_id and self.DB_MANAGER:
                try:
                    pottery_rec = self.DB_MANAGER.get_pottery_by_id_rep(pottery_id)
                    if pottery_rec and hasattr(pottery_rec, 'datazione') and pottery_rec.datazione:
                        dat = pottery_rec.datazione
                        datazioni[dat] = datazioni.get(dat, 0) + 1
                except:
                    pass

        if datazioni:
            stats_text = "<b>Datazione summary:</b> "
            sorted_dat = sorted(datazioni.items(), key=lambda x: -x[1])
            stats_text += ", ".join([f"{d}: {c}" for d, c in sorted_dat[:5]])
            stats_label = QLabel(stats_text)
            stats_label.setTextFormat(Qt.RichText)
            stats_layout.addWidget(stats_label)

        stats_layout.addStretch()
        layout.addLayout(stats_layout)

        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        btn_go_to = QPushButton("Go to Selected")
        btn_go_to.setToolTip("Navigate to the selected pottery record")
        btn_go_to.clicked.connect(lambda: self._go_to_result_from_table(table, dialog))
        button_layout.addWidget(btn_go_to)

        btn_close = QPushButton("Close")
        btn_close.clicked.connect(dialog.close)
        button_layout.addWidget(btn_close)

        layout.addLayout(button_layout)

        dialog.exec_()

    def _show_result_image(self, result):
        """Show the matched image in a popup"""
        image_path = result.get('image_path', '')
        if not image_path or not os.path.exists(image_path):
            # Try to get from relative path
            relative_path = result.get('relative_path', '')
            if relative_path and self.similarity_engine:
                image_path = self.similarity_engine._build_image_path(relative_path)

        if image_path and os.path.exists(image_path):
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Match - {result.get('similarity_percent', 0):.1f}%")
            layout = QVBoxLayout(dialog)

            label = QLabel()
            pixmap = QPixmap(image_path)
            if not pixmap.isNull():
                # Scale to reasonable size
                scaled = pixmap.scaled(600, 600, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                label.setPixmap(scaled)
            else:
                label.setText("Could not load image")

            layout.addWidget(label)
            dialog.exec_()
        else:
            QMessageBox.warning(self, "Image Not Found",
                f"Could not find image file:\n{image_path or 'No path available'}")

    def _go_to_result_from_table(self, table, dialog):
        """Navigate to selected result from table"""
        selected = table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "No Selection", "Please select a row first")
            return

        row = selected[0].row()
        # Get pottery_id from the View button in that row
        btn = table.cellWidget(row, 8)
        if btn:
            pottery_id = btn.property('pottery_id')
            if pottery_id:
                dialog.close()
                self._navigate_to_pottery(pottery_id)

    def _navigate_to_pottery(self, pottery_id):
        """Navigate to a specific pottery record by id_rep"""
        try:
            # Find the record in DATA_LIST
            for i, record in enumerate(self.DATA_LIST):
                if record.id_rep == pottery_id:
                    self.rec_num = i
                    self.fill_fields()
                    self.set_rec_counter(len(self.DATA_LIST), self.rec_num + 1)
                    return

            # If not found in current list, search database
            search_dict = {'id_rep': pottery_id}
            results = self.DB_MANAGER.query_bool(search_dict, 'POTTERY')
            if results:
                pottery = results[0]
                # Re-search to get full data
                self.empty_fields()
                search_dict = {
                    'sito': "'" + pottery.sito + "'",
                    'id_number': "'" + str(pottery.id_number) + "'"
                }
                self.search_rec(search_dict)

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not navigate to record: {e}")

    def on_threshold_changed(self, value):
        """Update threshold label when slider changes"""
        self.label_threshold_value.setText(f"{value}%")

    def on_similarity_model_changed(self, index):
        """Show/hide custom prompt based on selected model"""
        model_text = self.combo_similarity_model.currentText()
        is_openai = 'OpenAI' in model_text
        if hasattr(self, 'custom_prompt_widget'):
            self.custom_prompt_widget.setVisible(is_openai)
            # Clear the prompt when switching away from OpenAI
            if not is_openai and hasattr(self, 'lineEdit_custom_prompt'):
                self.lineEdit_custom_prompt.clear()

    def get_similarity_model_name(self):
        """Get model name from combo box selection"""
        text = self.combo_similarity_model.currentText()
        if 'KhutmML' in text or 'Fine-tuned' in text:
            return 'khutm_clip'
        elif 'CLIP' in text:
            return 'clip'
        elif 'DINOv2' in text:
            return 'dinov2'
        elif 'OpenAI' in text:
            return 'openai'
        return 'clip'

    def get_similarity_search_type(self):
        """Get search type from combo box selection"""
        text = self.combo_similarity_type.currentText().lower()
        if text in ['general', 'decoration', 'shape']:
            return text
        return 'general'

    def on_find_similar_clicked(self):
        """Handle find similar button click"""
        if not HAS_SIMILARITY_SEARCH:
            QMessageBox.warning(self, "Error", "Similarity search module not available")
            return

        # Initialize engine if needed
        if self.similarity_engine is None:
            self.similarity_engine = PotterySimilaritySearchEngine(self.DB_MANAGER)

        # Get custom prompt first - determines search mode
        custom_prompt = ''
        if hasattr(self, 'lineEdit_custom_prompt'):
            custom_prompt = self.lineEdit_custom_prompt.text().strip()

        model_name = self.get_similarity_model_name()
        search_type = self.get_similarity_search_type()
        threshold = self.slider_similarity_threshold.value() / 100.0

        # Get advanced preprocessing options
        auto_crop = getattr(self, 'chk_auto_crop', None) and self.chk_auto_crop.isChecked()
        edge_preproc = getattr(self, 'chk_edge_preproc', None) and self.chk_edge_preproc.isChecked()
        segment_decoration = getattr(self, 'chk_segment_decoration', None) and self.chk_segment_decoration.isChecked()
        remove_background = getattr(self, 'chk_remove_background', None) and self.chk_remove_background.isChecked()

        # THREE MODES:
        # 1. Global Search (with custom prompt, OpenAI): text-only search in entire DB
        # 2. Combined Search (with custom prompt + image, OpenAI): analyze image with custom focus
        # 3. Standard Search (no custom prompt): image-based similarity search

        # Check search mode when custom prompt is provided
        is_global_search = hasattr(self, 'radio_search_global') and self.radio_search_global.isChecked()

        if custom_prompt and is_global_search:
            # MODE 1: Global text-based semantic search (no image needed)
            print(f"[SIMILARITY] GLOBAL TEXT SEARCH - Prompt: {custom_prompt[:50]}...")
            print(f"[SIMILARITY] Model={model_name}, Threshold={threshold}")

            if model_name != 'openai':
                QMessageBox.warning(self, "OpenAI Required",
                    "Global text search requires OpenAI model.\n"
                    "Please select 'OpenAI Vision (cloud)' from the model dropdown.")
                return

            # For global text search, use lower default threshold (text vs image descriptions = lower scores)
            if threshold > 0.5:
                threshold = 0.40  # Override to 40% for text search
                print(f"[SIMILARITY] Threshold adjusted to {threshold*100:.0f}% for text search")

            # Update status
            self.label_similarity_status.setText("Global search by description...")
            self.btn_find_similar.setEnabled(False)

            # Create worker for text-based search (no image needed)
            self.similarity_worker = PotterySimilarityWorker(
                self.similarity_engine,
                'search_by_text',
                custom_prompt=custom_prompt,
                model_name=model_name,
                search_type=search_type,
                threshold=threshold
            )
            # Connect both signals - with_meta has top scores for feedback
            self.similarity_worker.search_complete.connect(self.on_similarity_search_complete)
            self.similarity_worker.search_complete_with_meta.connect(self.on_similarity_search_complete_with_meta)
            self.similarity_worker.error_occurred.connect(self.on_similarity_error)
            self.similarity_worker.start()
            return

        # MODE 2 & 3: Need an image - either combined search or standard search
        # Check if we have a current record
        if not self.DATA_LIST or self.REC_CORR >= len(self.DATA_LIST):
            QMessageBox.warning(self, "Error", "No pottery record selected")
            return

        current_record = self.DATA_LIST[self.REC_CORR]
        pottery_id = current_record.id_rep  # Use id_rep for DB queries (FK to media_to_entity_table)
        pottery_id_number = getattr(current_record, 'id_number', 'N/A')  # Display ID for user

        # Log which record we're searching from
        print(f"[SIMILARITY] IMAGE SEARCH MODE from pottery id_number={pottery_id_number} (id_rep={pottery_id})")
        print(f"[SIMILARITY] Record info: sito={getattr(current_record, 'sito', 'N/A')}, "
              f"area={getattr(current_record, 'area', 'N/A')}, "
              f"us={getattr(current_record, 'us', 'N/A')}")

        # Get all images for this pottery record
        all_images = self.DB_MANAGER.get_all_pottery_images(pottery_id)

        if not all_images:
            QMessageBox.warning(self, "No Images", f"No images found for pottery ID {pottery_id_number}")
            return

        # If multiple images, let user choose which one to search with
        selected_image_path = None
        if len(all_images) > 1:
            selected_image_path = self.show_image_selection_dialog(all_images, pottery_id_number)
            if not selected_image_path:
                return  # User cancelled
        else:
            # Only one image, use it directly
            relative_path = all_images[0].get('path_resize')
            if relative_path:
                selected_image_path = self.similarity_engine._build_image_path(relative_path)

        if not selected_image_path:
            QMessageBox.warning(self, "Error", "Could not determine image path")
            return

        # Determine search mode
        is_combined_search = custom_prompt and not is_global_search

        if is_combined_search:
            print(f"[SIMILARITY] COMBINED SEARCH - Image + Custom prompt: {custom_prompt[:50]}...")
        else:
            print(f"[SIMILARITY] STANDARD IMAGE SEARCH")

        print(f"[SIMILARITY] Selected image: {selected_image_path}")
        print(f"[SIMILARITY] Model={model_name}, Type={search_type}, Threshold={threshold}")
        print(f"[SIMILARITY] Auto-crop={auto_crop}, Edge-preproc={edge_preproc}")
        print(f"[SIMILARITY] Segment-decoration={segment_decoration}, Remove-background={remove_background}")

        # Update status
        if is_combined_search:
            self.label_similarity_status.setText(f"Combined search with {model_name}...")
        else:
            self.label_similarity_status.setText(f"Searching with {model_name}...")
        self.btn_find_similar.setEnabled(False)

        # Create worker for background search - use search_similar with specific image
        self.similarity_worker = PotterySimilarityWorker(
            self.similarity_engine,
            'search',  # Use direct search instead of search_by_id
            image_path=selected_image_path,
            model_name=model_name,
            search_type=search_type,
            threshold=threshold,
            exclude_pottery_id=pottery_id,  # Exclude current pottery from results
            auto_crop=auto_crop,
            edge_preprocessing=edge_preproc,
            segment_decoration=segment_decoration,
            remove_background=remove_background,
            custom_prompt=custom_prompt  # For OpenAI combined search - analyzes image with this focus
        )
        self.similarity_worker.search_complete.connect(self.on_similarity_search_complete)
        # Connect with_meta signal for detailed feedback
        self.similarity_worker.search_complete_with_meta.connect(self.on_similarity_search_complete_with_meta)
        self.similarity_worker.error_occurred.connect(self.on_similarity_error)
        self.similarity_worker.start()

    def show_image_selection_dialog(self, images, pottery_id_number):
        """Show dialog to let user select which image to use for similarity search"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Select Image for Pottery {pottery_id_number}")
        dialog.setMinimumSize(600, 400)

        layout = QVBoxLayout()

        label = QLabel(f"This pottery has {len(images)} images. Select one to search for similar:")
        layout.addWidget(label)

        # Image list with thumbnails
        list_widget = QListWidget()
        list_widget.setViewMode(QListWidget.IconMode)
        list_widget.setIconSize(QSize(120, 120))
        list_widget.setSpacing(10)
        list_widget.setResizeMode(QListWidget.Adjust)

        # Get config for path building
        conn = Connection()
        thumb_path_config = conn.thumb_path()
        thumb_path_str = thumb_path_config.get('thumb_path', '')
        is_cloudinary = thumb_path_str.lower().startswith('cloudinary://')

        for img in images:
            filename = img.get('filename', img.get('path_resize', 'Unknown'))
            relative_path = img.get('path_resize', '')

            item = QListWidgetItem(filename)

            # Try to load thumbnail
            if relative_path:
                if is_cloudinary and HAS_REMOTE_LOADER:
                    cloudinary_path = f"{thumb_path_str}/{relative_path}"
                    pixmap = RemoteImageLoader.load_pixmap(cloudinary_path, 120, 120)
                else:
                    full_path = self.similarity_engine._build_image_path(relative_path)
                    if os.path.exists(full_path):
                        pixmap = QPixmap(full_path)
                        if not pixmap.isNull():
                            pixmap = pixmap.scaled(120, 120, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    else:
                        pixmap = None

                if pixmap and not pixmap.isNull():
                    item.setIcon(QIcon(pixmap))

            # Store full path in item data
            if relative_path:
                full_path = self.similarity_engine._build_image_path(relative_path)
                item.setData(Qt.UserRole, full_path)

            list_widget.addItem(item)

        layout.addWidget(list_widget)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_ok = QPushButton("Search with Selected")
        btn_cancel = QPushButton("Cancel")
        btn_ok.clicked.connect(dialog.accept)
        btn_cancel.clicked.connect(dialog.reject)
        btn_layout.addWidget(btn_ok)
        btn_layout.addWidget(btn_cancel)
        layout.addLayout(btn_layout)

        dialog.setLayout(layout)

        # Select first item by default
        if list_widget.count() > 0:
            list_widget.setCurrentRow(0)

        if dialog.exec_() == QDialog.Accepted:
            selected_item = list_widget.currentItem()
            if selected_item:
                return selected_item.data(Qt.UserRole)
        return None

    def on_similarity_search_complete(self, results):
        """Handle search completion"""
        self.btn_find_similar.setEnabled(True)

        # Get current settings for informative message
        current_threshold = self.slider_similarity_threshold.value()
        model_name = self.get_similarity_model_name()
        search_type = self.get_similarity_search_type()
        custom_prompt = ''
        if hasattr(self, 'lineEdit_custom_prompt'):
            custom_prompt = self.lineEdit_custom_prompt.text().strip()
        is_global_search = hasattr(self, 'radio_search_global') and self.radio_search_global.isChecked()

        if not results:
            self.label_similarity_status.setText("No similar pottery found")

            # Build informative message with suggestions
            msg_parts = [f"No pottery found above {current_threshold}% threshold."]

            # Try to get max score from index to show what's available
            if self.similarity_engine and hasattr(self.similarity_engine, 'index_manager'):
                try:
                    index, mapping = self.similarity_engine.index_manager.get_index(model_name, search_type)
                    if index and index.ntotal > 0:
                        msg_parts.append(f"\nIndex contains {index.ntotal} pottery images.")
                    else:
                        msg_parts.append(f"\nIndex '{model_name}_{search_type}' is empty or not built.")
                        msg_parts.append("→ Try: Build Index first!")
                except:
                    pass

            # Suggestions based on settings
            msg_parts.append("\n\n📌 Suggestions:")

            if current_threshold > 70:
                msg_parts.append(f"• Lower threshold (currently {current_threshold}%)")

            if custom_prompt and is_global_search:
                msg_parts.append("• Global text search typically has lower scores (40-60%)")
                msg_parts.append("• Try threshold ≤ 50%")

            if model_name == 'openai' and search_type == 'general':
                msg_parts.append("• For decorations: build 'Decoration' index")

            if not custom_prompt and model_name in ['clip', 'dinov2']:
                chk_auto = getattr(self, 'chk_auto_crop', None) and self.chk_auto_crop.isChecked()
                chk_edge = getattr(self, 'chk_edge_preproc', None) and self.chk_edge_preproc.isChecked()
                if not chk_auto and not chk_edge:
                    msg_parts.append("• Try enabling 'Auto-crop' or 'Edge-enhance'")

            QMessageBox.information(self, "No Results", ''.join(msg_parts))
            return

        # Filter results if "Only decorated" is checked
        original_count = len(results)
        if hasattr(self, 'chk_only_decorated') and self.chk_only_decorated.isChecked():
            filtered_results = []
            for r in results:
                pottery_data = r.get('pottery_data', {})
                exdeco = pottery_data.get('exdeco', '')
                intdeco = pottery_data.get('intdeco', '')
                # Keep only if exdeco OR intdeco has a value (not empty/None)
                has_decoration = (exdeco and str(exdeco).strip()) or (intdeco and str(intdeco).strip())
                if has_decoration:
                    filtered_results.append(r)
            results = filtered_results
            self.label_similarity_status.setText(f"Found {len(results)} decorated (filtered from {original_count})")
        else:
            self.label_similarity_status.setText(f"Found {len(results)} similar items")

        if not results:
            QMessageBox.information(self, "Results", "No decorated pottery found above the threshold.\nTry unchecking 'Only decorated pottery' filter.")
            return

        # Show results dialog
        self.show_similarity_results_dialog(results)

    def on_similarity_search_complete_with_meta(self, results, meta):
        """Handle search completion with metadata (top scores)"""
        self.btn_find_similar.setEnabled(True)

        # Get current settings
        current_threshold = self.slider_similarity_threshold.value()
        top_scores = meta.get('top_scores', [])
        model_name = self.get_similarity_model_name()
        search_type = self.get_similarity_search_type()
        custom_prompt = ''
        if hasattr(self, 'lineEdit_custom_prompt'):
            custom_prompt = self.lineEdit_custom_prompt.text().strip()
        is_global_search = hasattr(self, 'radio_search_global') and self.radio_search_global.isChecked()

        if not results:
            self.label_similarity_status.setText("No similar pottery found")

            # Build detailed message with top scores
            msg_parts = [f"No pottery found above {current_threshold}% threshold."]

            if top_scores:
                max_score = max(top_scores)
                msg_parts.append(f"\n\n📊 Best matches found:")
                msg_parts.append(f"• Top score: {max_score:.1f}%")
                msg_parts.append(f"• Top 5: {', '.join([f'{s:.1f}%' for s in top_scores[:5]])}")

                if max_score < current_threshold:
                    suggested = int(max_score - 5)  # Suggest 5% below best match
                    suggested = max(40, suggested)  # But not below 40%
                    msg_parts.append(f"\n\n💡 Suggestion: Lower threshold to {suggested}% to see results")

            # Model-specific tips
            msg_parts.append("\n\n📌 Tips:")

            if custom_prompt and is_global_search:
                msg_parts.append("• Global text search: scores typically 40-60%")
                msg_parts.append("• Try more specific descriptions")
            elif model_name == 'openai':
                if search_type == 'general':
                    msg_parts.append("• For decorations: use 'Decoration' search type")
                    msg_parts.append("• Build 'Decoration' index for better results")
                msg_parts.append("• Combined search with prompt focuses the analysis")
            elif model_name == 'clip':
                msg_parts.append("• CLIP scores typically 60-85%")
                msg_parts.append("• Try 'Edge-enhance' for line decorations")
                msg_parts.append("• Try 'Auto-crop' to focus on decoration details")
            elif model_name == 'dinov2':
                msg_parts.append("• DINOv2 scores typically 50-80%")
                msg_parts.append("• Best for texture/surface patterns")
                msg_parts.append("• Try 'Edge-enhance' for clearer features")
            elif model_name == 'khutm_clip':
                msg_parts.append("• KhutmML-CLIP is fine-tuned for archaeological pottery")
                msg_parts.append("• Scores typically 55-90% for similar pottery")
                msg_parts.append("• Use 'decoration' search for painted/incised patterns")
                msg_parts.append("• Use 'shape' search for vessel forms")
                msg_parts.append("• Re-index after adding new pottery images")

            # Check preprocessing options
            chk_auto = getattr(self, 'chk_auto_crop', None) and self.chk_auto_crop.isChecked()
            chk_edge = getattr(self, 'chk_edge_preproc', None) and self.chk_edge_preproc.isChecked()
            if not chk_auto and not chk_edge and model_name in ['clip', 'dinov2', 'khutm_clip']:
                msg_parts.append("\n• Enable preprocessing options for better matching")

            QMessageBox.information(self, "No Results", ''.join(msg_parts))
            return

        # If we have results, proceed normally
        self.on_similarity_search_complete(results)

    def on_similarity_error(self, error_msg):
        """Handle search error"""
        self.btn_find_similar.setEnabled(True)
        self.label_similarity_status.setText("Error occurred")
        QMessageBox.warning(self, "Error", f"Similarity search error: {error_msg}")

    def show_similarity_results_dialog(self, results):
        """Show dialog with similarity search results"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Similar Pottery ({len(results)} results)")
        dialog.setMinimumSize(800, 600)

        layout = QVBoxLayout()

        # Info label
        info_label = QLabel("Double-click on an item to navigate to that pottery record")
        layout.addWidget(info_label)

        # Results list with thumbnails
        list_widget = QListWidget()
        list_widget.setViewMode(QListWidget.IconMode)
        list_widget.setIconSize(QSize(150, 150))
        list_widget.setSpacing(10)
        list_widget.setResizeMode(QListWidget.Adjust)
        list_widget.setWordWrap(True)

        # Get config for Cloudinary detection
        conn = Connection()
        thumb_path_config = conn.thumb_path()
        thumb_path_str = thumb_path_config.get('thumb_path', '')
        is_cloudinary = thumb_path_str.lower().startswith('cloudinary://')

        for result in results:
            pottery_data = result.get('pottery_data', {})
            similarity = result.get('similarity_percent', 0)

            # Create item label with more info - show id_number (user-visible ID), not id_rep (DB key)
            label = f"{similarity:.1f}%\n"
            id_number = pottery_data.get('id_number', '') if pottery_data else ''
            label += f"ID: {id_number if id_number else result.get('pottery_id', 'N/A')}\n"
            if pottery_data:
                # Show US
                us_val = pottery_data.get('us', '')
                if us_val:
                    label += f"US: {us_val}\n"
                # Show form and specific_form
                if pottery_data.get('form'):
                    label += f"Form: {pottery_data.get('form', '')}\n"
                if pottery_data.get('specific_form'):
                    label += f"Spec: {pottery_data.get('specific_form', '')}\n"
                # Show decoration
                if pottery_data.get('exdeco'):
                    label += f"Deco: {pottery_data.get('exdeco', '')}"

            item = QListWidgetItem(label)

            # Try to load thumbnail - supports both local and Cloudinary
            image_path = result.get('image_path', '')
            pixmap = None

            if image_path:
                if is_cloudinary and HAS_REMOTE_LOADER:
                    # For Cloudinary: build cloudinary URL from relative path
                    relative_path = result.get('relative_path', '')
                    if relative_path:
                        cloudinary_path = f"{thumb_path_str}/{relative_path}"
                        pixmap = RemoteImageLoader.load_pixmap(cloudinary_path, 150, 150)
                elif os.path.exists(image_path):
                    # Local file
                    pixmap = QPixmap(image_path)
                    if not pixmap.isNull():
                        pixmap = pixmap.scaled(150, 150, Qt.KeepAspectRatio, Qt.SmoothTransformation)

            if pixmap and not pixmap.isNull():
                item.setIcon(QIcon(pixmap))

            item.setData(Qt.UserRole, result)
            list_widget.addItem(item)

        # Double-click to navigate to record
        list_widget.itemDoubleClicked.connect(lambda item: self.navigate_to_pottery(item.data(Qt.UserRole)))

        layout.addWidget(list_widget)

        # Summary statistics
        if results:
            similarities = [r.get('similarity_percent', 0) for r in results]
            avg_sim = sum(similarities) / len(similarities)
            max_sim = max(similarities)
            min_sim = min(similarities)
            summary_label = QLabel(f"Statistics: {len(results)} results | Avg: {avg_sim:.1f}% | Max: {max_sim:.1f}% | Min: {min_sim:.1f}%")
            summary_label.setStyleSheet("font-weight: bold; padding: 5px; background-color: #f0f0f0;")
            layout.addWidget(summary_label)

        # Buttons row
        buttons_layout = QHBoxLayout()

        # Export to Excel button
        btn_export = QPushButton("Export to Excel")
        btn_export.setToolTip("Export results to Excel with thumbnails and chart")
        btn_export.clicked.connect(lambda: self.export_similarity_results(results, thumb_path_str, is_cloudinary))
        buttons_layout.addWidget(btn_export)

        # Show Chart button
        btn_chart = QPushButton("Show Chart")
        btn_chart.setToolTip("Show similarity distribution chart")
        btn_chart.clicked.connect(lambda: self.show_similarity_chart(results))
        buttons_layout.addWidget(btn_chart)

        # Close button
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(dialog.close)
        buttons_layout.addWidget(btn_close)

        layout.addLayout(buttons_layout)

        dialog.setLayout(layout)

        # Store dialog reference
        self._similarity_dialog = dialog

        # Make dialog non-modal so user can interact with main form
        dialog.setModal(False)
        dialog.setAttribute(Qt.WA_DeleteOnClose)  # Clean up when closed
        dialog.show()  # Non-blocking show instead of exec_()

    def navigate_to_pottery(self, result_data):
        """Navigate to a pottery record from search results using id_number"""
        if not result_data:
            print("[SIMILARITY] ERROR: result_data is None or empty")
            return

        pottery_data = result_data.get('pottery_data', {})
        pottery_id_number = pottery_data.get('id_number', None) if pottery_data else None

        print(f"[SIMILARITY] Navigating to pottery id_number={pottery_id_number}")

        if not pottery_id_number:
            print("[SIMILARITY] ERROR: id_number is None or empty")
            return

        # First, load ALL records (like View All button) to ensure record is findable
        print("[SIMILARITY] Loading all records (view_all)...")
        self.empty_fields()
        self.charge_records()  # Reload all records from DB
        self.BROWSE_STATUS = "b"
        self.label_status.setText(self.STATUS_ITEMS[self.BROWSE_STATUS])
        self.SORT_STATUS = "n"
        self.label_sort.setText(self.SORTED_ITEMS[self.SORT_STATUS])

        print(f"[SIMILARITY] Loaded {len(self.DATA_LIST)} records")

        # Now find the record in DATA_LIST using id_number
        found = False
        for i, record in enumerate(self.DATA_LIST):
            record_id_number = str(getattr(record, 'id_number', ''))
            if record_id_number == str(pottery_id_number):
                print(f"[SIMILARITY] Found record at index {i}, id_number={record_id_number}")
                print(f"[SIMILARITY] Record details: id_rep={getattr(record, 'id_rep', 'N/A')}, sito={getattr(record, 'sito', 'N/A')}")

                # Debug: check what's at index i vs index 0
                print(f"[SIMILARITY] DATA_LIST[{i}].id_number = {getattr(self.DATA_LIST[i], 'id_number', 'N/A')}")
                print(f"[SIMILARITY] DATA_LIST[0].id_number = {getattr(self.DATA_LIST[0], 'id_number', 'N/A')}")

                self.REC_CORR = i
                self.REC_TOT = len(self.DATA_LIST)
                self.DATA_LIST_REC_CORR = self.DATA_LIST[i]  # Set current record reference
                self.DATA_LIST_REC_TEMP = self.DATA_LIST[i]
                self.set_rec_counter(self.REC_TOT, self.REC_CORR + 1)

                print(f"[SIMILARITY] Before fill_fields: REC_CORR={self.REC_CORR}")
                self.fill_fields(self.REC_CORR)  # Use REC_CORR explicitly

                found = True
                # Verify what was actually filled
                filled_id = self.lineEdit_id_number.text() if hasattr(self, 'lineEdit_id_number') else 'N/A'
                print(f"[SIMILARITY] After fill_fields: lineEdit_id_number='{filled_id}'")
                print(f"[SIMILARITY] Navigation complete. REC_CORR={self.REC_CORR}, showing record {self.REC_CORR + 1}/{self.REC_TOT}")
                break

        if not found:
            print(f"[SIMILARITY] WARNING: pottery id_number={pottery_id_number} not found even after view_all (len={len(self.DATA_LIST)})")
            # Debug: print first 5 id_numbers to check format
            for j, r in enumerate(self.DATA_LIST[:5]):
                print(f"[SIMILARITY] Sample record {j}: id_number='{getattr(r, 'id_number', 'N/A')}'")
            QMessageBox.warning(self, "Not Found",
                f"Pottery ID {pottery_id_number} not found in database.")

        # Keep dialog open and on top - don't close it
        if hasattr(self, '_similarity_dialog') and self._similarity_dialog:
            self._similarity_dialog.raise_()
            self._similarity_dialog.activateWindow()

    def export_similarity_results(self, results, thumb_path_str, is_cloudinary):
        """Export similarity results to Excel with thumbnails and embedded chart"""
        if not results:
            QMessageBox.warning(self, "Export", "No results to export")
            return

        # Ask for save location
        from qgis.PyQt.QtWidgets import QFileDialog
        import datetime

        default_name = f"pottery_similarity_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel Report", default_name, "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            # Try to use openpyxl for Excel with images
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.chart import BarChart, Reference
            from openpyxl.utils import get_column_letter
            import tempfile
            import urllib.request

            wb = Workbook()
            ws = wb.active
            ws.title = "Similarity Results"

            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = ["Thumbnail", "Similarity %", "ID Number", "Site", "Area", "US",
                      "Form", "Specific Form", "Ext. Decoration", "Int. Decoration", "Specific Shape"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # Set column widths
            ws.column_dimensions['A'].width = 15  # Thumbnail
            ws.column_dimensions['B'].width = 12  # Similarity
            ws.column_dimensions['C'].width = 12  # ID
            ws.column_dimensions['D'].width = 15  # Site
            ws.column_dimensions['E'].width = 8   # Area
            ws.column_dimensions['F'].width = 8   # US
            ws.column_dimensions['G'].width = 15  # Form
            ws.column_dimensions['H'].width = 15  # Specific Form
            ws.column_dimensions['I'].width = 20  # Ext Deco
            ws.column_dimensions['J'].width = 20  # Int Deco
            ws.column_dimensions['K'].width = 15  # Specific Shape

            # Data rows
            row_num = 2
            temp_files = []  # Track temp files for cleanup

            for result in results:
                pottery_data = result.get('pottery_data', {})
                similarity = result.get('similarity_percent', 0)

                # Set row height for thumbnail
                ws.row_dimensions[row_num].height = 80

                # Try to add thumbnail
                image_path = result.get('image_path', '')
                thumb_added = False

                if image_path:
                    try:
                        local_path = None
                        if is_cloudinary and HAS_REMOTE_LOADER:
                            # Download from Cloudinary
                            relative_path = result.get('relative_path', '')
                            if relative_path:
                                cloudinary_url = RemoteImageLoader.cloudinary_to_url(f"{thumb_path_str}/{relative_path}")
                                if cloudinary_url:
                                    temp_file = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                                    temp_files.append(temp_file.name)
                                    urllib.request.urlretrieve(cloudinary_url, temp_file.name)
                                    local_path = temp_file.name
                        elif os.path.exists(image_path):
                            local_path = image_path

                        if local_path and os.path.exists(local_path):
                            img = XLImage(local_path)
                            img.width = 80
                            img.height = 80
                            ws.add_image(img, f'A{row_num}')
                            thumb_added = True
                    except Exception as e:
                        print(f"[EXPORT] Could not add thumbnail: {e}")

                if not thumb_added:
                    ws.cell(row=row_num, column=1, value="[No image]")

                # Data cells
                ws.cell(row=row_num, column=2, value=round(similarity, 1))
                ws.cell(row=row_num, column=3, value=pottery_data.get('id_number', ''))
                ws.cell(row=row_num, column=4, value=pottery_data.get('sito', ''))
                ws.cell(row=row_num, column=5, value=pottery_data.get('area', ''))
                ws.cell(row=row_num, column=6, value=pottery_data.get('us', ''))
                ws.cell(row=row_num, column=7, value=pottery_data.get('form', ''))
                ws.cell(row=row_num, column=8, value=pottery_data.get('specific_form', ''))
                ws.cell(row=row_num, column=9, value=pottery_data.get('exdeco', ''))
                ws.cell(row=row_num, column=10, value=pottery_data.get('intdeco', ''))
                ws.cell(row=row_num, column=11, value=pottery_data.get('specific_shape', ''))

                # Apply borders to data cells
                for col in range(1, 12):
                    ws.cell(row=row_num, column=col).border = thin_border
                    ws.cell(row=row_num, column=col).alignment = Alignment(vertical="center")

                row_num += 1

            # Add summary statistics
            row_num += 1
            ws.cell(row=row_num, column=1, value="STATISTICS").font = Font(bold=True)
            row_num += 1

            similarities = [r.get('similarity_percent', 0) for r in results]
            ws.cell(row=row_num, column=1, value="Total Results:")
            ws.cell(row=row_num, column=2, value=len(results))
            row_num += 1
            ws.cell(row=row_num, column=1, value="Average Similarity:")
            ws.cell(row=row_num, column=2, value=round(sum(similarities)/len(similarities), 1))
            row_num += 1
            ws.cell(row=row_num, column=1, value="Max Similarity:")
            ws.cell(row=row_num, column=2, value=round(max(similarities), 1))
            row_num += 1
            ws.cell(row=row_num, column=1, value="Min Similarity:")
            ws.cell(row=row_num, column=2, value=round(min(similarities), 1))

            # Create chart sheet
            ws_chart = wb.create_sheet("Similarity Chart")

            # Create histogram data (similarity distribution)
            bins = list(range(50, 105, 5))  # 50-55, 55-60, ..., 95-100
            bin_counts = [0] * (len(bins) - 1)
            for sim in similarities:
                for i in range(len(bins) - 1):
                    if bins[i] <= sim < bins[i+1]:
                        bin_counts[i] += 1
                        break
                else:
                    if sim >= 100:
                        bin_counts[-1] += 1

            # Write chart data
            ws_chart['A1'] = "Similarity Range"
            ws_chart['B1'] = "Count"
            for i, (start, count) in enumerate(zip(bins[:-1], bin_counts), 2):
                ws_chart[f'A{i}'] = f"{start}-{start+5}%"
                ws_chart[f'B{i}'] = count

            # Create bar chart
            chart = BarChart()
            chart.title = "Similarity Distribution"
            chart.x_axis.title = "Similarity Range"
            chart.y_axis.title = "Count"
            chart.style = 10

            data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(bins))
            cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(bins))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            chart.width = 15
            chart.height = 10

            ws_chart.add_chart(chart, "D2")

            # Save workbook
            wb.save(file_path)

            # Cleanup temp files
            for temp_file in temp_files:
                try:
                    os.unlink(temp_file)
                except:
                    pass

            QMessageBox.information(self, "Export Complete",
                f"Results exported to:\n{file_path}\n\n"
                f"Contains {len(results)} pottery records with thumbnails and similarity chart.")

        except ImportError:
            # Fallback to CSV if openpyxl not available
            QMessageBox.warning(self, "Missing Library",
                "openpyxl is required for Excel export with images.\n"
                "Install with: pip install openpyxl\n\n"
                "Falling back to CSV export.")
            self.export_similarity_to_csv(results, file_path.replace('.xlsx', '.csv'))

        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {str(e)}")
            print(f"[EXPORT] Error: {e}")
            import traceback
            traceback.print_exc()

    def export_similarity_to_csv(self, results, file_path):
        """Fallback CSV export without thumbnails"""
        import csv

        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Similarity %", "ID Number", "Site", "Area", "US",
                           "Form", "Specific Form", "Ext. Decoration", "Int. Decoration", "Specific Shape"])

            for result in results:
                pottery_data = result.get('pottery_data', {})
                writer.writerow([
                    round(result.get('similarity_percent', 0), 1),
                    pottery_data.get('id_number', ''),
                    pottery_data.get('sito', ''),
                    pottery_data.get('area', ''),
                    pottery_data.get('us', ''),
                    pottery_data.get('form', ''),
                    pottery_data.get('specific_form', ''),
                    pottery_data.get('exdeco', ''),
                    pottery_data.get('intdeco', ''),
                    pottery_data.get('specific_shape', '')
                ])

        QMessageBox.information(self, "Export Complete", f"Results exported to:\n{file_path}")

    def show_similarity_chart(self, results):
        """Show similarity distribution chart in a popup dialog"""
        if not results:
            return

        try:
            import matplotlib
            matplotlib.use('Qt5Agg')
            from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
            from matplotlib.figure import Figure
            import numpy as np

            similarities = [r.get('similarity_percent', 0) for r in results]

            # Create dialog
            chart_dialog = QDialog(self)
            chart_dialog.setWindowTitle("Similarity Distribution")
            chart_dialog.setMinimumSize(600, 450)

            layout = QVBoxLayout()

            # Create matplotlib figure
            fig = Figure(figsize=(8, 5), dpi=100)
            canvas = FigureCanvas(fig)

            # Create subplot
            ax = fig.add_subplot(111)

            # Histogram
            bins = list(range(50, 105, 5))
            ax.hist(similarities, bins=bins, edgecolor='black', alpha=0.7, color='#4472C4')

            ax.set_xlabel('Similarity (%)', fontsize=11)
            ax.set_ylabel('Count', fontsize=11)
            ax.set_title(f'Similarity Distribution ({len(results)} results)', fontsize=12, fontweight='bold')

            # Add statistics text
            avg_sim = sum(similarities) / len(similarities)
            stats_text = f'Avg: {avg_sim:.1f}%  |  Max: {max(similarities):.1f}%  |  Min: {min(similarities):.1f}%'
            ax.text(0.5, 0.95, stats_text, transform=ax.transAxes, ha='center',
                   fontsize=10, bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))

            ax.grid(axis='y', alpha=0.3)
            fig.tight_layout()

            layout.addWidget(canvas)

            # Close button
            btn_close = QPushButton("Close")
            btn_close.clicked.connect(chart_dialog.close)
            layout.addWidget(btn_close)

            chart_dialog.setLayout(layout)
            chart_dialog.exec_()

        except ImportError as e:
            QMessageBox.warning(self, "Missing Library",
                f"matplotlib is required for charts.\n"
                f"Install with: pip install matplotlib\n\n"
                f"Error: {e}")

    def on_build_index_clicked(self):
        """Handle build index button click"""
        if not HAS_SIMILARITY_SEARCH:
            QMessageBox.warning(self, "Error", "Similarity search module not available")
            return

        model_name = self.get_similarity_model_name()
        search_type = self.get_similarity_search_type()

        reply = QMessageBox.question(
            self,
            "Build Index",
            f"Build similarity index for {model_name} ({search_type})?\n\n"
            "This may take several minutes depending on the number of images.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        # Initialize engine if needed
        if self.similarity_engine is None:
            self.similarity_engine = PotterySimilaritySearchEngine(self.DB_MANAGER)

        # Update UI
        self.btn_build_index.setEnabled(False)
        self.btn_find_similar.setEnabled(False)
        self.progress_similarity.setVisible(True)
        self.progress_similarity.setValue(0)
        self.label_similarity_status.setText("Building index...")

        # Create worker
        self.similarity_worker = PotterySimilarityWorker(
            self.similarity_engine,
            'build_index',
            model_name=model_name,
            search_type=search_type
        )
        self.similarity_worker.index_progress.connect(self.on_index_progress)
        self.similarity_worker.operation_complete.connect(self.on_index_complete)
        self.similarity_worker.error_occurred.connect(self.on_similarity_error)
        self.similarity_worker.start()

    def on_index_progress(self, current, total, message):
        """Update progress during index building"""
        if total > 0:
            self.progress_similarity.setMaximum(total)
            self.progress_similarity.setValue(current)
        self.label_similarity_status.setText(message)

    def on_index_complete(self, message):
        """Handle index building completion"""
        self.btn_build_index.setEnabled(True)
        self.btn_find_similar.setEnabled(True)
        self.progress_similarity.setVisible(False)
        self.label_similarity_status.setText(message)
        QMessageBox.information(self, "Index Built", message)

    def on_export_indexes_clicked(self):
        """Export all similarity indexes to a ZIP file for sharing"""
        import zipfile
        import datetime
        from qgis.PyQt.QtWidgets import QFileDialog

        index_dir = os.path.expanduser('~/pyarchinit/bin/pottery_similarity')

        if not os.path.exists(index_dir):
            QMessageBox.warning(self, "No Indexes",
                "No similarity indexes found.\nBuild indexes first before exporting.")
            return

        # Find all index files
        index_files = []
        for f in os.listdir(index_dir):
            if f.endswith('.faiss') or f.endswith('_mapping.pkl'):
                index_files.append(f)

        if not index_files:
            QMessageBox.warning(self, "No Indexes",
                "No similarity indexes found in the index directory.")
            return

        # Show what will be exported
        file_list = "\n".join(f"  - {f}" for f in sorted(index_files))
        reply = QMessageBox.question(
            self,
            "Export Indexes",
            f"Export the following {len(index_files)} files?\n\n{file_list}\n\n"
            "These indexes can be imported on other PCs with the same database.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        # Ask for save location
        default_name = f"pottery_similarity_indexes_{datetime.datetime.now().strftime('%Y%m%d')}.zip"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Indexes", default_name, "ZIP Files (*.zip)"
        )

        if not file_path:
            return

        try:
            with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for f in index_files:
                    full_path = os.path.join(index_dir, f)
                    zipf.write(full_path, f)

            # Get file size
            size_mb = os.path.getsize(file_path) / (1024 * 1024)

            QMessageBox.information(self, "Export Complete",
                f"Indexes exported successfully!\n\n"
                f"File: {file_path}\n"
                f"Size: {size_mb:.1f} MB\n"
                f"Files: {len(index_files)}\n\n"
                "You can share this ZIP with other users who have the same database.")

        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export indexes:\n{str(e)}")

    def on_import_indexes_clicked(self):
        """Import similarity indexes from a ZIP file"""
        import zipfile
        from qgis.PyQt.QtWidgets import QFileDialog

        index_dir = os.path.expanduser('~/pyarchinit/bin/pottery_similarity')

        # Ask for ZIP file
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import Indexes", "", "ZIP Files (*.zip)"
        )

        if not file_path:
            return

        try:
            # Check ZIP contents first
            with zipfile.ZipFile(file_path, 'r') as zipf:
                contents = zipf.namelist()

                # Verify it contains index files
                faiss_files = [f for f in contents if f.endswith('.faiss')]
                mapping_files = [f for f in contents if f.endswith('_mapping.pkl')]

                if not faiss_files:
                    QMessageBox.warning(self, "Invalid ZIP",
                        "This ZIP file does not contain any similarity index files (.faiss)")
                    return

                # Show what will be imported
                file_list = "\n".join(f"  - {f}" for f in sorted(contents))

                # Check for existing files
                existing = []
                for f in contents:
                    if os.path.exists(os.path.join(index_dir, f)):
                        existing.append(f)

                warning_msg = ""
                if existing:
                    warning_msg = f"\n\nWARNING: {len(existing)} existing files will be overwritten!"

                reply = QMessageBox.question(
                    self,
                    "Import Indexes",
                    f"Import the following {len(contents)} files?\n\n{file_list}{warning_msg}\n\n"
                    "Make sure these indexes were created with the SAME database!",
                    QMessageBox.Yes | QMessageBox.No
                )

                if reply != QMessageBox.Yes:
                    return

                # Create directory if needed
                os.makedirs(index_dir, exist_ok=True)

                # Extract files
                zipf.extractall(index_dir)

            QMessageBox.information(self, "Import Complete",
                f"Indexes imported successfully!\n\n"
                f"FAISS indexes: {len(faiss_files)}\n"
                f"Mapping files: {len(mapping_files)}\n\n"
                "The indexes are now ready to use.")

            # Update status
            self.label_similarity_status.setText(f"Imported {len(faiss_files)} indexes")

        except zipfile.BadZipFile:
            QMessageBox.warning(self, "Invalid File", "The selected file is not a valid ZIP file.")
        except Exception as e:
            QMessageBox.warning(self, "Import Error", f"Failed to import indexes:\n{str(e)}")

    def on_update_index_clicked(self):
        """Update existing indexes with new/modified/deleted images"""
        from modules.utility.pottery_similarity.similarity_search import (
            PotterySimilaritySearchEngine, PotterySimilarityWorker
        )

        # Ask user what to update
        reply = QMessageBox.question(
            self,
            "Update Indexes",
            "Update all existing indexes?\n\n"
            "This will:\n"
            "• Add new images not yet indexed\n"
            "• Update embeddings for modified images\n"
            "• Remove embeddings for deleted images\n\n"
            "Only indexes that already exist will be updated.\n"
            "(Use 'Build Index' to create new indexes)",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        # Disable buttons during update
        self.btn_find_similar.setEnabled(False)
        self.btn_build_index.setEnabled(False)
        self.btn_update_index.setEnabled(False)
        self.progress_similarity.setVisible(True)
        self.progress_similarity.setMaximum(100)
        self.label_similarity_status.setText("Analyzing changes...")

        # Create search engine
        search_engine = PotterySimilaritySearchEngine(self.DB_MANAGER)

        # Create worker for updating all indexes
        self.similarity_worker = PotterySimilarityWorker(
            search_engine,
            'update_all_indexes'
        )

        # Connect signals
        self.similarity_worker.index_progress.connect(self.on_index_progress)
        self.similarity_worker.operation_complete.connect(self.on_update_complete)
        self.similarity_worker.error_occurred.connect(self.on_similarity_error)

        # Start worker
        self.similarity_worker.start()

    def on_update_complete(self, message):
        """Handle update completion"""
        self.btn_build_index.setEnabled(True)
        self.btn_find_similar.setEnabled(True)
        self.btn_update_index.setEnabled(True)
        self.progress_similarity.setVisible(False)
        self.label_similarity_status.setText(message)
        QMessageBox.information(self, "Update Complete", message)

    def on_train_khutm_clicked(self):
        """Open training dialog for KhutmML-CLIP model"""
        dialog = KhutmTrainingDialog(self.DB_MANAGER, self)
        dialog.training_complete.connect(self.on_khutm_training_complete)
        dialog.exec_()

    def on_prepare_dataset_clicked(self):
        """Prepare dataset for KhutmML training"""
        dialog = DatasetPreparationDialog(self.DB_MANAGER, self)
        dialog.exec_()

    def on_khutm_training_complete(self, success, message):
        """Handle training completion"""
        if success:
            # Ask if user wants to rebuild indexes automatically
            reply = QMessageBox.question(
                self, "Training Complete",
                f"KhutmML-CLIP training completed successfully!\n\n{message}\n\n"
                "Do you want to rebuild the similarity indexes now?\n\n"
                "(This will index all pottery images with the new model)",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )

            if reply == QMessageBox.Yes:
                self.run_khutm_indexing_pipeline()
            else:
                QMessageBox.information(self, "Training Complete",
                    "You can rebuild indexes later using the 'Build Index' button\n"
                    "with 'KhutmML-CLIP (Fine-tuned)' model selected.")
        else:
            QMessageBox.warning(self, "Training Failed", f"Training failed:\n{message}")

    def run_khutm_indexing_pipeline(self):
        """Run the KhutmML indexing pipeline after training"""
        import subprocess

        venv_python = os.path.expanduser('~/pyarchinit/bin/pottery_venv/bin/python')
        indexer_script = os.path.expanduser('~/pyarchinit/bin/khutm_clip_indexer.py')
        faiss_script = os.path.expanduser('~/pyarchinit/bin/import_embeddings_to_faiss.py')
        model_dir = os.path.expanduser('~/pyarchinit/bin/models/khutm_clip')

        # Check scripts exist
        if not os.path.exists(indexer_script) or not os.path.exists(faiss_script):
            QMessageBox.warning(self, "Scripts Not Found",
                "Indexing scripts not found. Please rebuild indexes manually.")
            return

        self.label_similarity_status.setText("Indexing with new model...")
        self.progress_similarity.setVisible(True)
        self.progress_similarity.setMaximum(0)  # Indeterminate
        QApplication.processEvents()

        # Clean environment
        clean_env = os.environ.copy()
        for var in ['PYTHONHOME', 'PYTHONPATH', 'PYTHONEXECUTABLE']:
            clean_env.pop(var, None)

        search_types = ['general', 'decoration', 'shape']
        errors = []

        try:
            for i, search_type in enumerate(search_types):
                self.label_similarity_status.setText(f"Indexing {search_type} ({i+1}/3)...")
                QApplication.processEvents()

                # Run indexer
                cmd = [venv_python, indexer_script,
                       '--search-type', search_type,
                       '--model-dir', model_dir]

                result = subprocess.run(cmd, capture_output=True, text=True,
                                        timeout=600, env=clean_env)

                if result.returncode != 0:
                    errors.append(f"{search_type}: {result.stderr[:200]}")

            # Convert to FAISS
            self.label_similarity_status.setText("Converting to FAISS format...")
            QApplication.processEvents()

            cmd = [venv_python, faiss_script, '--search-type', 'all']
            result = subprocess.run(cmd, capture_output=True, text=True,
                                    timeout=300, env=clean_env)

            if result.returncode != 0:
                errors.append(f"FAISS conversion: {result.stderr[:200]}")

        except subprocess.TimeoutExpired:
            errors.append("Indexing timed out")
        except Exception as e:
            errors.append(str(e))

        self.progress_similarity.setVisible(False)

        if errors:
            self.label_similarity_status.setText("Indexing completed with errors")
            QMessageBox.warning(self, "Indexing Errors",
                f"Indexing completed with some errors:\n\n" + "\n".join(errors))
        else:
            self.label_similarity_status.setText("Indexing complete!")
            QMessageBox.information(self, "Indexing Complete",
                "All indexes rebuilt successfully!\n\n"
                "You can now use KhutmML-CLIP for similarity search.")

    def on_export_khutm_model_clicked(self):
        """Export the trained KhutmML-CLIP model to a ZIP file"""
        import zipfile
        from datetime import datetime

        model_dir = os.path.expanduser('~/pyarchinit/bin/models/khutm_clip')

        # Check if model exists
        if not os.path.exists(model_dir):
            QMessageBox.warning(self, "No Model Found",
                "No trained KhutmML-CLIP model found.\n\n"
                "Train a model first using the 'Train KhutmML' button.")
            return

        # Check for model files
        model_files = []
        for f in os.listdir(model_dir):
            if f.endswith(('.pt', '.json', '.txt', '.pkl')):
                model_files.append(f)

        if not model_files:
            QMessageBox.warning(self, "No Model Files",
                f"No model files found in:\n{model_dir}\n\n"
                "Train a model first using the 'Train KhutmML' button.")
            return

        # Ask for save location
        default_name = f"khutm_clip_model_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export KhutmML-CLIP Model",
            os.path.join(os.path.expanduser('~'), default_name),
            "ZIP Archive (*.zip)"
        )

        if not save_path:
            return

        try:
            self.label_similarity_status.setText("Exporting model...")
            QApplication.processEvents()

            # Create ZIP file with model files
            with zipfile.ZipFile(save_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for f in model_files:
                    file_path = os.path.join(model_dir, f)
                    zf.write(file_path, f)

                # Add a manifest with export info
                manifest = {
                    'export_date': datetime.now().isoformat(),
                    'model_type': 'khutm_clip',
                    'files': model_files
                }
                import json
                zf.writestr('manifest.json', json.dumps(manifest, indent=2))

            self.label_similarity_status.setText("Model exported successfully")
            QMessageBox.information(self, "Export Complete",
                f"KhutmML-CLIP model exported to:\n{save_path}\n\n"
                f"Files included: {len(model_files)}")

        except Exception as e:
            self.label_similarity_status.setText("Export failed")
            QMessageBox.critical(self, "Export Error", f"Failed to export model:\n{str(e)}")

    def on_import_khutm_model_clicked(self):
        """Import a KhutmML-CLIP model from a ZIP file"""
        import zipfile
        import shutil

        model_dir = os.path.expanduser('~/pyarchinit/bin/models/khutm_clip')

        # Ask for ZIP file
        zip_path, _ = QFileDialog.getOpenFileName(
            self,
            "Import KhutmML-CLIP Model",
            os.path.expanduser('~'),
            "ZIP Archive (*.zip)"
        )

        if not zip_path:
            return

        # Confirm if model already exists
        if os.path.exists(model_dir) and os.listdir(model_dir):
            reply = QMessageBox.question(
                self, "Confirm Import",
                "A trained model already exists.\n\n"
                "Do you want to replace it with the imported model?\n\n"
                "(The existing model will be backed up)",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return

        try:
            self.label_similarity_status.setText("Importing model...")
            QApplication.processEvents()

            # Verify ZIP contains expected files
            with zipfile.ZipFile(zip_path, 'r') as zf:
                file_list = zf.namelist()

                # Check for model files (at least one .pt file expected)
                pt_files = [f for f in file_list if f.endswith('.pt')]
                if not pt_files:
                    QMessageBox.warning(self, "Invalid Model Archive",
                        "The ZIP file does not contain valid model files (.pt).\n\n"
                        "Please select a valid KhutmML-CLIP model archive.")
                    self.label_similarity_status.setText("Import cancelled")
                    return

                # Backup existing model if present
                if os.path.exists(model_dir) and os.listdir(model_dir):
                    from datetime import datetime
                    backup_dir = model_dir + f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    shutil.move(model_dir, backup_dir)

                # Create model directory
                os.makedirs(model_dir, exist_ok=True)

                # Extract files
                for f in file_list:
                    if f != 'manifest.json':  # Skip manifest
                        zf.extract(f, model_dir)

            self.label_similarity_status.setText("Model imported successfully")
            QMessageBox.information(self, "Import Complete",
                f"KhutmML-CLIP model imported successfully!\n\n"
                f"Files imported: {len(pt_files)} model file(s)\n\n"
                "You may need to rebuild the similarity indexes to use the new model.")

        except zipfile.BadZipFile:
            self.label_similarity_status.setText("Import failed")
            QMessageBox.critical(self, "Import Error", "The selected file is not a valid ZIP archive.")
        except Exception as e:
            self.label_similarity_status.setText("Import failed")
            QMessageBox.critical(self, "Import Error", f"Failed to import model:\n{str(e)}")


class KhutmTrainingDialog(QDialog):
    """Dialog for fine-tuning the KhutmML-CLIP model"""
    training_complete = pyqtSignal(bool, str)  # success, message

    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.training_process = None
        self.training_output = ""
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Train KhutmML-CLIP Model")
        self.setMinimumSize(600, 500)

        layout = QVBoxLayout(self)

        # Header
        header_label = QLabel(
            "<h3>Fine-tune CLIP on Your Pottery Dataset</h3>"
            "<p>Train a specialized model that better recognizes your archaeological pottery.</p>"
        )
        header_label.setWordWrap(True)
        layout.addWidget(header_label)

        # Data source selection
        source_group = QGroupBox("Training Data Source")
        source_layout = QVBoxLayout()

        self.radio_database = QRadioButton("Use pottery images from database")
        self.radio_database.setChecked(True)
        self.radio_database.setToolTip("Uses all pottery images linked in the database for training")
        source_layout.addWidget(self.radio_database)

        self.radio_folder = QRadioButton("Use images from folder")
        self.radio_folder.setToolTip("Select a folder with pottery images organized by category")
        source_layout.addWidget(self.radio_folder)

        folder_layout = QHBoxLayout()
        self.lineEdit_folder = QLineEdit()
        self.lineEdit_folder.setEnabled(False)
        self.lineEdit_folder.setPlaceholderText("Select folder with images...")
        folder_layout.addWidget(self.lineEdit_folder)
        self.btn_browse_folder = QPushButton("Browse...")
        self.btn_browse_folder.setEnabled(False)
        self.btn_browse_folder.clicked.connect(self.browse_training_folder)
        folder_layout.addWidget(self.btn_browse_folder)
        source_layout.addLayout(folder_layout)

        self.radio_folder.toggled.connect(lambda checked: self.lineEdit_folder.setEnabled(checked))
        self.radio_folder.toggled.connect(lambda checked: self.btn_browse_folder.setEnabled(checked))

        source_group.setLayout(source_layout)
        layout.addWidget(source_group)

        # Training parameters
        params_group = QGroupBox("Training Parameters")
        params_layout = QGridLayout()

        params_layout.addWidget(QLabel("Epochs:"), 0, 0)
        self.spinBox_epochs = QSpinBox()
        self.spinBox_epochs.setRange(1, 100)
        self.spinBox_epochs.setValue(10)
        self.spinBox_epochs.setToolTip("Number of training epochs (more = longer but potentially better)")
        params_layout.addWidget(self.spinBox_epochs, 0, 1)

        params_layout.addWidget(QLabel("Batch Size:"), 0, 2)
        self.spinBox_batch = QSpinBox()
        self.spinBox_batch.setRange(4, 64)
        self.spinBox_batch.setValue(16)
        self.spinBox_batch.setToolTip("Batch size (lower = less memory, higher = faster training)")
        params_layout.addWidget(self.spinBox_batch, 0, 3)

        params_layout.addWidget(QLabel("Learning Rate:"), 1, 0)
        self.comboBox_lr = QComboBox()
        self.comboBox_lr.addItems(['1e-3', '1e-4', '1e-5', '5e-5'])
        self.comboBox_lr.setCurrentIndex(1)  # Default 1e-4
        self.comboBox_lr.setToolTip("Learning rate (smaller = safer, larger = faster)")
        params_layout.addWidget(self.comboBox_lr, 1, 1)

        params_layout.addWidget(QLabel("Min images per class:"), 1, 2)
        self.spinBox_min_images = QSpinBox()
        self.spinBox_min_images.setRange(2, 50)
        self.spinBox_min_images.setValue(5)
        self.spinBox_min_images.setToolTip("Minimum images required per pottery type for training")
        params_layout.addWidget(self.spinBox_min_images, 1, 3)

        params_group.setLayout(params_layout)
        layout.addWidget(params_group)

        # Output location
        output_group = QGroupBox("Model Output")
        output_layout = QHBoxLayout()
        self.lineEdit_output = QLineEdit()
        self.lineEdit_output.setText(os.path.expanduser('~/pyarchinit/bin/models/khutm_clip'))
        self.lineEdit_output.setToolTip("Directory where trained model will be saved")
        output_layout.addWidget(self.lineEdit_output)
        self.btn_browse_output = QPushButton("Browse...")
        self.btn_browse_output.clicked.connect(self.browse_output_folder)
        output_layout.addWidget(self.btn_browse_output)
        output_group.setLayout(output_layout)
        layout.addWidget(output_group)

        # Progress section
        progress_group = QGroupBox("Training Progress")
        progress_layout = QVBoxLayout()

        self.label_status = QLabel("Ready to train")
        self.label_status.setStyleSheet("color: gray; font-style: italic;")
        progress_layout.addWidget(self.label_status)

        self.progressBar = QProgressBar()
        self.progressBar.setVisible(False)
        progress_layout.addWidget(self.progressBar)

        self.textEdit_log = QTextEdit()
        self.textEdit_log.setReadOnly(True)
        self.textEdit_log.setMaximumHeight(150)
        self.textEdit_log.setVisible(False)
        progress_layout.addWidget(self.textEdit_log)

        progress_group.setLayout(progress_layout)
        layout.addWidget(progress_group)

        # Buttons
        btn_layout = QHBoxLayout()
        self.btn_train = QPushButton("Start Training")
        self.btn_train.clicked.connect(self.start_training)
        btn_layout.addWidget(self.btn_train)

        self.btn_cancel = QPushButton("Cancel")
        self.btn_cancel.clicked.connect(self.cancel_or_close)
        btn_layout.addWidget(self.btn_cancel)

        layout.addLayout(btn_layout)

    def browse_training_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Training Folder")
        if folder:
            self.lineEdit_folder.setText(folder)

    def browse_output_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.lineEdit_output.setText(folder)

    def start_training(self):
        """Start the training process"""
        import subprocess
        import json

        # Validate
        if self.radio_folder.isChecked() and not self.lineEdit_folder.text():
            QMessageBox.warning(self, "Error", "Please select a training folder")
            return

        output_dir = self.lineEdit_output.text()
        if not output_dir:
            QMessageBox.warning(self, "Error", "Please specify output directory")
            return

        # Prepare command
        trainer_script = os.path.expanduser('~/pyarchinit/bin/khutm_clip_trainer.py')
        python_exec = os.path.expanduser('~/pyarchinit/bin/pottery_venv/bin/python')

        if not os.path.exists(trainer_script):
            QMessageBox.warning(self, "Error",
                f"Training script not found:\n{trainer_script}")
            return

        if not os.path.exists(python_exec):
            QMessageBox.warning(self, "Error",
                f"Python environment not found:\n{python_exec}")
            return

        # Build command arguments
        cmd = [
            python_exec,
            trainer_script,
            '--epochs', str(self.spinBox_epochs.value()),
            '--batch-size', str(self.spinBox_batch.value()),
            '--learning-rate', self.comboBox_lr.currentText(),
            '--output-dir', output_dir,
            '--min-images-per-class', str(self.spinBox_min_images.value())
        ]

        if self.radio_database.isChecked():
            # Get database path from connection settings
            from ..modules.db.pyarchinit_conn_strings import Connection
            conn = Connection()
            thumb_path = conn.thumb_path().get('thumb_path', '')
            cmd.extend(['--data-source', 'database', '--thumb-path', thumb_path])
        else:
            cmd.extend(['--data-source', 'folder', '--data-path', self.lineEdit_folder.text()])

        # Update UI
        self.btn_train.setEnabled(False)
        self.progressBar.setVisible(True)
        self.progressBar.setRange(0, 0)  # Indeterminate
        self.textEdit_log.setVisible(True)
        self.textEdit_log.clear()
        self.label_status.setText("Training in progress...")
        self.label_status.setStyleSheet("color: blue;")

        # Create timer for reading output
        self.output_timer = QTimer(self)
        self.output_timer.timeout.connect(self.read_training_output)

        try:
            # Clean environment to avoid QGIS Python interference
            clean_env = os.environ.copy()
            # Remove QGIS-specific Python paths that interfere with virtualenv
            keys_to_remove = ['PYTHONHOME', 'PYTHONPATH', 'PYTHONEXECUTABLE']
            for key in keys_to_remove:
                clean_env.pop(key, None)

            # Start subprocess with clean environment
            self.training_process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1,
                env=clean_env
            )

            self.output_timer.start(500)  # Check every 500ms

            # Start monitoring in background
            self.monitor_timer = QTimer(self)
            self.monitor_timer.timeout.connect(self.check_training_status)
            self.monitor_timer.start(1000)

        except Exception as e:
            self.label_status.setText(f"Error starting training: {str(e)}")
            self.label_status.setStyleSheet("color: red;")
            self.btn_train.setEnabled(True)

    def read_training_output(self):
        """Read output from training process"""
        if self.training_process and self.training_process.stdout:
            try:
                # Non-blocking read
                import select
                if hasattr(select, 'select'):
                    ready, _, _ = select.select([self.training_process.stdout], [], [], 0)
                    if ready:
                        line = self.training_process.stdout.readline()
                        if line:
                            self.training_output += line
                            self.textEdit_log.append(line.strip())

                            # Parse progress if available
                            if 'Epoch' in line and '/' in line:
                                try:
                                    parts = line.split('Epoch')[1].split('/')
                                    current = int(parts[0].strip())
                                    total = int(parts[1].split()[0])
                                    self.progressBar.setRange(0, total)
                                    self.progressBar.setValue(current)
                                except:
                                    pass
            except:
                pass

    def check_training_status(self):
        """Check if training is complete"""
        if self.training_process:
            retcode = self.training_process.poll()
            if retcode is not None:
                # Training finished
                self.output_timer.stop()
                self.monitor_timer.stop()

                # Read any remaining output
                try:
                    remaining = self.training_process.stdout.read()
                    if remaining:
                        self.training_output += remaining
                        self.textEdit_log.append(remaining.strip())
                except:
                    pass

                self.progressBar.setRange(0, 100)
                self.progressBar.setValue(100)
                self.btn_train.setEnabled(True)

                if retcode == 0:
                    self.label_status.setText("Training completed successfully!")
                    self.label_status.setStyleSheet("color: green;")
                    self.training_complete.emit(True, "Model saved to " + self.lineEdit_output.text())
                else:
                    self.label_status.setText(f"Training failed (exit code {retcode})")
                    self.label_status.setStyleSheet("color: red;")
                    self.training_complete.emit(False, self.training_output[-500:])

                self.training_process = None

    def cancel_or_close(self):
        """Cancel training or close dialog"""
        if self.training_process:
            reply = QMessageBox.question(
                self, "Cancel Training",
                "Training is in progress. Are you sure you want to cancel?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                try:
                    self.training_process.terminate()
                    self.training_process.wait(timeout=5)
                except:
                    self.training_process.kill()
                self.training_process = None
                self.reject()
        else:
            self.accept()


class DatasetPreparationDialog(QDialog):
    """Dialog for preparing training dataset"""

    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Prepare Training Dataset")
        self.setMinimumSize(550, 400)

        layout = QVBoxLayout(self)

        # Description
        desc_label = QLabel(
            "<h3>Dataset Preparation</h3>"
            "<p>Organize pottery images for training. This will analyze your database "
            "and help you prepare a dataset with proper groupings.</p>"
        )
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)

        # Grouping strategy
        group_box = QGroupBox("Grouping Strategy")
        group_layout = QVBoxLayout()

        self.radio_fabric = QRadioButton("Group by Fabric")
        self.radio_fabric.setToolTip("Group pottery by fabric type for material similarity")
        group_layout.addWidget(self.radio_fabric)

        self.radio_form = QRadioButton("Group by Form")
        self.radio_form.setChecked(True)
        self.radio_form.setToolTip("Group pottery by general form for shape similarity")
        group_layout.addWidget(self.radio_form)

        self.radio_specific_form = QRadioButton("Group by Specific Form")
        self.radio_specific_form.setToolTip("Group pottery by specific typological form for detailed shape similarity")
        group_layout.addWidget(self.radio_specific_form)

        self.radio_decoration_type = QRadioButton("Group by Decoration Type")
        self.radio_decoration_type.setToolTip("Group pottery by decoration type (geometric, figurative, etc.)")
        group_layout.addWidget(self.radio_decoration_type)

        self.radio_decoration_motif = QRadioButton("Group by Decoration Motif")
        self.radio_decoration_motif.setToolTip("Group pottery by decorative motif pattern")
        group_layout.addWidget(self.radio_decoration_motif)

        self.radio_decoration_combined = QRadioButton("Group by Decoration (Type+Motif+Position)")
        self.radio_decoration_combined.setToolTip("Group pottery by combined decoration attributes for detailed pattern similarity")
        group_layout.addWidget(self.radio_decoration_combined)

        self.radio_ware = QRadioButton("Group by Ware")
        self.radio_ware.setToolTip("Group pottery by ware type")
        group_layout.addWidget(self.radio_ware)

        self.radio_site = QRadioButton("Group by Site")
        self.radio_site.setToolTip("Group pottery by archaeological site")
        group_layout.addWidget(self.radio_site)

        group_box.setLayout(group_layout)
        layout.addWidget(group_box)

        # Output location
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("Output Folder:"))
        self.lineEdit_output = QLineEdit()
        self.lineEdit_output.setText(os.path.expanduser('~/pyarchinit/bin/training_data'))
        output_layout.addWidget(self.lineEdit_output)
        self.btn_browse = QPushButton("Browse...")
        self.btn_browse.clicked.connect(self.browse_output)
        output_layout.addWidget(self.btn_browse)
        layout.addLayout(output_layout)

        # Statistics
        self.label_stats = QLabel("Analyzing database...")
        self.label_stats.setStyleSheet("color: gray;")
        layout.addWidget(self.label_stats)

        # Progress
        self.progressBar = QProgressBar()
        self.progressBar.setVisible(False)
        layout.addWidget(self.progressBar)

        # Buttons
        btn_layout = QHBoxLayout()
        self.btn_analyze = QPushButton("Analyze Database")
        self.btn_analyze.clicked.connect(self.analyze_database)
        btn_layout.addWidget(self.btn_analyze)

        self.btn_prepare = QPushButton("Prepare Dataset")
        self.btn_prepare.clicked.connect(self.prepare_dataset)
        self.btn_prepare.setEnabled(False)
        btn_layout.addWidget(self.btn_prepare)

        self.btn_close = QPushButton("Close")
        self.btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(self.btn_close)

        layout.addLayout(btn_layout)

        # Run initial analysis
        QTimer.singleShot(100, self.analyze_database)

    def browse_output(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.lineEdit_output.setText(folder)

    def get_grouping_field(self):
        """Return the field to use for grouping based on radio selection"""
        if self.radio_fabric.isChecked():
            return 'fabric'
        elif self.radio_form.isChecked():
            return 'form'
        elif self.radio_specific_form.isChecked():
            return 'specific_form'
        elif self.radio_decoration_type.isChecked():
            return 'decoration_type'
        elif self.radio_decoration_motif.isChecked():
            return 'decoration_motif'
        elif self.radio_decoration_combined.isChecked():
            return 'decoration_combined'  # Special case handled in queries
        elif self.radio_ware.isChecked():
            return 'ware'
        elif self.radio_site.isChecked():
            return 'sito'
        return 'form'

    def analyze_database(self):
        """Analyze database to show statistics"""
        try:
            field = self.get_grouping_field()

            # Handle combined decoration field specially
            if field == 'decoration_combined':
                # Concatenate decoration_type, decoration_motif, decoration_position
                query = """
                    SELECT CONCAT_WS('_', p.decoration_type, p.decoration_motif, p.decoration_position) as combined,
                           COUNT(DISTINCT p.id_rep) as count
                    FROM pottery_table p
                    INNER JOIN media_to_entity_table mte ON p.id_rep = mte.id_entity AND mte.entity_type = 'CERAMICA'
                    INNER JOIN media_table m ON mte.id_media = m.id_media
                    WHERE (p.decoration_type IS NOT NULL AND p.decoration_type != ''
                           OR p.decoration_motif IS NOT NULL AND p.decoration_motif != ''
                           OR p.decoration_position IS NOT NULL AND p.decoration_position != '')
                          AND m.mediatype = 'image'
                    GROUP BY combined
                    ORDER BY count DESC
                """
            else:
                # Count pottery with images by grouping field
                # Use correct join through media_to_entity_table
                query = f"""
                    SELECT p.{field}, COUNT(DISTINCT p.id_rep) as count
                    FROM pottery_table p
                    INNER JOIN media_to_entity_table mte ON p.id_rep = mte.id_entity AND mte.entity_type = 'CERAMICA'
                    INNER JOIN media_table m ON mte.id_media = m.id_media
                    WHERE p.{field} IS NOT NULL AND p.{field} != '' AND m.mediatype = 'image'
                    GROUP BY p.{field}
                    ORDER BY count DESC
                """

            results = []
            total_count = 0

            # Create session from db_manager
            session = self.db_manager.Session()
            try:
                from sqlalchemy import text
                results = session.execute(text(query)).fetchall()

                # Count total pottery with images
                total_query = """
                    SELECT COUNT(DISTINCT p.id_rep)
                    FROM pottery_table p
                    INNER JOIN media_to_entity_table mte ON p.id_rep = mte.id_entity AND mte.entity_type = 'CERAMICA'
                    INNER JOIN media_table m ON mte.id_media = m.id_media
                    WHERE m.mediatype = 'image'
                """
                total_count = session.execute(text(total_query)).scalar() or 0
            except Exception as e:
                print(f"Query error: {e}")
            finally:
                session.close()

            # Build stats text
            stats_parts = [f"<b>Total pottery with images:</b> {total_count}"]

            if results:
                valid_groups = [(r[0], r[1]) for r in results if r[1] >= 5]
                stats_parts.append(f"<b>Groups with >= 5 images:</b> {len(valid_groups)}")
                if valid_groups[:5]:
                    top_groups = ", ".join([f"{g[0]}({g[1]})" for g in valid_groups[:5]])
                    stats_parts.append(f"<b>Top groups:</b> {top_groups}")
            else:
                stats_parts.append("(Could not analyze groupings)")

            self.label_stats.setText("<br>".join(stats_parts))
            self.label_stats.setStyleSheet("")
            self.btn_prepare.setEnabled(total_count > 10)

        except Exception as e:
            self.label_stats.setText(f"Error analyzing database: {str(e)}")
            self.label_stats.setStyleSheet("color: red;")

    def prepare_dataset(self):
        """Prepare the training dataset"""
        import shutil

        output_dir = self.lineEdit_output.text()
        if not output_dir:
            QMessageBox.warning(self, "Error", "Please specify output directory")
            return

        field = self.get_grouping_field()

        try:
            os.makedirs(output_dir, exist_ok=True)

            # Handle combined decoration field specially
            if field == 'decoration_combined':
                query = """
                    SELECT p.id_rep,
                           CONCAT_WS('_', p.decoration_type, p.decoration_motif, p.decoration_position) as combined,
                           mt.path_resize, m.filepath
                    FROM pottery_table p
                    INNER JOIN media_to_entity_table mte ON p.id_rep = mte.id_entity AND mte.entity_type = 'CERAMICA'
                    INNER JOIN media_table m ON mte.id_media = m.id_media
                    LEFT JOIN media_thumb_table mt ON m.id_media = mt.id_media
                    WHERE (p.decoration_type IS NOT NULL AND p.decoration_type != ''
                           OR p.decoration_motif IS NOT NULL AND p.decoration_motif != ''
                           OR p.decoration_position IS NOT NULL AND p.decoration_position != '')
                          AND m.mediatype = 'image'
                """
            else:
                # Get all pottery with images, grouped by field
                # Use correct join through media_to_entity_table and media_thumb_table
                query = f"""
                    SELECT p.id_rep, p.{field}, mt.path_resize, m.filepath
                    FROM pottery_table p
                    INNER JOIN media_to_entity_table mte ON p.id_rep = mte.id_entity AND mte.entity_type = 'CERAMICA'
                    INNER JOIN media_table m ON mte.id_media = m.id_media
                    LEFT JOIN media_thumb_table mt ON m.id_media = mt.id_media
                    WHERE p.{field} IS NOT NULL AND p.{field} != '' AND m.mediatype = 'image'
                """

            from sqlalchemy import text
            session = self.db_manager.Session()
            try:
                results = session.execute(text(query)).fetchall()
            finally:
                session.close()

            if not results:
                QMessageBox.warning(self, "No Data", "No pottery with images found in database")
                return

            # Group by field value
            groups = {}
            for row in results:
                id_rep, group_value, thumb_path_db, original_path = row
                # Prefer thumbnail, fallback to original
                image_path = thumb_path_db or original_path
                if not image_path:
                    continue
                if group_value not in groups:
                    groups[group_value] = []
                groups[group_value].append((id_rep, image_path))

            # Create directories and copy images
            self.progressBar.setVisible(True)
            self.progressBar.setMaximum(len(results))

            copied = 0
            skipped_groups = 0

            # Get base path for images
            from ..modules.db.pyarchinit_conn_strings import Connection
            conn = Connection()
            thumb_resize_path = conn.thumb_resize().get('thumb_resize', '')

            for group_name, images in groups.items():
                if len(images) < 5:
                    skipped_groups += 1
                    continue

                # Create safe directory name
                safe_name = "".join(c for c in str(group_name) if c.isalnum() or c in (' ', '-', '_')).strip()
                if not safe_name:
                    safe_name = f"group_{hash(group_name) % 10000}"

                group_dir = os.path.join(output_dir, safe_name)
                os.makedirs(group_dir, exist_ok=True)

                for id_rep, rel_path in images:
                    # Build full path - try thumb_resize first
                    full_path = os.path.join(thumb_resize_path, rel_path) if thumb_resize_path else rel_path
                    if not os.path.exists(full_path):
                        # Try as absolute path
                        full_path = rel_path
                    if os.path.exists(full_path):
                        dest_name = f"pottery_{id_rep}_{os.path.basename(rel_path)}"
                        dest_path = os.path.join(group_dir, dest_name)
                        try:
                            shutil.copy2(full_path, dest_path)
                            copied += 1
                        except Exception as copy_err:
                            print(f"Copy error: {copy_err}")

                    self.progressBar.setValue(copied)

            self.progressBar.setVisible(False)

            QMessageBox.information(self, "Dataset Prepared",
                f"Dataset prepared successfully!\n\n"
                f"Images copied: {copied}\n"
                f"Groups created: {len(groups) - skipped_groups}\n"
                f"Groups skipped (< 5 images): {skipped_groups}\n\n"
                f"Output: {output_dir}\n\n"
                "You can now use this folder for training.")

        except Exception as e:
            self.progressBar.setVisible(False)
            QMessageBox.warning(self, "Error", f"Failed to prepare dataset:\n{str(e)}")


class PotteryFilterDialog(QDialog):
    """Dialog for filtering Pottery records by ID Number and Year with checkboxes"""
    L = QgsSettings().value("locale/userLocale")[0:2]

    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.selected_ids = []
        self.selected_year = None
        self.pottery_records = []
        self.initUI()

    def initUI(self):
        if self.L == 'it':
            self.setWindowTitle("Filtra Record Ceramica per Anno e ID Number")
        elif self.L == 'de':
            self.setWindowTitle("Keramikdatensätze nach Jahr und ID-Nummer filtern")
        else:
            self.setWindowTitle("Filter Pottery Records by Year and ID Number")

        self.setMinimumSize(450, 550)
        layout = QVBoxLayout(self)

        # Year filter section
        year_layout = QHBoxLayout()
        year_label = QLabel(self)
        if self.L == 'it':
            year_label.setText("Anno:")
        elif self.L == 'de':
            year_label.setText("Jahr:")
        else:
            year_label.setText("Year:")
        year_layout.addWidget(year_label)

        self.comboBox_year = QComboBox(self)
        self.comboBox_year.setMinimumWidth(100)
        self.comboBox_year.currentIndexChanged.connect(self.on_year_changed)
        year_layout.addWidget(self.comboBox_year)
        year_layout.addStretch()
        layout.addLayout(year_layout)

        # Search bar for ID Number
        self.search_bar = QLineEdit(self)
        if self.L == 'it':
            self.search_bar.setPlaceholderText("Cerca ID Number...")
        elif self.L == 'de':
            self.search_bar.setPlaceholderText("ID-Nummer suchen...")
        else:
            self.search_bar.setPlaceholderText("Search ID Number...")
        self.search_bar.textChanged.connect(self.filter_list)
        layout.addWidget(self.search_bar)

        # Select All / Deselect All buttons
        btn_layout = QHBoxLayout()
        self.btn_select_all = QPushButton(self)
        self.btn_deselect_all = QPushButton(self)

        if self.L == 'it':
            self.btn_select_all.setText("Seleziona Tutti")
            self.btn_deselect_all.setText("Deseleziona Tutti")
        elif self.L == 'de':
            self.btn_select_all.setText("Alle auswählen")
            self.btn_deselect_all.setText("Alle abwählen")
        else:
            self.btn_select_all.setText("Select All")
            self.btn_deselect_all.setText("Deselect All")

        self.btn_select_all.clicked.connect(self.select_all)
        self.btn_deselect_all.clicked.connect(self.deselect_all)
        btn_layout.addWidget(self.btn_select_all)
        btn_layout.addWidget(self.btn_deselect_all)
        layout.addLayout(btn_layout)

        # List widget with checkboxes
        self.list_widget = QListWidget(self)
        layout.addWidget(self.list_widget)

        # Populate year combobox and list widget
        self.populate_years_and_ids()

        # Status label
        self.label_status = QLabel(self)
        self.update_status_label()
        layout.addWidget(self.label_status)

        # Filter button
        filter_button = QPushButton(self)
        if self.L == 'it':
            filter_button.setText("Applica Filtro")
        elif self.L == 'de':
            filter_button.setText("Filter anwenden")
        else:
            filter_button.setText("Apply Filter")
        filter_button.clicked.connect(self.apply_filter)
        layout.addWidget(filter_button)

        self.setLayout(layout)

    def natural_sort_key(self, text):
        """Natural sorting key that handles alphanumeric values correctly"""
        return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', str(text))]

    def populate_years_and_ids(self):
        """Fetch pottery records and populate year combobox and ID list"""
        try:
            # Get all pottery records
            self.pottery_records = self.db_manager.query_all('pottery_table')

            # Get unique years and sort them
            unique_years = sorted(
                set(record.anno for record in self.pottery_records if record.anno is not None),
                reverse=True  # Most recent first
            )

            # Populate year combobox
            self.comboBox_year.blockSignals(True)
            self.comboBox_year.clear()
            if self.L == 'it':
                self.comboBox_year.addItem("Tutti gli anni", None)
            elif self.L == 'de':
                self.comboBox_year.addItem("Alle Jahre", None)
            else:
                self.comboBox_year.addItem("All years", None)

            for year in unique_years:
                self.comboBox_year.addItem(str(year), year)
            self.comboBox_year.blockSignals(False)

            # Populate list with all IDs initially
            self.update_id_list()

        except Exception as e:
            print(f"Error populating pottery filter: {e}")

    def on_year_changed(self, index):
        """Handle year selection change"""
        self.selected_year = self.comboBox_year.currentData()
        self.search_bar.clear()
        self.update_id_list()

    def get_filtered_records(self):
        """Get records filtered by selected year"""
        if self.selected_year is None:
            return self.pottery_records
        return [r for r in self.pottery_records if r.anno == self.selected_year]

    def update_id_list(self):
        """Update the ID list based on current year filter"""
        filtered_records = self.get_filtered_records()

        # Get unique id_numbers from filtered records and sort them naturally
        unique_ids = sorted(
            set(record.id_number for record in filtered_records if record.id_number is not None),
            key=self.natural_sort_key
        )

        self.update_list_widget(unique_ids, filtered_records)

    def update_list_widget(self, id_numbers, records=None):
        """Update the list widget with the given ID numbers"""
        self.list_widget.clear()

        if records is None:
            records = self.get_filtered_records()

        # Create a dict to count occurrences per id_number
        id_count = {}
        for record in records:
            if record.id_number is not None:
                id_count[record.id_number] = id_count.get(record.id_number, 0) + 1

        for id_num in id_numbers:
            list_item = QListWidgetItem(self.list_widget)
            count = id_count.get(id_num, 0)
            checkbox = QCheckBox(f"ID: {id_num} ({count} record{'s' if count != 1 else ''})")
            checkbox.id_number = id_num
            checkbox.stateChanged.connect(self.update_status_label)
            self.list_widget.addItem(list_item)
            self.list_widget.setItemWidget(list_item, checkbox)

    def filter_list(self, text):
        """Filter the list based on search text"""
        filtered_records = self.get_filtered_records()

        if not text:
            # Show all if search is empty
            unique_ids = sorted(
                set(record.id_number for record in filtered_records if record.id_number is not None),
                key=self.natural_sort_key
            )
        else:
            # Filter by search text
            unique_ids = sorted(
                set(record.id_number for record in filtered_records
                    if record.id_number is not None and str(record.id_number).startswith(text)),
                key=self.natural_sort_key
            )
        self.update_list_widget(unique_ids, filtered_records)

    def select_all(self):
        """Select all visible checkboxes"""
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            checkbox = self.list_widget.itemWidget(item)
            if checkbox:
                checkbox.setChecked(True)

    def deselect_all(self):
        """Deselect all visible checkboxes"""
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            checkbox = self.list_widget.itemWidget(item)
            if checkbox:
                checkbox.setChecked(False)

    def update_status_label(self):
        """Update the status label with count of selected items"""
        count = self.get_selected_count()
        total = self.list_widget.count()
        year_text = ""
        if self.selected_year:
            year_text = f" ({self.selected_year})"

        if self.L == 'it':
            self.label_status.setText(f"Selezionati: {count} di {total}{year_text}")
        elif self.L == 'de':
            self.label_status.setText(f"Ausgewählt: {count} von {total}{year_text}")
        else:
            self.label_status.setText(f"Selected: {count} of {total}{year_text}")

    def get_selected_count(self):
        """Get the count of selected checkboxes"""
        count = 0
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            checkbox = self.list_widget.itemWidget(item)
            if checkbox and checkbox.isChecked():
                count += 1
        return count

    def apply_filter(self):
        """Apply the filter and close the dialog"""
        self.selected_ids.clear()

        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            checkbox = self.list_widget.itemWidget(item)
            if checkbox and checkbox.isChecked():
                self.selected_ids.append(checkbox.id_number)

        print(f"Selected ID Numbers: {self.selected_ids}, Year: {self.selected_year}")
        self.accept()

    def get_selected_ids(self):
        """Return the list of selected ID numbers"""
        return self.selected_ids

    def get_selected_year(self):
        """Return the selected year (None if 'All years')"""
        return self.selected_year

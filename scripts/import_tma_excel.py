#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_tma_excel.py — Parser/importer per le schede TMA di pyArchInit.

Riversa un Excel "TMA Elenco" (foglio ELENCO) nella tabella
``tma_materiali_archeologici`` del database PostgreSQL indicato in ``config.cfg``.

Regole (concordate con Marianna, vedi "Istruzioni per PARSING_TMA.docx"):

* Mapping colonne Excel -> campi TMA (vedi SRC_MAP / MAT_CATEGORIES).
* Colonna US (col. L) colorata:
    - GIALLO  (FFFF00): la US contiene piu' valori separati da VIRGOLA ->
      una scheda TMA per ogni US (token tenuti verbatim: "22a", "35.36", ...).
    - AZZURRO (00B0F0): piu' righe con la stessa (US, AREA, VANO/LOCUS) ->
      una sola scheda TMA, unificando TUTTI i campi (valori distinti uniti
      con " / ", identici saltati; materiali = unione).
    - nessun colore: la US e' presa cosi' com'e' (puo' essere vuota).
* Materiali: 12 colonne 1/0 -> ``ogtm`` = elenco delle categorie presenti.
* Campi fissi: sito = --site, ldct = "Magazzino", dtzg = "" (non in Excel).
* deso = Descrizione + Note.
* Import in APPEND (non tocca le righe esistenti).

Uso:
    python3 import_tma_excel.py --excel "<file.xlsx>"           # dry-run (default)
    python3 import_tma_excel.py --excel "<file.xlsx>" --apply   # scrive sul DB

Il dry-run genera "<excel>_TMA_preview.xlsx" + statistiche, senza toccare il DB.
"""

import argparse
import ast
import csv
import datetime as _dt
import os
import re
import sys
import uuid
from collections import OrderedDict

# Separatori multipli di US nelle celle GIALLE: virgola, punto, punto e virgola.
# (Excel salva i "35.36" come numeri float -> uno zero finale eventuale e' perso,
#  es. "2004.2010" diventa "2004.201": questi casi vanno verificati a mano.)
US_SPLIT_RE = re.compile(r"[.,;]+")

# Override manuali per celle dove Excel, salvando la US come numero, ha perso
# lo zero finale (es. "2004.2010" memorizzato come "2004.201"). La chiave e' il
# testo US letto dalla cella; il valore e' la lista corretta di US. Confermato
# con Marianna.
US_OVERRIDES = {
    "2004.201": ["2004", "2010"],
}

import openpyxl

# --- Mapping: header Excel (stripped) -> colonna TMA -------------------------
SRC_MAP = OrderedDict([
    ("Luogo",                    "localita"),
    ("Area/Quartiere",           "area"),
    ("SETTORE",                  "settore"),
    ("SAGGIO",                   "saggio"),
    ("VANO / LOCUS",             "vano_locus"),
    ("Data",                     "dscd"),
    ("Inventariati",             "inventario"),
    ("Cassa (nuova numerazione)", "cassetta"),
    ("Cassa",                    "vecchia_collocazione"),
    ("Magazzino",                "ldcn"),
    ("Campagna",                 "scan"),
])
US_HEADER = "US"
DESC_HEADER = "Descrizione"
NOTE_HEADER = "Note"

# Categorie materiali (header Excel), nell'ordine; 1 = presente.
MAT_CATEGORIES = [
    "CERAMICA", "LITICA", "VASI IN PIETRA", "METALLO", "STATUINE", "FUSERUOLE",
    "RIVESTIMENTI PARIETALI", "OSSA UMANE", "OSSA ANIMALI", "CARBONI", "TERRA",
    "ALTRO",
]

YELLOW = "FFFFFF00"
BLUE = "FF00B0F0"

# Colonne TMA scritte dallo script (in ordine), escluso id/system auto.
TMA_FIELDS = [
    "sito", "area", "localita", "settore", "inventario", "ogtm", "ldct",
    "ldcn", "vecchia_collocazione", "cassetta", "scan", "saggio", "vano_locus",
    "dscd", "dscu", "dtzg", "deso",
]
# Campi testuali unificabili in un gruppo AZZURRO (tutti tranne i fissi e la US).
MERGE_FIELDS = [
    "area", "localita", "settore", "inventario", "ldcn", "vecchia_collocazione",
    "cassetta", "scan", "saggio", "vano_locus", "dscd", "deso",
]


def to_text(v):
    """Converte un valore di cella in stringa pulita preservando i numeri/US."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "1" if v else ""
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    if isinstance(v, int):
        return str(v)
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def is_present(v):
    """True se la cella materiale indica presenza (1), False per 0/vuoto."""
    s = to_text(v)
    return s not in ("", "0")


def us_fill(cell):
    """Ritorna 'yellow' / 'blue' / None in base al riempimento della cella US."""
    f = cell.fill
    if not f or f.patternType != "solid":
        return None
    rgb = getattr(f.fgColor, "rgb", None)
    if rgb == YELLOW:
        return "yellow"
    if rgb == BLUE:
        return "blue"
    return None


def merge_values(values):
    """Unisce valori distinti non vuoti preservando l'ordine, con ' / '."""
    seen = OrderedDict()
    for v in values:
        v = (v or "").strip()
        if v and v not in seen:
            seen[v] = True
    return " / ".join(seen.keys())


def cat_list(materials):
    """Lista ordinata-unica di categorie in sentence-case (solo la 1a lettera
    maiuscola), es. "CERAMICA" -> "Ceramica". Usata sia per ``ogtm`` (testo)
    sia per le righe di ``tma_materiali_ripetibili`` (campo ``macc``)."""
    seen = OrderedDict()
    for m in materials:
        m = (m or "").strip()
        if m:
            k = m.capitalize()
            seen.setdefault(k, True)
    return list(seen.keys())


def render_ogtm(materials):
    """Testo riepilogo materiali: categorie sentence-case separate da ', '."""
    return ", ".join(cat_list(materials))


def build_column_index(ws):
    """header (stripped) -> column letter, dalla riga 1."""
    idx = {}
    for c in ws[1]:
        if c.value is not None:
            idx[str(c.value).strip()] = c.column_letter
    return idx


def parse_excel(path, sheet, site):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Foglio '{sheet}' non trovato. Fogli: {wb.sheetnames}")
    ws = wb[sheet]
    col = build_column_index(ws)

    missing = [h for h in list(SRC_MAP) + [US_HEADER, DESC_HEADER, NOTE_HEADER]
               + MAT_CATEGORIES if h not in col]
    if missing:
        print(f"[WARN] header mancanti nell'Excel (saltati): {missing}")

    def cellval(row, header):
        letter = col.get(header)
        return ws[f"{letter}{row}"].value if letter else None

    records = []          # record non-colorati e gialli (gia' espansi)
    blue_groups = OrderedDict()   # (us, area, vano) -> lista di record per-riga
    stats = dict(input_rows=0, yellow=0, blue=0, plain=0, empty_us=0,
                 split_extra=0, blue_groups=0)

    us_letter = col.get(US_HEADER)

    for row in range(2, ws.max_row + 1):
        # Salta righe completamente vuote
        if all(cellval(row, h) in (None, "") for h in list(SRC_MAP)
               + [US_HEADER, DESC_HEADER, NOTE_HEADER] + MAT_CATEGORIES):
            continue
        stats["input_rows"] += 1

        base = {f: "" for f in TMA_FIELDS}
        base["sito"] = site
        base["ldct"] = "Magazzino"
        base["dtzg"] = ""
        for header, field in SRC_MAP.items():
            base[field] = to_text(cellval(row, header))
        # "Inventariati": in `inventario` (RA) tieni SOLO i veri numeri
        # d'inventario (contengono cifre). Il testo descrittivo — es.
        # "Vasi inventariati", "Scoria", "Osso lavorato" — NON e' un numero RA:
        # va in `deso` e `inventario` resta vuoto (evita anche i conflitti sul
        # vincolo UNIQUE parziale (sito, area, inventario, dscu)).
        deso_parts = [to_text(cellval(row, DESC_HEADER)),
                      to_text(cellval(row, NOTE_HEADER))]
        inv = base.get("inventario", "")
        if inv and not any(ch.isdigit() for ch in inv):
            deso_parts.append(inv)
            base["inventario"] = ""
        base["deso"] = merge_values(deso_parts)
        materials = [m for m in MAT_CATEGORIES if is_present(cellval(row, m))]

        us_raw = to_text(cellval(row, US_HEADER))
        kind = us_fill(ws[f"{us_letter}{row}"]) if us_letter else None

        if not us_raw:
            stats["empty_us"] += 1

        if kind == "yellow":
            stats["yellow"] += 1
            if us_raw in US_OVERRIDES:
                tokens = list(US_OVERRIDES[us_raw])
            else:
                tokens = [t.strip() for t in US_SPLIT_RE.split(us_raw) if t.strip()]
            if not tokens:
                tokens = [""]
            stats["split_extra"] += max(0, len(tokens) - 1)
            for tok in tokens:
                rec = dict(base)
                rec["dscu"] = tok
                rec["_materials"] = cat_list(materials)
                rec["ogtm"] = ", ".join(rec["_materials"])
                rec["_origine"] = f"giallo/split US riga {row}"
                records.append(rec)

        elif kind == "blue":
            stats["blue"] += 1
            key = (us_raw, base["area"], base["vano_locus"])
            rec = dict(base)
            rec["dscu"] = us_raw
            rec["_materials"] = materials
            rec["_row"] = row
            blue_groups.setdefault(key, []).append(rec)

        else:
            stats["plain"] += 1
            rec = dict(base)
            rec["dscu"] = us_raw
            rec["_materials"] = cat_list(materials)
            rec["ogtm"] = ", ".join(rec["_materials"])
            rec["_origine"] = f"riga {row}"
            records.append(rec)

    # Unifica i gruppi azzurri
    for (us, area, vano), group in blue_groups.items():
        stats["blue_groups"] += 1
        merged = {f: "" for f in TMA_FIELDS}
        merged["sito"] = site
        merged["ldct"] = "Magazzino"
        merged["dtzg"] = ""
        merged["dscu"] = us
        for f in MERGE_FIELDS:
            merged[f] = merge_values([g.get(f, "") for g in group])
        all_mats = []
        for g in group:
            all_mats.extend(g.get("_materials", []))
        merged["_materials"] = cat_list(all_mats)
        merged["ogtm"] = ", ".join(merged["_materials"])
        rows = ", ".join(str(g["_row"]) for g in group)
        merged["_origine"] = f"azzurro/unifica {len(group)} righe ({rows})"
        records.append(merged)

    return records, stats


# Etichette leggibili (nome scheda TMA) + colonna DB, per l'anteprima.
LABELS = {
    "dscu": "US (dscu)",
    "sito": "Sito (sito)",
    "area": "Area (area)",
    "localita": "Località (localita)",
    "settore": "Settore (settore)",
    "saggio": "Saggio (saggio)",
    "vano_locus": "Vano/Locus (vano_locus)",
    "dscd": "Data scavo (dscd)",
    "scan": "Denominazione scavo (scan)",
    "cassetta": "Cassetta (cassetta)",
    "vecchia_collocazione": "Vecchia collocazione (vecchia_collocazione)",
    "ldcn": "Denominazione collocazione (ldcn)",
    "ldct": "Tipologia collocazione (ldct)",
    "inventario": "Inventario RA (inventario)",
    "ogtm": "Materiale (ogtm)",
    "dtzg": "Fascia cronologica (dtzg)",
    "deso": "Indicazioni oggetti = Descrizione+Note (deso)",
    "_origine": "Origine (riga/regola Excel)",
}
# Ordine leggibile per l'anteprima: identificazione US in testa.
PREVIEW_ORDER = [
    "dscu", "sito", "area", "localita", "settore", "saggio", "vano_locus",
    "dscd", "scan", "cassetta", "vecchia_collocazione", "ldcn", "ldct",
    "inventario", "ogtm", "dtzg", "deso", "_origine",
]


def write_preview(records, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TMA_preview"
    ws.append([LABELS.get(h, h) for h in PREVIEW_ORDER])
    for rec in records:
        ws.append([rec.get(h, "") for h in PREVIEW_ORDER])
    wb.save(out_path)


def load_config(path):
    return ast.literal_eval(open(path, "r", encoding="utf-8").read())


def apply_to_db(records, cfg):
    import psycopg2
    conn = psycopg2.connect(host=cfg["HOST"], port=cfg["PORT"],
                            dbname=cfg["DATABASE"], user=cfg["USER"],
                            password=cfg["PASSWORD"], connect_timeout=15)
    conn.autocommit = False
    cur = conn.cursor()
    table = "tma_materiali_archeologici"

    # --- Backup della tabella su CSV (timestamp) ---
    ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = os.path.join(os.path.dirname(os.path.abspath(cfg["__path__"])),
                          f"backup_{table}_{ts}.csv")
    cur.execute(f'SELECT * FROM "{table}" ORDER BY id')
    bcols = [d[0] for d in cur.description]
    with open(backup, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(bcols)
        w.writerows(cur.fetchall())
    print(f"[backup] {table} -> {backup}")

    now = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # --- gestione id ---
    # Se la colonna id ha una sequenza (serial), risincronizziamo la sequenza
    # al max(id) corrente e lasciamo che sia il DEFAULT ad assegnare gli id:
    # cosi' non desincronizziamo la sequenza per i futuri insert da GUI. Se non
    # c'e' sequenza, ripieghiamo sull'assegnazione esplicita max+1.
    cur.execute("SELECT pg_get_serial_sequence(%s, 'id')", (table,))
    seq = cur.fetchone()[0]
    use_explicit_id = seq is None
    next_id = None
    if seq:
        cur.execute(f'SELECT setval(%s, (SELECT COALESCE(MAX(id), 1) FROM "{table}"))', (seq,))
        print(f"[apply] sequenza {seq} risincronizzata a max(id).")
        insert_cols = TMA_FIELDS + ["created_at", "created_by", "entity_uuid"]
    else:
        cur.execute(f'SELECT COALESCE(MAX(id), 0) + 1 FROM "{table}"')
        next_id = cur.fetchone()[0]
        insert_cols = ["id"] + TMA_FIELDS + ["created_at", "created_by", "entity_uuid"]

    placeholders = ", ".join(["%s"] * len(insert_cols))
    quoted = ", ".join(f'"{c}"' for c in insert_cols)
    # Vincolo UNIQUE parziale idx_tma_unique_record (sito, area, inventario,
    # dscu) WHERE inventario non vuoto: salta in silenzio le poche duplicazioni
    # genuine (stesso numero d'inventario, stessa area/US) invece di abortire.
    # RETURNING id serve per collegare le righe materiali ripetibili.
    sql = (f'INSERT INTO "{table}" ({quoted}) VALUES ({placeholders}) '
           f"ON CONFLICT (sito, area, inventario, dscu) "
           f"WHERE inventario IS NOT NULL AND inventario <> '' DO NOTHING "
           f"RETURNING id")

    # Materiali ripetibili: una riga per categoria, SOLO la categoria
    # (macc = madi = categoria sentence-case), collegata via id_tma. La form TMA
    # legge la griglia "Materiali" da qui.
    rip_table = "tma_materiali_ripetibili"
    cur.execute("SELECT pg_get_serial_sequence(%s, 'id')", (rip_table,))
    rip_seq = cur.fetchone()[0]
    if rip_seq:
        cur.execute(f'SELECT setval(%s, (SELECT COALESCE(MAX(id), 1) FROM "{rip_table}"))',
                    (rip_seq,))
    rip_sql = (f'INSERT INTO "{rip_table}" '
               f'(id_tma, macc, madi, created_at, created_by, entity_uuid) '
               f'VALUES (%s, %s, %s, %s, %s, %s)')

    inserted = 0
    skipped = 0
    rip_rows = 0
    for rec in records:
        base = [rec.get(f, "") for f in TMA_FIELDS] + [
            now, "tma_excel_import", str(uuid.uuid4())]
        vals = ([next_id] + base) if use_explicit_id else base
        if use_explicit_id:
            next_id += 1
        cur.execute(sql, vals)
        row = cur.fetchone()
        if row is None:          # conflitto -> saltata, niente ripetibili
            skipped += 1
            continue
        inserted += 1
        pid = row[0]
        for cat in rec.get("_materials", []):
            cur.execute(rip_sql, (pid, cat, cat, now, "tma_excel_import",
                                  str(uuid.uuid4())))
            rip_rows += 1
    conn.commit()
    cur.execute(f'SELECT COUNT(*), MAX(id) FROM "{table}"')
    tot, mx = cur.fetchone()
    cur.close()
    conn.close()
    print(f"[apply] inserite {inserted} righe in '{table}' "
          f"(saltate {skipped} dup su vincolo UNIQUE) + {rip_rows} righe in "
          f"'{rip_table}'. Totale {table}: {tot} (max id {mx}).")


def main():
    ap = argparse.ArgumentParser(description="Parser/importer schede TMA da Excel")
    ap.add_argument("--excel", required=True, help="file .xlsx sorgente")
    ap.add_argument("--sheet", default="ELENCO", help="nome foglio (default ELENCO)")
    ap.add_argument("--config", default=os.path.expanduser(
        "~/pyarchinit/pyarchinit_DB_folder/config.cfg"),
        help="percorso config.cfg")
    ap.add_argument("--site", default="Festòs_2025", help="valore campo 'sito'")
    ap.add_argument("--apply", action="store_true",
                    help="scrive sul DB (default: dry-run, nessuna scrittura)")
    args = ap.parse_args()

    records, stats = parse_excel(args.excel, args.sheet, args.site)

    print("\n=== STATISTICHE ===")
    print(f"  righe input (non vuote) : {stats['input_rows']}")
    print(f"  righe non colorate      : {stats['plain']}")
    print(f"  US vuote                : {stats['empty_us']}")
    print(f"  righe GIALLE (split)    : {stats['yellow']}  (+{stats['split_extra']} schede extra)")
    print(f"  righe AZZURRE (unifica) : {stats['blue']}  -> {stats['blue_groups']} schede unificate")
    print(f"  >>> SCHEDE TMA totali   : {len(records)}")

    if not args.apply:
        out = os.path.splitext(args.excel)[0] + "_TMA_preview.xlsx"
        write_preview(records, out)
        print(f"\n[dry-run] anteprima scritta in:\n  {out}")
        print("[dry-run] nessuna scrittura sul DB. Usa --apply per importare.")
    else:
        cfg = load_config(args.config)
        cfg["__path__"] = args.config
        apply_to_db(records, cfg)


if __name__ == "__main__":
    main()
